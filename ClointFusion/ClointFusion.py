# Project Name: ClointFusion
# Project Description: A Python based RPA Automation Framework for Desktop GUI, Citrix, Web and basic Excel operations.
# Code Authors: Mayur Patil, Murali M V P
# Code Contributors: Avinash, Fharook

# Code Structure
# 1. All imports
# 2. All global variables
# 3. All function definitions
# 4. All test cases
# 5. All default services

# 1. All imports

# Python Inbuilt Libraries
import subprocess, os, sys, platform, sqlite3, time, webbrowser, traceback, shutil, socket
from dateutil import parser
import urllib.request
from functools import lru_cache
import threading
from threading import Timer
import re
import json
import pyinspect as pi
from rich import pretty
from pathlib import Path
from datetime import timedelta
import logging
import tempfile
import warnings
from pywebio.output import put_text
import datetime
import PySimpleGUI as sg
import openpyxl as op
import pandas as pd

try:
    import pyautogui as pg
except:
    from pywebio.output import popup, put_html
    
from openpyxl import Workbook
from openpyxl import load_workbook
import clipboard
import helium as browser
from PIL import Image
import requests
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.common.exceptions import InvalidArgumentException, TimeoutException, WebDriverException, NoSuchWindowException, SessionNotCreatedException, ElementNotInteractableException, UnexpectedAlertPresentException, InvalidSelectorException
from tabloo import show
from colored import fg, attr
import click
from rich.console import Console
import random
import speech_recognition as sr
import pyttsx3

temp_current_working_dir = tempfile.mkdtemp(prefix="cloint_",suffix="_fusion")
temp_current_working_dir = Path(temp_current_working_dir)

windows_os = "windows"
linux_os = "linux"
mac_os = "darwin"
os_name = str(platform.system()).lower()

python_exe_path = os.path.join(os.path.dirname(sys.executable), "python.exe")
pythonw_exe_path = os.path.join(os.path.dirname(sys.executable), "pythonw.exe")
python_version = str(sys.version_info.major)

if os.getlogin() in python_exe_path:
    python_exe_path = python_exe_path.replace(os.getlogin(), f'"{os.getlogin()}"')

if os.getlogin() in pythonw_exe_path:
    pythonw_exe_path = pythonw_exe_path.replace(os.getlogin(), f'"{os.getlogin()}"')

if os_name == windows_os:
    clointfusion_directory = r"C:\Users\{}\ClointFusion".format(str(os.getlogin()))
elif os_name == linux_os:
    clointfusion_directory = r"/home/{}/ClointFusion".format(str(os.getlogin()))
elif os_name == mac_os:
    clointfusion_directory = r"/Users/{}/ClointFusion".format(str(os.getlogin()))
else:
    clointfusion_directory = temp_current_working_dir

current_working_dir = os.path.dirname(os.path.realpath(__file__)) #get cwd
config_folder_path = Path(os.path.join(clointfusion_directory, "Config_Files"))
img_folder_path =  Path(os.path.join(clointfusion_directory, "Images"))
cf_splash_png_path = Path(os.path.join(clointfusion_directory,"Logo_Icons","Splash.PNG"))
cf_icon_cdt_file_path = os.path.join(clointfusion_directory,"Logo_Icons","Cloint-ICON-CDT.ico")
log_path = Path(os.path.join(clointfusion_directory, "Logs"))
batch_file_path = Path(os.path.join(clointfusion_directory, "Batch_File"))    
output_folder_path = Path(os.path.join(clointfusion_directory, "Output")) 
error_screen_shots_path = Path(os.path.join(clointfusion_directory, "Error_Screenshots"))
status_log_excel_filepath = Path(os.path.join(clointfusion_directory,"StatusLogExcel"))
cf_icon_file_path = os.path.join(clointfusion_directory,"Logo_Icons","Cloint-ICON.ico")
cf_logo_file_path = os.path.join(clointfusion_directory,"Logo_Icons","Cloint-LOGO.PNG")
engine = ""

def _get_site_packages_path():
    """
    Returns Site-Packages Path
    """
    try:
        import site
        if os_name == windows_os:
            site_packages_path = next(p for p in site.getsitepackages() if 'site-packages' in p)
        else:
            site_packages_path = site.getsitepackages()[0]
    except:
        site_packages_path = site.getsitepackages()[0]
    return str(site_packages_path)

site_packages_path = _get_site_packages_path()

#Bol Related
if os_name == windows_os:   
    engine = pyttsx3.init('sapi5')
    voices = engine.getProperty('voices')
    voice = random.choice(voices)# Randomly decide male/female voice
    engine.setProperty('voice', voice.id)
    r = sr.Recognizer()
    energy_threshold = [3000]
    
    from unicodedata import name
    import pygetwindow as gw
    try:
        import win32gui
    except:
        cmd = python_exe_path + site_packages_path  + "/Scripts/pywin32_postinstall.py -install"
        os.system(cmd)
        import win32gui
elif os_name == linux_os:
    engine = pyttsx3.init('espeak')
    voices = engine.getProperty('voices')
    voice_int = random.randint(11, 17)# Randomly decide male/female voice
    engine.setProperty('voice', voices[voice_int].id)
    r = sr.Recognizer()
    energy_threshold = [3000]
else:
    engine = pyttsx3.init()
    voices = engine.getProperty('voices')
    voice = random.choice(voices)# Randomly decide male/female voice
    engine.setProperty('voice', voice.id)

pi.install_traceback(hide_locals=True,relevant_only=True,enable_prompt=True)
pretty.install()

console = Console()
sg.theme('Dark') # for PySimpleGUI FRONT END

try:
    from ClointFusion import selft
except Exception as e:
    import selft
    # print("Please re-install ClointFusion.")
    # sys.exit()

# 2. All global variables

log_path = ""
log_file_path = ""

bot_name = "My_BOT"
user_name, c_version, s_version, user_email = selft.get_details()

browser_driver = ""

ss_path_b = Path(os.path.join(temp_current_working_dir,"my_screen_shot_before.png")) #before search
ss_path_a = Path(os.path.join(temp_current_working_dir,"my_screen_shot_after.png")) #after search

enable_semi_automatic_mode = False # Default is to GUI Mode
Browser_Service_Started = False
ai_screenshot = ""
ai_processes = []
helium_service_launched=False

contribution_messages = ["We appreciate your contribution to a better tomorrow.", "Today's def (Python) creates tomorrow's software. Please contribute.", "A bug fix a day, keeps the Exceptions away. I'd be honored to have you as a contributor.", "Contributors are the true givers of the modern age. Please make a contribution.", "Money can merely buy a server; contributors are the true wealth creators.", "Contributors are the backbone of Open Source Software. Please keep contributing."]


# 3. All function definitions

# ---------  Methods ---------

def show_emoji(strInput=""):
    """
    Function which prints Emojis

    Usage: 
    print(show_emoji('thumbsup'))
    print("OK",show_emoji('thumbsup'))
    Default: thumbsup
    """
    import emoji
    
    if not strInput:
        return(emoji.emojize(":{}:".format(str('thumbsup').lower()),use_aliases=True,variant="emoji_type"))
    else:
        return(emoji.emojize(":{}:".format(str(strInput).lower()),use_aliases=True,variant="emoji_type"))

def print_with_magic_color(strMsg:str="",magic:bool=False)->None:
    """
    Function used for printing strings in color
    Args : Message (str): string which needs to be printed in color
           magic : (bool) :If each letter/character in the string is to be printed with different color
    Returns : None

    """
    import random
    
    try:
        if magic == False:
            fg_random = random.randint(2,255)
            
            while fg_random in [8,*range(15,28),22,23,*range(51,68),77,*range(87,99),114,149,*range(231,250)]:
                fg_random = random.randint(2,255)            

            print ('%s %s  %s' % (fg(fg_random), strMsg, attr(1)))
        else:
            for ch in strMsg:
                try:
                    fg_random = random.randint(2,255)
                    while fg_random in [8,*range(15,28),22,23,*range(51,68),77,*range(87,99),114,149,*range(231,250)]:
                        fg_random = random.randint(2,255)
                    print ('%s%s%s' % (fg(fg_random), ch,attr(1)),sep='',end='')
                except:
                    print ('%s' % (fg(1), attr('reset')),ch,sep='',end='')
        
        reset = attr('reset')    
        print (reset)
    except UnicodeEncodeError:
        pass
    except TypeError:
        pass
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in print_with_magic_color="+str(ex))

def read_semi_automatic_log(key):
    """
    Function to read a value from semi_automatic_log for a given key
    """
    try:
        if config_folder_path:
            bot_config_path = os.path.join(config_folder_path,bot_name + ".xlsx")
            bot_config_path = Path(bot_config_path)
        else:
            bot_config_path = os.path.join(clointfusion_directory,"Config_Files","First_Run.xlsx")
            bot_config_path = Path(bot_config_path)
            
            if not os.path.exists(bot_config_path):
                df = pd.DataFrame({'SNO': [],'KEY': [], 'VALUE':[]})
                append_df_to_excel(bot_config_path, df, index=False, startrow=0)

        df = pd.read_excel(bot_config_path,engine='openpyxl')
        
        value = df[df['KEY'] == key]['VALUE'].to_list()
        value = str(value[0])
        return value

    except:
        return None

def update_semi_automatic_log(key, value):
    """
    Update semi automatic excel log 
    """
    try:
        if config_folder_path:
            bot_config_path = os.path.join(config_folder_path, bot_name + ".xlsx")
            
        else:
            bot_config_path = os.path.join(clointfusion_directory,"Config_Files","First_Run.xlsx")
        
        bot_config_path = Path(bot_config_path)
        
        if _excel_if_value_exists(bot_config_path,usecols=['KEY'],value=key):
            df = pd.read_excel(bot_config_path,engine='openpyxl')
            row_index = df.index[df['KEY'] == key].tolist()[0]
            
            df.loc[row_index,'VALUE'] = value
            df.to_excel(bot_config_path,index=False)
        else:
            # print(bot_config_path)
            reader = pd.read_excel(bot_config_path,engine='openpyxl')
            df = pd.DataFrame({'SNO': [len(reader)+1], 'KEY': [key], 'VALUE':[value]})
            append_df_to_excel(bot_config_path, df, index=False,startrow=None,header=None)
            
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in update_semi_automatic_log="+str(ex))

def OFF_semi_automatic_mode():
    """
    This function sets semi_automatic_mode as False => OFF
    """
    global enable_semi_automatic_mode
    semi_automatic_config_file_path = os.path.join(config_folder_path,"Semi_Automatic_Mode.txt")
    semi_automatic_config_file_path = Path(semi_automatic_config_file_path)

    try:    
        with open(semi_automatic_config_file_path, 'w') as f:
            f.write('False')
        enable_semi_automatic_mode = False
        print("Semi Automatic Mode is DISABLED "+ show_emoji())
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in OFF_semi_automatic_mode="+str(ex))        

def ON_semi_automatic_mode():
    """
    This function sets semi_automatic_mode as True => ON
    """
    global enable_semi_automatic_mode
    semi_automatic_config_file_path = os.path.join(config_folder_path,"Semi_Automatic_Mode.txt")
    semi_automatic_config_file_path = Path(semi_automatic_config_file_path)

    try:    
        with open(semi_automatic_config_file_path, 'w') as f:
            f.write('True')    
            
        enable_semi_automatic_mode = True
        print("Semi Automatic Mode is ENABLED "+ show_emoji())
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in ON_semi_automatic_mode="+str(ex))

def text_to_speech(audio, show=True, rate=170):
    """
    Text to Speech using Google's Generic API
    Rate is the speed of speech. Default is 150
    Actual default : 200
    """
    global engine
    if type(audio) is list:
        if show:
            print(' '.join(audio))
    else:
        if show:
            print(str(audio))
    if os_name == linux_os:
        rate -= 10
    engine.setProperty('rate', rate)
    engine.say(audio)   
    engine.runAndWait()

# ---------  Methods Ends ---------


# ---------  Private Functions ---------
# Function in use, Dont Delete
def _perform_self_test(tour=False):
    try:
        data = cursr.execute("SELECT updating from CF_IMP_VALUES")

        for row in data:
            updating =  row[0]
            if updating == "False":
                _welcome_to_clointfusion()
                clointfusion_self_test(tour)
            else:
                sys.exit()
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in Self Test="+str(ex))
        _rerun_clointfusion_first_run(ex)

def _welcome_to_clointfusion():
    global user_name
    """
    Internal Function to display welcome message & push a notification to ClointFusion Slack
    """
    from pyfiglet import Figlet
    version = "(Version: 1.1.4)"

    hour = datetime.datetime.now().hour
    
    if hour >= 0 and hour < 12:
        greeting = "Good Morning"
    elif hour >= 12 and hour < 16:
        greeting = "Good Afternoon"
    else:
        greeting = "Good Evening"

    messages_list = ['Where would I be without a friend like you?', 'I appreciate what you did.', 'Thank you for thinking of me.', 'Thank you for your time today.', 'I am so thankful for what you did here', 'I really appreciate your help. Thank you.', "You know, if you're reading this, you're in the top 1% of smart people.", 'We know the world is full of choices. Yet you picked us, Thank you very much.', 'Thank you. We hope your experience was excellent and we can’t wait to see you again soon.', 'We hope you are happy with our tool, if not we are just an e-mail away at clointfusion@cloint.com. We will be pleased to hear from you.', 'ClointFusion would like to thank excellent users like you for your support. We couldn’t do it without you!', 'Thank you for your business, your trust, and your confidence. It is our pleasure to work with you.', 'We take pride in your business with us. Thank you!', 'It has been our pleasure to serve you, and we hope we see you again soon.', 'We value your trust and confidence in us and sincerely appreciate you!', 'Your satisfaction is our greatest concern!', 'Your confidence in us is greatly appreciated!', 'We are excited to serve you first!', 'Thank you for keeping us informed about how best to serve your needs. Together, we can make this history.', 'Our brand innovation wouldn’t have been possible if you didn’t give us feedback about our services.', 'Thank you so much for playing a pivotal role in our growth. We’ll make sure we continue to put your needs first as our company expands and improves.', 'We are exceedingly pleased to find people we can always count on. Thank you for being one of our loyal and trusted clients.', ]
    message = random.choice(messages_list)
    
    welcome_msg = f"\n{greeting} {str(user_name).title()} !  Welcome to ClointFusion, Made in India with " + show_emoji("red_heart") + f". {version}"
    print_with_magic_color(welcome_msg,magic=True)
    print()
    print_with_magic_color(message,magic=False)
    f = Figlet(font='small', width=150)
    console.print(f.renderText("ClointFusion Community Edition"))

def _create_status_log_file(xtLogFilePath):
    """
    Internal Function to create Status Log File
    """
    try:
        if not os.path.exists(xtLogFilePath):
            df = pd.DataFrame({'Timestamp': [], 'Status':[]})
            writer = pd.ExcelWriter(xtLogFilePath) # pylint: disable=abstract-class-instantiated
            df.to_excel(writer, sheet_name='Sheet1', index=False)
            writer.save()
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in _create_status_log_file = " +str(ex))

def _init_log_file():
    """
    Generates the log and saves it to the file in the given base directory. Internal function
    """
    global status_log_excel_filepath
    
    try:
        excelFileName = "StatusLog.xlsx"

        if not os.path.exists(status_log_excel_filepath):
            os.makedirs(status_log_excel_filepath)

        status_log_excel_filepath = os.path.join(status_log_excel_filepath,excelFileName)
        status_log_excel_filepath = Path(status_log_excel_filepath)
        
        _create_status_log_file(status_log_excel_filepath)   

    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in _init_log_file="+str(ex))

def _folder_write_text_file(txt_file_path="",contents=""):
    """
    Writes given contents to a text file
    """
    try:
        
        f = open(txt_file_path,'w',encoding="utf-8")
        f.write(str(contents))
        f.close()
        
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in folder_write_text_file="+str(ex))

def _ask_user_semi_automatic_mode():
    """
    Ask user to 'Enable Semi Automatic Mode'
    """
    try:
        # global enable_semi_automatic_mode
        # values = []
        
        # file_path = os.path.join(config_folder_path, 'Dont_Ask_Again.txt')
        # file_path = Path(file_path)
        # stored_do_not_ask_user_preference = _folder_read_text_file(file_path)

        # file_path = os.path.join(config_folder_path, 'Semi_Automatic_Mode.txt')
        # file_path = Path(file_path)
        # enable_semi_automatic_mode = _folder_read_text_file(file_path)
        
        bot_config_path = os.path.join(config_folder_path,bot_name + ".xlsx")
        bot_config_path = Path(bot_config_path)
        # pg.alert(bot_config_path)

        # if stored_do_not_ask_user_preference is None or str(stored_do_not_ask_user_preference).lower() == 'false':

        #     layout = [[sg.Text('Do you want me to store GUI responses & use them next time when you run this BOT ?',text_color='orange',font='Courier 13')],
        #             [sg.Submit('Yes',bind_return_key=True,button_color=('white','green'),font='Courier 14', focus=True), sg.CloseButton('No', button_color=('white','firebrick'),font='Courier 14')],
        #             [sg.Checkbox('Do not ask me again', key='-DONT_ASK_AGAIN-',default=True, text_color='yellow',enable_events=True)],
        #             [sg.Text("To see this message again, goto 'Config_Files' folder of your BOT and change 'Dont_Ask_Again.txt' to False. \n Please find path here: {}".format(Path(os.path.join(config_folder_path, 'Dont_Ask_Again.txt'))),key='-DND-',visible=False,font='Courier 8')]]

        #     window = sg.Window('ClointFusion - Enable Semi Automatic Mode ?',layout,return_keyboard_events=True,use_default_focus=False,disable_close=False,element_justification='c',keep_on_top=True,finalize=True,icon=cf_icon_file_path)
            
        #     while True:
        #         event, values = window.read()
        #         if event == '-DONT_ASK_AGAIN-':
        #             stored_do_not_ask_user_preference = values['-DONT_ASK_AGAIN-']
        #             file_path = os.path.join(config_folder_path, 'Dont_Ask_Again.txt')
        #             file_path = Path(file_path)
        #             _folder_write_text_file(file_path,str(stored_do_not_ask_user_preference))

        #             if values['-DONT_ASK_AGAIN-']:
        #                 window.Element('-DND-').Update(visible=True)
        #             else:
        #                 window.Element('-DND-').Update(visible=False)

        #         file_path = os.path.join(config_folder_path, 'Dont_Ask_Again.txt')
        #         file_path = Path(file_path)
        #         _folder_write_text_file(file_path,str(stored_do_not_ask_user_preference))

        #         if event in (sg.WINDOW_CLOSED , 'No'): #ask me every time
        #             enable_semi_automatic_mode = False
        #             break
        #         elif event == 'Yes': #do not ask me again
        #             enable_semi_automatic_mode = True
        #             stored_do_not_ask_user_preference = values['-DONT_ASK_AGAIN-']
        #             file_path = os.path.join(config_folder_path, 'Dont_Ask_Again.txt')
        #             file_path = Path(file_path)
        #             _folder_write_text_file(file_path,str(stored_do_not_ask_user_preference))
        #             break
        
        #     window.close()

        if not os.path.exists(bot_config_path):
            df = pd.DataFrame({'SNO': [],'KEY': [], 'VALUE':[]})
            append_df_to_excel(bot_config_path, df, index=False, startrow=0)
                
            # if enable_semi_automatic_mode:
            #     print("Semi Automatic Mode is ENABLED "+ show_emoji())
            # else:
            #     print("Semi Automatic Mode is DISABLED "+ show_emoji())
            
            # file_path = os.path.join(config_folder_path, 'Semi_Automatic_Mode.txt')
            # file_path = Path(file_path)
            # _folder_write_text_file(file_path,str(enable_semi_automatic_mode))
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in _ask_user_semi_automatic_mode " + str(ex))

def _excel_if_value_exists(excel_path="",sheet_name='Sheet1',header=0,usecols="",value=""):
    """
    Check if a given value exists in given excel. Returns True / False
    """
    try:
        
        if usecols:
            df = pd.read_excel(excel_path, sheet_name=sheet_name, header=header, usecols=usecols,engine='openpyxl')
        else:
            df = pd.read_excel(excel_path, sheet_name=sheet_name, header=header,engine='openpyxl')
        
        if value in df.values:
            df = ''
            return True
        else:
            df = ''
            return False

    except Exception:
        # print("Error in _excel_if_value_exists="+str(ex))
        return False

def _extract_filename_from_filepath(strFilePath=""):
    """
    Function which extracts file name from the given filepath
    """
    if strFilePath:
        try:
            strFileName = Path(strFilePath).name
            strFileName = str(strFileName).split(".")[0]

            return strFileName
        except Exception as ex:
            selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
            print("Error in _extract_filename_from_filepath="+str(ex))

    else:
        print("Please enter the value="+str(strFilePath))    

def _window_find_exact_name(windowName=""):
    """
    Gives you the exact window name you are looking for.

    Parameters:
        windowName  (str) : Name of the window to find.

    Returns:
        win (str)              : Exact window name.
        window_found (boolean) : A boolean TRUE if the window is found
    """
    win = ""
    window_found = False

    if not windowName:
        windowName = gui_get_any_input_from_user("Partial Window Name")

    try:
        lst = gw.getAllTitles()
        
        for item in lst:
            if str(item).strip():
                if str(windowName).lower() in str(item).lower():
                    win = item
                    window_found = True
                    break
        return win, window_found
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in _window_find_exact_name="+str(ex))

def _excel_copy_range(startCol=1, startRow=1, endCol=1, endRow=1, sheet='Sheet1'):
    """
    Copies the specific range from the given excel sheet.
    """
    try:
        rangeSelected = []
        #Loops through selected Rows
        for k in range(startRow,endRow + 1,1):
            #Appends the row to a RowSelected list
            rowSelected = []
            for l in range(startCol,endCol+1,1):
                rowSelected.append(sheet.cell(row = k, column = l).value)
            #Adds the RowSelected List and nests inside the rangeSelected
            rangeSelected.append(rowSelected)

        return rangeSelected

    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in _excel_copy_range="+str(ex))
    
def _excel_paste_range(startCol=1, startRow=1, endCol=1, endRow=1, sheetReceiving='Sheet1',copiedData=[]):
    """
    Pastes the specific range to the given excel sheet.
    """
    try:
        countRow = 0
        for k in range(startRow,endRow+1,1):
            countCol = 0
            for l in range(startCol,endCol+1,1):
                sheetReceiving.cell(row = k, column = l).value = copiedData[countRow][countCol]
                countCol += 1
            countRow += 1
        return countRow

    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in _excel_paste_range="+str(ex))

def get_public_ip():
    try:
        public_ip = str(requests.get('https://checkip.amazonaws.com').text.strip())
        return public_ip
    except:
        public_ip = str(requests.get('http://ip.42.pl/raw').text)
        return public_ip

class DisableLogger():
    def __enter__(self):
       logging.disable(logging.CRITICAL)
    def __exit__(self, exit_type, exit_value, exit_traceback):
       logging.disable(logging.NOTSET)

# ---------  Private Functions Ends ---------


# ---------  GUI Functions ---------

def gui_get_consent_from_user(msgForUser="Continue ?"):    
    """
    Generic function to get consent from user using GUI. Returns the yes or no

    Default Text: "Do you want to "
    """
    values = []
    try:
        oldValue = ""
        oldKey = msgForUser
        show_gui = False
        existing_value = read_semi_automatic_log(msgForUser)

        # if existing_value is None:
        #     show_gui = True

        if str(enable_semi_automatic_mode).lower() == 'false' :# and existing_value:
            show_gui = True
            oldValue = existing_value
            
        if show_gui:
            layout = [[sg.Text("ClointFusion - Set Yourself Free for Better Work", font='Courier 16',text_color='orange')],
                [sg.Text('Do you want to '),sg.Text(text=oldKey,font=('Courier 12'),text_color='yellow')],
                [sg.Submit('Yes',button_color=('white','green'),font=('Courier 14'),bind_return_key=True),sg.Submit('No',button_color=('white','firebrick'),font=('Courier 14'))]]

            window = sg.Window('ClointFusion',layout, return_keyboard_events=True,use_default_focus=True,disable_close=False,element_justification='c',keep_on_top=True, finalize=True,icon=cf_icon_file_path)

            while True:
                event, values = window.read()
                if event == 'No':
                    oldValue = 'No'
                    break
                if event == 'Yes':
                    oldValue = 'Yes'
                    break
                        
            window.close()
            values['-KEY-'] = msgForUser

            if str(values['-KEY-']):
                update_semi_automatic_log(str(values['-KEY-']).strip(),str(oldValue))
        
            return oldValue

        else:
            return str(existing_value)

    except ValueError:
        print("Please check the input values, and try again.")
        text_to_speech("Please check the input values, and try again.", show=False)
        sys.exit()
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in gui_get_consent_from_user="+str(ex))

def gui_get_dropdownlist_values_from_user(msgForUser="",dropdown_list=[],multi_select=True): 
    """
    Generic function to accept one of the drop-down value from user using GUI. Returns all chosen values in list format.

    Default Text: "Please choose the item(s) from "
    """
    try:
        values = []
        dropdown_list = dropdown_list

        if dropdown_list:
            try:
                oldValue = []
                oldKey = msgForUser
                show_gui = False
                existing_value = read_semi_automatic_log(msgForUser)
                
                # if existing_value is None:
                #     show_gui = True

                if str(enable_semi_automatic_mode).lower() == 'false' :#and existing_value:
                    show_gui = True
                    oldValue = existing_value
                                
                if show_gui:
                    if multi_select:
                        layout = [[sg.Text("ClointFusion - Set Yourself Free for Better Work", font='Courier 16', text_color='orange')],
                                [sg.Text('Please choose the item(s) from '),sg.Text(text=oldKey,font=('Courier 12'),text_color='yellow'),sg.Listbox(dropdown_list,size=(30, 5),key='-EXCELCOL-',default_values=oldValue,select_mode=sg.LISTBOX_SELECT_MODE_MULTIPLE,enable_events=True,change_submits=True)],#oldExcelCols
                                [sg.Submit('OK',button_color=('white','green'),bind_return_key=True),sg.CloseButton('Cancel',button_color=('white','firebrick'))]]

                    else:
                        layout = [[sg.Text("ClointFusion - Set Yourself Free for Better Work", font='Courier 16', text_color='orange')],
                                [sg.Text('Please choose an item from '),sg.Text(text=oldKey,font=('Courier 12'),text_color='yellow'),sg.Listbox(dropdown_list,size=(30, 5),key='-EXCELCOL-',default_values=oldValue,select_mode=sg.LISTBOX_SELECT_MODE_SINGLE,enable_events=True,change_submits=True)],#oldExcelCols
                                [sg.Submit('OK',button_color=('white','green'),bind_return_key=True),sg.CloseButton('Cancel',button_color=('white','firebrick'))]]

                    window = sg.Window('ClointFusion',layout, return_keyboard_events=True,use_default_focus=False,disable_close=False,element_justification='c',keep_on_top=True, finalize=True,icon=cf_icon_file_path)

                    while True:                
                        event, values = window.read()
                        
                        if event is None or event == 'Cancel' or event == "Escape:27":
                            values = []
                            break

                        if event == 'OK':
                            if values and values['-EXCELCOL-']:
                                break
                            else:
                                message_pop_up("Please enter all the values")

                    window.close()

                    if values and event == 'OK':
                        values['-KEY-'] = msgForUser
                        
                        if str(values['-KEY-']) and str(values['-EXCELCOL-']):
                            update_semi_automatic_log(str(values['-KEY-']).strip(),str(values['-EXCELCOL-']).strip())

                        return values['-EXCELCOL-']
                    else:
                        return oldValue
                
                else:
                    return oldValue
                    
            except Exception as ex:
                selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
                print("Error in gui_get_dropdownlist_values_from_user="+str(ex))
        else:
            print('gui_get_dropdownlist_values_from_user - List is empty')
    except ValueError:
        print("Please check the input values, and try again.")
        text_to_speech("Please check the input values, and try again.", show=False)
        sys.exit()
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in gui_get_dropdownlist_values_from_user="+str(ex))

def gui_get_excel_sheet_header_from_user(msgForUser=""): 
    """
    Generic function to accept excel path, sheet name and header from user using GUI. Returns all these values in disctionary format.

    Default Text: "Please choose the excel "
    """
    values = []
    sheet_namesLst = []
    try:
        oldValue = "" + "," + "Sheet1" + "," + "0"
        oldKey = msgForUser
        show_gui = False
        existing_value = read_semi_automatic_log(msgForUser)
        
        if existing_value is None:
            show_gui = True

        if str(enable_semi_automatic_mode).lower() == 'false' and existing_value:
            show_gui = True
            oldValue = existing_value
            
        if show_gui:
            oldFilePath, oldSheet , oldHeader = str(oldValue).split(",")
    
            layout = [[sg.Text("ClointFusion - Set Yourself Free for Better Work", font='Courier 16', text_color='orange')],
                    [sg.Text('Please choose the excel '),sg.Text(text=oldKey,font=('Courier 12'),text_color='yellow'),sg.Input(default_text=oldFilePath,key="-FILEPATH-",enable_events=True,change_submits=True), sg.FileBrowse(file_types=(("Excel File", "*.xlsx"),("Excel File", "*.xlsx")))], 
                    [sg.Text('Sheet Name'), sg.Combo(sheet_namesLst,default_value=oldSheet,size=(20, 0),key="-SHEET-",enable_events=True)], 
                    [sg.Text('Choose the header row'),sg.Spin(values=('0', '1', '2', '3', '4', '5'),initial_value=int(oldHeader),key="-HEADER-",enable_events=True,change_submits=True)],
                    # [sg.Checkbox('Use this excel file for all the excel related operations of this BOT',enable_events=True, key='-USE_THIS_EXCEL-',default=old_Use_This_excel, text_color='yellow')],
                    [sg.Submit('OK',button_color=('white','green'),bind_return_key=True),sg.CloseButton('Cancel',button_color=('white','firebrick'))]]
        
            window = sg.Window('ClointFusion',layout, return_keyboard_events=True,use_default_focus=False,disable_close=False,element_justification='c',keep_on_top=True, finalize=True,icon=cf_icon_file_path)

            while True:
                if oldFilePath: 
                    sheet_namesLst = excel_get_all_sheet_names(oldFilePath)
                    window['-SHEET-'].update(values=sheet_namesLst)   
                
                event, values = window.read()
                
                if event is None or event == 'Cancel' or event == "Escape:27":
                    values = []
                    break

                if event == 'OK':
                    if values and values['-FILEPATH-'] and values['-SHEET-']:
                        break
                    else:
                        message_pop_up("Please enter all the values")

                if event == '-FILEPATH-':
                    sheet_namesLst = excel_get_all_sheet_names(values['-FILEPATH-'])
                    window['-SHEET-'].update(values=sheet_namesLst)   
                    window.refresh()
                    oldFilePath = ""

                    if len(sheet_namesLst) >= 1:
                        window['-SHEET-'].update(value=sheet_namesLst[0]) 

                if event == '-SHEET-':
                    window['-SHEET-'].update(value=values['-SHEET-'])

            window.close()

            if values: 
                values['-KEY-'] = msgForUser
                
                concatenated_value = values['-FILEPATH-'] + "," +  values ['-SHEET-'] + "," + values['-HEADER-']
                
                if str(values['-KEY-']) and concatenated_value:
                    update_semi_automatic_log(str(values['-KEY-']).strip(),str(concatenated_value))

                return values['-FILEPATH-'] , values ['-SHEET-'] , int(values['-HEADER-'])

            else:    
                oldFilePath, oldSheet , oldHeader = str(existing_value).split(",")
                return oldFilePath, oldSheet , int(oldHeader)
        
        else:
            oldFilePath, oldSheet , oldHeader = str(existing_value).split(",")
            return oldFilePath, oldSheet , int(oldHeader)
            
    except TypeError:
        print("Please check the input values, and try again.")
        text_to_speech("Please check the input values, and try again.", show=False)
    except ValueError:
        print("Please check the input values, and try again.")
        text_to_speech("Please check the input values, and try again.", show=False)
        sys.exit()
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in gui_get_excel_sheet_header_from_user="+str(ex))
    
def gui_get_folder_path_from_user(msgForUser="the folder : "):    
    """
    Generic function to accept folder path from user using GUI. Returns the folderpath value in string format.

    Default text: "Please choose "
    """
    values = []
    try:
        oldValue = ""
        oldKey = msgForUser
        show_gui = False
        existing_value = read_semi_automatic_log(msgForUser)

        if existing_value is None:
            show_gui = True

        if str(enable_semi_automatic_mode).lower() == 'false' and existing_value:
            show_gui = True
            oldValue = existing_value
            
        if show_gui:
            layout = [[sg.Text("ClointFusion - Set Yourself Free for Better Work", font='Courier 16',text_color='orange')],
                [sg.Text('Please choose '),sg.Text(text=oldKey,font=('Courier 12'),text_color='yellow'),sg.Input(default_text=oldValue ,key='-FOLDER-', enable_events=True), sg.FolderBrowse()],
                [sg.Submit('OK',button_color=('white','green'),bind_return_key=True),sg.CloseButton('Cancel',button_color=('white','firebrick'))]]

            window = sg.Window('ClointFusion',layout, return_keyboard_events=True,use_default_focus=True,disable_close=False,element_justification='c',keep_on_top=True,finalize=True,icon=cf_icon_file_path)

            while True:
                event, values = window.read()

                if event == sg.WINDOW_CLOSED or event == 'Cancel':
                    break
                if event == 'OK':
                    if values and values['-FOLDER-']:
                        break
                    else:
                        message_pop_up("Please enter the required values")
            
            window.close()

            if values and event == 'OK':
                values['-KEY-'] = msgForUser

                if str(values['-KEY-']) and str(values['-FOLDER-']):
                    update_semi_automatic_log(str(values['-KEY-']).strip(),str(values['-FOLDER-']).strip())
            
                if values is not None:
                    return str(values['-FOLDER-']).strip()
            else:
                return None

        else:
            return str(existing_value)

    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in gui_get_folder_path_from_user="+str(ex))

def gui_get_any_input_from_user(msgForUser="Please enter : ",password=False,multi_line=False,mandatory_field=True):   
    import cryptocode 
    from pywebio.input import PASSWORD, TEXT, textarea 

    """
    Generic function to accept any input (text / numeric) from user using GUI. Returns the value in string format.
    Please use unique message (key) for each value.

    Default Text: "Please enter "
    """
    values = []
    try:
        oldValue = ""
        oldKey = msgForUser
        show_gui = False
        existing_value = read_semi_automatic_log(msgForUser)
        
        if existing_value == "nan":
            existing_value = None
            
        if existing_value is None:
            show_gui = True
        
        if str(enable_semi_automatic_mode).lower() == 'false' and existing_value:
            show_gui = True
            oldValue = existing_value
        
        layout = ""
        if show_gui:
            if password:
                layout = [[sg.Text("ClointFusion - Set Yourself Free for Better Work", font='Courier 16',text_color='orange')],
                    [sg.Text('Please enter '),sg.Text(text=oldKey,font=('Courier 12'),text_color='yellow'),sg.Input(default_text=str(cryptocode.decrypt(oldValue, "ClointFusion")).strip(),key='-VALUE-', justification='c',password_char='*')],
                    [sg.Text('This field is mandatory',text_color='red')],
                    [sg.Submit('OK',button_color=('white','green'),bind_return_key=True),sg.CloseButton('Cancel',button_color=('white','firebrick'))]]

            elif not password and mandatory_field and multi_line:
                layout = [[sg.Text("ClointFusion - Set Yourself Free for Better Work", font='Courier 16',text_color='orange')],
                    [sg.Text('Please enter '),sg.Text(text=oldKey,font=('Courier 12'),text_color='yellow'),sg.MLine(default_text=oldValue,key='-VALUE-', size=(40,8),justification='l')],
                    [sg.Text('This field is mandatory',text_color='red')],
                    [sg.Submit('OK',button_color=('white','green'),bind_return_key=True),sg.CloseButton('Cancel',button_color=('white','firebrick'))]]

            elif not password and mandatory_field and not multi_line:
                layout = [[sg.Text("ClointFusion - Set Yourself Free for Better Work", font='Courier 16',text_color='orange')],
                    [sg.Text('Please enter '),sg.Text(text=oldKey,font=('Courier 12'),text_color='yellow'),sg.Input(default_text=oldValue,key='-VALUE-', justification='c')],
                    [sg.Text('This field is mandatory',text_color='red')],
                    [sg.Submit('OK',button_color=('white','green'),bind_return_key=True),sg.CloseButton('Cancel',button_color=('white','firebrick'))]]

            elif not password and not mandatory_field and multi_line:
                layout = [[sg.Text("ClointFusion - Set Yourself Free for Better Work", font='Courier 16',text_color='orange')],
                    [sg.Text('Please enter '),sg.Text(text=oldKey,font=('Courier 12'),text_color='yellow'),sg.MLine(default_text=oldValue,key='-VALUE-',size=(40,8), justification='l')],
                    [sg.Text('You may leave this field empty',text_color='orange')],
                    [sg.Submit('OK',button_color=('white','green'),bind_return_key=True),sg.CloseButton('Cancel',button_color=('white','firebrick'))]]

            elif not password and not mandatory_field and not multi_line:
                layout = [[sg.Text("ClointFusion - Set Yourself Free for Better Work", font='Courier 16',text_color='orange')],
                    [sg.Text('Please enter '),sg.Text(text=oldKey,font=('Courier 12'),text_color='yellow'),sg.Input(default_text=oldValue,key='-VALUE-', justification='c')],
                    [sg.Text('You may leave this field empty',text_color='orange')],
                    [sg.Submit('OK',button_color=('white','green'),bind_return_key=True),sg.CloseButton('Cancel',button_color=('white','firebrick'))]]


            window = sg.Window('ClointFusion',layout, return_keyboard_events=True,use_default_focus=True,disable_close=True,element_justification='c',keep_on_top=True,finalize=True,icon=cf_icon_file_path)

            while True:
                
                event, values = window.read()

                if event == sg.WINDOW_CLOSED  or event == 'Cancel':
                    
                    if oldValue or (values and values['-VALUE-']):
                        break

                    else:
                        if mandatory_field:
                            message_pop_up("Its a mandatory field !.. Cannot proceed, exiting now..")
                            print("Exiting ClointFusion, as Mandatory field is missing")
                            sys.exit(0)
                        else:
                            print("Mandatory field is missing, continuing with None/Empty value")
                            break
                
                if event == 'OK':
                    if values and values['-VALUE-']:
                        break
                    else:
                        if mandatory_field:
                            message_pop_up("This value is required. Please enter the value..")
                        else:
                            break
            
            window.close()

            if values and event == 'OK':
                values['-KEY-'] = msgForUser
            
            if values is not None and str(values['-KEY-']) and str(values['-VALUE-']):
                if password:
                    update_semi_automatic_log(str(values['-KEY-']).strip(),cryptocode.encrypt(str(values['-VALUE-']).strip(),"ClointFusion"))
                else:
                    update_semi_automatic_log(str(values['-KEY-']).strip(),str(values['-VALUE-']).strip())

            if values is not None and str(values['-VALUE-']):
                return str(values['-VALUE-']).strip()
            else:
                return None
        
        else:
            return str(existing_value)

    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        try:
            msgForUser = "Please enter " + msgForUser

            if not password and not mandatory_field and not multi_line:
                existing_value = input(msgForUser, type=TEXT, required=False)

            elif not password and mandatory_field and multi_line:
                existing_value = textarea(msgForUser,rows=6, required=True)
                
            elif not password and mandatory_field and not multi_line:
                existing_value = input(msgForUser, type=TEXT,required=mandatory_field)    
            
            elif not password and not mandatory_field and multi_line:
                existing_value = textarea(msgForUser,rows=6, required=False)
            
            elif password:
                existing_value = input(msgForUser, type=PASSWORD, required=mandatory_field)

            return existing_value
        except:
            print("Error in gui_get_any_input_from_user=" + str(ex))

def gui_get_any_file_from_user(msgForUser="the file : ",Extension_Without_Dot="*"):    
    """
    Generic function to accept file path from user using GUI. Returns the filepath value in string format.Default allows all files i.e *

    Default Text: "Please choose "
    """
    values = []
    try:
        oldValue = ""
        oldKey = msgForUser
        show_gui = False
        existing_value = read_semi_automatic_log(msgForUser)

        if existing_value is None:
            show_gui = True

        if str(enable_semi_automatic_mode).lower() == 'false' and existing_value:
            show_gui = True
            oldValue = existing_value
            
        if show_gui:
            layout = [[sg.Text("ClointFusion - Set Yourself Free for Better Work", font='Courier 16',text_color='orange')],
                [sg.Text('Please choose '),sg.Text(text=oldKey + " (ending with .{})".format(str(Extension_Without_Dot).lower()),font=('Courier 12'),text_color='yellow'),sg.Input(default_text=oldValue ,key='-FILE-', enable_events=True), sg.FileBrowse(file_types=((".{} File".format(Extension_Without_Dot), "*.{}".format(Extension_Without_Dot)),))],
                [sg.Submit('OK',button_color=('white','green'),bind_return_key=True),sg.CloseButton('Cancel',button_color=('white','firebrick'))]]

            window = sg.Window('ClointFusion',layout, return_keyboard_events=True,use_default_focus=True,disable_close=False,element_justification='c',keep_on_top=True, finalize=True,icon=cf_icon_file_path)

            while True:
                    event, values = window.read()
                    if event == sg.WINDOW_CLOSED  or event == 'Cancel':
                        break
                    if event == 'OK':
                        if values and values['-FILE-']:
                            break
                        else:
                            message_pop_up("Please enter the required values")
                            # print("Please enter the values")
            window.close()

            if values and event == 'OK':
                values['-KEY-'] = msgForUser

                if str(values['-KEY-']) and str(values['-FILE-']):
                    update_semi_automatic_log(str(values['-KEY-']).strip(),str(values['-FILE-']).strip())
            
                if values is not None and str(values['-FILE-']):
                    return str(values['-FILE-']).strip()
            else:
                return None

        else:
            return str(existing_value)

    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in gui_get_any_file_from_user="+str(ex))

# ---------  GUI Functions Ends---------


# ---------  Mouse Functions --------- 
    
def mouse_click(x='', y='', left_or_right="left", no_of_clicks=1):
    """Clicks at the given X Y Co-ordinates on the screen using single / double / triple click(s). Default clicks on current position.

    Args:
        x (int): x-coordinate on screen.
        Eg: 369 or 435, Defaults: ''.
        y (int): y-coordinate on screen.
        Eg: 369 or 435, Defaults: ''.
        left_or_right (str, optional): Which mouse button.
        Eg: right or left, Defaults: left.
        no_of_click (int, optional): Number of times specified mouse button to be clicked.
        Eg: 1 or 2, Max 3. Defaults: 1.

    Returns:
        bool: Whether the function is successful or failed.
    """
    status = False
    try:
        
        if not x or not y:
            x, y = pg.position()
            
        time.sleep(1)
        
        if x and y:
            x,y = int(x), int(y)
            no_of_clicks = 3 if no_of_clicks > 3 else no_of_clicks
            pg.click(x,y, clicks=no_of_clicks, button=left_or_right)
            status = True
            time.sleep(1)
        status = True
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in mouseClick="+str(ex))
    finally:
        return status

def mouse_move(x="",y=""):
    """Moves the cursor to the given X Y Co-ordinates.

    Args:
        x (int): x-coordinate on screen.
        Eg: 369 or 435, Defaults: ''.
        y (int): y-coordinate on screen.
        Eg: 369 or 435, Defaults: ''.

    Returns:
        bool: Whether the function is successful or failed.
    """
    status = False
    try:
        
        if not x or not y:
            if not x and not y:
                x_y = str(gui_get_any_input_from_user("X,Y co-ordinates to the move Mouse to. Ex: 369,435"))
                if "," in x_y:
                    x, y = x_y.split(",")
                elif " " in x_y:
                    x, y = x_y.split(" ")
            if not x and y:
                x = str(gui_get_any_input_from_user(f"Enter 'X' co-ordinate Y={y} to move mouse. Ex: 369"))
            if not y and x:
                y = str(gui_get_any_input_from_user(f"Enter 'Y' co-ordinate X={x} to move mouse. Ex: 369"))
        if x and y:
            x,y = int(x), int(y)
            time.sleep(0.2)
            pg.moveTo(x,y)
            time.sleep(0.2)
        status = True
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in mouse_move="+str(ex))
    finally:
        return status

def mouse_drag_from_to(x1="",y1="",x2="",y2="",delay=0.5):
    """Clicks and drags from x1 y1 co-ordinates to x2 y2 Co-ordinates on the screen

    Args:
        x1 or x2 (int): x-coordinate on screen.
        Eg: 369 or 435, Defaults: ''.
        y1 or y2 (int): y-coordinate on screen.
        Eg: 369 or 435, Defaults: ''.
        delay (float, optional): Seconds to wait while performing action. 
        Eg: 1 or 0.5, Defaults to 0.5.

    Returns:
        bool: Whether the function is successful or failed.
    """
    status = False
    try:
        
        if not x1 or not y1:
            if not x1 and not y1:
                x_y_1 = str(gui_get_any_input_from_user("FROM X,Y co-ordinates for Mouse Drag Ex: 200,215"))
                if "," in x_y_1:
                    x1, y1 = x_y_1.split(",")
                elif " " in x_y_1:
                    x1, y1 = x_y_1.split(" ")
            if not x1 and y1:
                x1 = str(gui_get_any_input_from_user(f"Enter 'X' co-ordinate Y={y1} to perform Mouse Drag FROM Ex: 369"))
            if not y1 and x1:
                y1 = str(gui_get_any_input_from_user(f"Enter 'Y' co-ordinate X={x1} to perform Mouse Drag FROM. Ex: 369"))

        if not x2 or not y2:
            if not x2 and not y2:
                x_y_2 = str(gui_get_any_input_from_user("TO X,Y co-ordinates for Mouse Drag Ex: 200,215"))
                if "," in x_y_2:
                    x2, y2 = x_y_2.split(",")
                elif " " in x_y_2:
                    x2, y2 = x_y_2.split(" ")
            if not x2 and y2:
                x2 = str(gui_get_any_input_from_user(f"Enter 'X' co-ordinate Y={y1} to perform Mouse Drag TO Ex: 369"))
            if not y2 and x2:
                y2 = str(gui_get_any_input_from_user(f"Enter 'Y' co-ordinate X={x1} to perform Mouse Drag TO. Ex: 369"))
                
        time.sleep(0.2)
        x1,y1,x2,y2 = int(x1), int(y1), int(x2), int(y2)
        pg.moveTo(x1,y1,duration=delay)
        pg.dragTo(x2,y2,duration=delay,button='left')
        time.sleep(0.2)
        status = True
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in mouse_drag_from_to="+str(ex))
    finally:
        return status

def mouse_search_snip_return_coordinates_x_y(img="", wait=180):
    """Searches the given image on the screen and returns its center of X Y co-ordinates.

    Args:
        img (str, optional): Path of the image. 
        Eg: D:\Files\Image.png, Defaults to "".
        wait (int, optional): Time you want to wait while program searches for image repeatably.
        Eg: 10 or 100 Defaults to 180.
        
    Returns:
        bool: If function is failed returns False.
        tuple (x, y): Image Center co-ordinates.
    """
    status = False
    try:
        region=(0,0,pg.size()[0],pg.size()[1])
        if not img:
            img = gui_get_any_file_from_user("snip image file, to get X,Y coordinates","png")

        time.sleep(1)

        pos = pg.locateOnScreen(img,region=region) 
        i = 0
        while pos == None and i < int(wait):
            pos = ()
            pos = pg.locateOnScreen(img,region=region)   
            time.sleep(1)
            i = i + 1

        time.sleep(1)

        if pos:
            x,y = pos.left + int(pos.width / 2), pos.top + int(pos.height / 2)
            status = (x,y)
    except AttributeError:
        print("Invalid Input. Please check the given input.")
        text_to_speech("Invalid Input. Please check the given input.", show=False)
        sys.exit()
    except FileNotFoundError:
        print("File not found. Please check the given input.")
        text_to_speech("File not found. Please check the given input.", show=False)
        sys.exit()
    except TypeError:
        print("Invalid Input. Please check the given input.")
        text_to_speech("Invalid Input. Please check the given input.", show=False)
        sys.exit()
    except OSError:
        print('File is missing, has improper permissions, or is an unsupported or invalid format. Please check and try again.')
        text_to_speech('File is missing, has improper permissions, or is an unsupported or invalid format. Please check and try again.', show=False)
        sys.exit()
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in mouse_search_snip_return_coordinates_x_y="+str(ex))
    finally:
        return status

# ---------  Mouse Functions Ends --------- 


# ---------  Keyboard Functions --------- 

def key_press(key_1='', key_2='', key_3='', write_to_window=""):
    """Emulates the given keystrokes.

    Args:
        key_1 (str, optional): Enter the 1st key 
        Eg: ctrl or shift. Defaults to ''.
        key_2 (str, optional): Enter the 2nd key in combination. 
        Eg: alt or A. Defaults to ''.
        key_3 (str, optional): Enter the 3rd key in combination. 
        Eg: del or tab. Defaults to ''.
        write_to_window (str, optional): (Only in Windows) Name of Window you want to activate. Defaults to "".
        
    Supported Keys:
        ['\\t', '\\n', '\\r', ' ', '!', '"', '#', '$', '%', '&', "'", '(',')', '*', '+', ',', '-', '.', '/', 
        '0', '1', '2', '3', '4', '5', '6', '7','8', '9', ':', ';', '<', '=', '>', '?', '@', '[', '\\', ']', '^', '_', '`', 
        'a', 'b', 'c', 'd', 'e','f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p', 'q', 'r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z', 
        '{', '|', '}', '~', 'accept', 'add', 'alt', 'altleft', 'altright', 'apps', 'backspace',
        'browserback', 'browserfavorites', 'browserforward', 'browserhome',
        'browserrefresh', 'browsersearch', 'browserstop', 'capslock', 'clear',
        'convert', 'ctrl', 'ctrlleft', 'ctrlright', 'decimal', 'del', 'delete',
        'divide', 'down', 'end', 'enter', 'esc', 'escape', 'execute', 'f1', 'f10',
        'f11', 'f12', 'f13', 'f14', 'f15', 'f16', 'f17', 'f18', 'f19', 'f2', 'f20',
        'f21', 'f22', 'f23', 'f24', 'f3', 'f4', 'f5', 'f6', 'f7', 'f8', 'f9',
        'final', 'fn', 'hanguel', 'hangul', 'hanja', 'help', 'home', 'insert', 'junja',
        'kana', 'kanji', 'launchapp1', 'launchapp2', 'launchmail',
        'launchmediaselect', 'left', 'modechange', 'multiply', 'nexttrack',
        'nonconvert', 'num0', 'num1', 'num2', 'num3', 'num4', 'num5', 'num6',
        'num7', 'num8', 'num9', 'numlock', 'pagedown', 'pageup', 'pause', 'pgdn',
        'pgup', 'playpause', 'prevtrack', 'print', 'printscreen', 'prntscrn',
        'prtsc', 'prtscr', 'return', 'right', 'scrolllock', 'select', 'separator',
        'shift', 'shiftleft', 'shiftright', 'sleep', 'space', 'stop', 'subtract', 'tab',
        'up', 'volumedown', 'volumemute', 'volumeup', 'win', 'winleft', 'winright', 'yen',
        'command', 'option', 'optionleft', 'optionright']
    
    Returns:
        bool: Whether the function is successful or failed.
    """
    status = False
    try:
        if not key_1:
            key_list = gui_get_any_input_from_user("keys combination using + as delimeter. Ex: ctrl or tab")
            
            if key_list.count("+") == 0:
                key_1 = key_list[0]
            elif key_list.count("+") == 1:
                key_1, key_2 = key_list.split("+")
            elif key_list.count("+") == 2:
                key_1, key_2, key_3 = key_list.split("+")
            else:
                print("Invalid keys combination")
                return False

        if os_name == windows_os:
            if write_to_window:
                    window_activate_and_maximize_windows(write_to_window)
        
        time.sleep(0.5)
        pg.hotkey(key_1,key_2,key_3)
        time.sleep(0.5)
        status = True
    except pg.FailSafeException:
        print('Mouse moving to a corner of the screen.')
        text_to_speech('Mouse moving to a corner of the screen.', show=False)
        sys.exit()
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in key_press="+str(ex))
    finally:
        return status
    
def key_write_enter(text_to_write="", write_to_window="", delay_after_typing=1, key="e"):
    """Writes/Types the given text.

    Args:
        text_to_write (str, optional): Text you wanted to type
        Eg: ClointFusion is awesone. Defaults to "".
        write_to_window (str, optional): (Only in Windows) Name of Window you want to activate
        Eg: Notepad. Defaults to "".
        delay_after_typing (int, optional): Seconds in time to wait after entering the text
        Eg: 5. Defaults to 1.
        key (str, optional): Press Enter key after typing.
        Eg: t for tab. Defaults to e

    Returns:
        bool: Whether the function is successful or failed.
    """
    
    status = False
    try:
        if not text_to_write:
            text_to_write = gui_get_any_input_from_user("message / username / any text")

        if os_name == windows_os:
            if write_to_window:
                    window_activate_and_maximize_windows(write_to_window)

        time.sleep(0.2)
        pg.write(text_to_write)
        if key.lower() ==  "e":
            pg.hotkey("enter")
        if key.lower() == "t":
            pg.hotkey("tab")
        time.sleep(delay_after_typing)
        status = True
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in key_write_enter="+str(ex))
    finally:
        return status

def key_hit_enter(write_to_window=""):
    """Enter key will be pressed once.

    Args:
        write_to_window (str, optional): (Only in Windows)Name of Window you want to activate.
        Eg: Notepad. Defaults to "".

    Returns:
        bool: Whether the function is successful or failed.
    """
    status = False
    try:
        time.sleep(0.5)
        key_press(key_1="enter", write_to_window=write_to_window)
        time.sleep(0.5)
        status = True
    except UnboundLocalError:
        print('Local variable referenced before assignment')
        text_to_speech('Local variable referenced before assignment', show=False)
        sys.exit()
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in key_hit_enter="+str(ex))
    finally:
        return status

# --------- Keyboard Functions Ends --------- 


# ---------  Message  Functions --------- 

def message_counter_down_timer(strMsg="Calling ClointFusion Function in (seconds)",start_value=5):
    """
    Function to show count-down timer. Default is 5 seconds.
    Ex: message_counter_down_timer()
    """
    CONTINUE = True
    layout = [[sg.Text(strMsg,justification='c')],[sg.Text('',size=(10, 0),font=('Helvetica', 20),justification='c', key='text')],
            [sg.Image(filename = str(cf_logo_file_path),size=(60,60))],
            [sg.Exit(button_color=('white', 'firebrick4'), key='Cancel')]]
    
    window = sg.Window('ClointFusion - Countdown Timer', layout, no_titlebar=True, auto_size_buttons=False,keep_on_top=True, grab_anywhere=False, element_justification='c',element_padding=(0, 0),finalize=True,icon=cf_icon_cdt_file_path)
    
    current_value = start_value + 1

    while True:
        event, _ = window.read(timeout=2)
        current_value = current_value - 1
        time.sleep(1)
            
        if current_value == 0:
            CONTINUE = True
            break
            
        if event in (sg.WINDOW_CLOSED , 'Cancel'):    
            CONTINUE = False  
            print("Action cancelled by user")
            break

        window['text'].update(value=current_value)

    window.close()
    return CONTINUE

def message_pop_up(strMsg="",delay=3):
    """
    Specified message will popup on the screen for a specified duration of time.

    Parameters:
        strMsg  (str) : message to popup.
        delay   (int) : duration of the popup.
    """
    try:
        # if not strMsg:
        #     strMsg = gui_get_any_input_from_user("pop-up message")
        sg.popup(strMsg,title='ClointFusion',auto_close_duration=delay, auto_close=True, keep_on_top=True,background_color="white",text_color="black")#,icon=cloint_ico_logo_base64)
    except:
        popup('', [put_html('<img src="https://raw.githubusercontent.com/ClointFusion/Image_ICONS_GIFs/main/Cloint-LOGO-New.png">'),
        put_html('<center><h3>' + strMsg + '</h3></center>'),])

def message_flash(msg="",delay=3):
    """
    specified msg will popup for a specified duration of time with OK button.

    Parameters:
        msg     (str) : message to popup.
        delay   (int) : duration of the popup.
    """
    try:
        if not msg:
            msg = gui_get_any_input_from_user("flash message")

        r = Timer(int(delay), key_hit_enter)
        r.start()
        pg.alert(text=msg, title='ClointFusion', button='OK')
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in message_flash="+str(ex))

def message_toast(message,website_url="", file_folder_path=""):
    """
    Function for displaying Windows 10 Toast Notifications.
    Pass website URL OR file / folder path that needs to be opened when user clicks on the toast notification.
    """
    
    if os_name == windows_os:

        if str(enable_semi_automatic_mode).lower() == 'false':
            from win10toast_click import ToastNotifier 
            toaster = ToastNotifier()

            if website_url:

                toaster.show_toast(
                    "ClointFusion", 
                    "{}. Click to open URL".format(message), 
                    icon_path=cf_icon_cdt_file_path,
                    duration=5, # for how many seconds toast should be visible; None = leave notification in Notification Center
                    threaded=True, # True = run other code in parallel; False = code execution will wait till notification disappears 
                    callback_on_click=lambda: webbrowser.open_new(website_url) # click notification to run function 
                )

            elif file_folder_path:
                toaster.show_toast(
                    "ClointFusion", 
                    "{}. Click to open".format(message), 
                    icon_path=cf_icon_cdt_file_path,
                    duration=5, # for how many seconds toast should be visible; None = leave notification in Notification Center
                    threaded=True, # True = run other code in parallel; False = code execution will wait till notification disappears 
                    callback_on_click=lambda: os.startfile(file_folder_path) # click notification to run function 
                )

            else:
                toaster.show_toast(
                    "ClointFusion", # title
                    message, # message 
                    icon_path=cf_icon_cdt_file_path, # 'icon_path' 
                    duration=5, # for how many seconds toast should be visible; None = leave notification in Notification Center
                    threaded=True, # True = run other code in parallel; False = code execution will wait till notification disappears 
            )
        else:
            print("This function works when semi-automatic mode is enabled")    

    else:
        print("This function works only on Windows OS")

# ---------  Message  Functions Ends ---------


# ---------  Browser Functions --------- 
    
def browser_activate(url="", files_download_path='', dummy_browser=True, incognito=False,
                     clear_previous_instances=False, profile="Default"):
    """Function to launch browser and start the session.

    Args:
        url (str, optional): Website you want to visit. Defaults to "".
        files_download_path (str, optional): Path to which the files need to be downloaded.
        Defaults: ''.
        dummy_browser (bool, optional): If it is false Default profile is opened. 
        Defaults: True.
        incognito (bool, optional): Opens the browser in incognito mode. 
        Defaults: False.
        clear_previous_instances (bool, optional): If true all the opened chrome instances are closed. 
        Defaults: False.
        profile (str, optional): By default it opens the 'Default' profile. 
        Eg : Profile 1, Profile 2

    Returns:
        bool: Whether the function is successful or failed.
    """
    status = False
    global browser_driver, helium_service_launched

    try:
        # To clear previous instances of chrome
        if clear_previous_instances:
            if os_name == windows_os:
                try:
                    subprocess.call('TASKKILL /IM chrome.exe', stdout=subprocess.DEVNULL, stderr=subprocess.STDOUT)
                except Exception as ex:
                    selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
                    print(f"Error while closing previous chrome instances. {ex}")
            elif os_name == mac_os:
                try:
                    subprocess.call('pkill "Google Chrome"', shell=True)
                except Exception as ex:
                    selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
                    print(f"Error while closing previous chrome instances. {ex}")
            elif os_name == linux_os:
                try:
                    subprocess.call('sudo pkill -9 chrome', shell=True)
                except Exception as ex:
                    selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
                    print(f"Error while closing previous chrome instances. {ex}")
            time.sleep(5)

        options = Options()
        options.add_argument("--start-maximized")
        options.add_experimental_option('excludeSwitches', ['enable-logging','enable-automation'])
        if os_name == linux_os:
            options.add_argument('--no-sandbox')
            options.add_argument('--disable-dev-shm-usage') 
        if incognito:
            options.add_argument("--incognito")
        if not dummy_browser:
            if os_name == windows_os:
                options.add_argument("user-data-dir=C:\\Users\\{}\\AppData\\Local\\Google\\Chrome\\User Data".format(os.getlogin()))
                options.add_argument(f"profile-directory={profile}")
            elif os_name == mac_os:
                options.add_argument("user-data-dir=/Users/{}/Library/Application/Support/Google/Chrome/User Data".format(os.getlogin()))
                options.add_argument(f"profile-directory={profile}")
            elif os_name == linux_os:
                options.add_argument("user-data-dir=/home/{}/.config/google-chrome".format(os.getlogin()))
                
            
        #  Set the download path
        if files_download_path != '':
            prefs = {
                'download.default_directory': files_download_path,
                "download.prompt_for_download": False,
                'download.directory_upgrade': True,
                "safebrowsing.enabled": False
            }
            options.add_experimental_option('prefs', prefs)

        try:
            with DisableLogger():
                browser_driver = webdriver.Chrome(ChromeDriverManager().install(), options=options)
            browser.set_driver(browser_driver)
            if url:
                browser.go_to(url)
            if not url:
                browser.go_to("https://dost.clointfusion.com")
            status = True
            browser.Config.implicit_wait_secs = 60
            helium_service_launched = True
        except InvalidArgumentException:
            print("Another Chrome Window is already in use. If you want to open your custom profile, please close all the chrome windows and try again.  Else make dummy_profile = True")
            text_to_speech("Another Chrome Window is already in use. If you want to open your custom profile, please close all the chrome windows and try again.  Else make dummy_profile = True", show=False)
            browser.kill_browser()
            sys.exit()
        except Exception as ex:
            selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
            print(f"Error while browser_activate: {str(ex)}")
    except InvalidArgumentException:
        print("Another Chrome Window is already in use. If you want to open your custom profile, please close all the chrome windows and try again.  Else make dummy_profile = True")
        text_to_speech("Another Chrome Window is already in use. If you want to open your custom profile, please close all the chrome windows and try again.  Else make dummy_profile = True", show=False)
        browser.kill_browser()
        sys.exit()
    except WebDriverException:
        print("Chrome instance not found. Try again using browser_activate()")
        text_to_speech("Chrome instance not found. Try again using browser_activate()", show=False)
        sys.exit()
    except SessionNotCreatedException:
        print("Session not created. Try again using browser_activate(). If the problem persists, check the version of Chrome")
        text_to_speech("Session not created. Try again using browser_activate(). If the problem persists, check the version of Chrome", show=False)
        sys.exit()
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in browser_activate_chrome = " + str(ex))
        browser.kill_browser()
    finally:
        return status

def browser_activate_firefox(url="", dummy_browser=True, profile="Default"):
    status = False
    global browser_driver, helium_service_launched
    
    try:
        from selenium.webdriver import FirefoxOptions, FirefoxProfile, Firefox, DesiredCapabilities
        from webdriver_manager.firefox import GeckoDriverManager
        
        options = FirefoxOptions()
        if profile == "Default":
            if os_name == windows_os:
                profile_name = [i for i in os.listdir("C:/Users/{}/AppData/Roaming/Mozilla/Firefox/Profiles/".format(os.getlogin())) if "default-release" in str(i)]
            elif os_name == linux_os:
                profile_name = [i for i in os.listdir("/home/{}/.mozilla/firefox/".format(os.getlogin())) if "default-release" in str(i)]
        else:
            profile_name = profile
        if dummy_browser == False:
            profile = FirefoxProfile('C:/Users/{}/AppData/Roaming/Mozilla/Firefox/Profiles/{}'.format(os.getlogin(),profile_name[0]))
            profile.update_preferences()
            desired = DesiredCapabilities.FIREFOX
            with DisableLogger():
                browser_driver = Firefox(executable_path=GeckoDriverManager().install(),firefox_profile=profile,options=options,desired_capabilities = desired)
                status = True
        else:
            with DisableLogger():
                browser_driver = Firefox(executable_path=GeckoDriverManager().install())
                status = True
        browser.set_driver(browser_driver)
        if not url:
            browser.go_to('https://dost.clointfusion.com/')
        else:
            browser.go_to(url)

    except WebDriverException:
        print("Firefox instance not found. Try again using browser_activate_firefox()")
        text_to_speech("Firefox instance not found. Try again using browser_activate_firefox()", show=False)
        sys.exit()
    except FileNotFoundError:
        print("Installation path is not found. Please reinstall the firefox browser with default settings.")
        text_to_speech("Installation path is not found. Please reinstall the firefox browser with default settings.", show=False)
        sys.exit()
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in browser_activate_firefox = " + str(ex))
        browser.kill_browser()
    finally:
        return status

def browser_navigate_h(url=""):
    """Navigate through the url after the session is started.

    Args:
        url (str, optional): Url which you want to visit. 
        Defaults: "".

    Returns:
        bool: Whether the function is successful or failed.
    """
    status = False
    try:
        if not url:
            url = gui_get_any_input_from_user("Website URL to Navigate using Helium functions. Ex: https://www.google.com")
        global helium_service_launched
        if not helium_service_launched:
            status = browser_activate(url=url.lower())
        browser.go_to(url.lower())
        status = True
    except WebDriverException:
        print("Chrome instance not found. Please restart the Bot using browser_activate()")
        text_to_speech("Chrome instance not found. Please restart the Bot using browser_activate()", show=False)
        sys.exit()
    except NoSuchWindowException:
        print("Chrome instance not found. Please restart the Bot using browser_activate()")
        text_to_speech("Chrome instance not found. Please restart the Bot using browser_activate()", show=False)
        sys.exit()
    except AttributeError:
        print("Invalid Input. Please check the given input.")
        text_to_speech("Invalid Input. Please check the given input.", show=False)
        sys.exit()
    except TimeoutException:
        print("Element not found. Please check the given input or change browser_set_waiting_time().")
        text_to_speech("Element not found. Please check the given input or change browser_set_waiting_time().", show=False)
        sys.exit()
    except RuntimeError:
        print("Chrome instance not found. Please restart the Bot using browser_activate()")
        text_to_speech("Chrome instance not found. Please restart the Bot using browser_activate()", show=False)
        sys.exit()
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in browser_navigate_h = " + str(ex))
    finally:
        return status

def browser_write_h(Value="", User_Visible_Text_Element=""):
    """Write a string in browser, if User_Visible_Text_Element is given it writes on the given element.

    Args:
        Value (str, optional): String which has be written. 
        Defaults: "".
        User_Visible_Text_Element (str, optional): The element which is visible(Like : Sign in). 
        Defaults: "".

    Returns:
        bool: Whether the function is successful or failed.
    """
    status = False
    try:
        if not Value:
            Value = gui_get_any_input_from_user('Value to be Written')
        if Value and User_Visible_Text_Element:
            browser.write(Value, into=User_Visible_Text_Element)
            status = True
        if Value and not User_Visible_Text_Element:
            browser.write(Value)
            status = True
    except LookupError:
        print("Element not found. Please check the given input.")
        text_to_speech("Element not found. Please check the given input.", show=False)
        sys.exit()
    except WebDriverException:
        print("Chrome instance not found. Please restart the Bot using browser_activate()")
        text_to_speech("Chrome instance not found. Please restart the Bot using browser_activate()", show=False)
        sys.exit()
    except RuntimeError:
        print("Chrome instance not found. Please restart the Bot using browser_activate()")
        text_to_speech("Chrome instance not found. Please restart the Bot using browser_activate()", show=False)
        sys.exit()
    except NoSuchWindowException:
        print("Chrome instance not found. Please restart the Bot using browser_activate()")
        text_to_speech("Chrome instance not found. Please restart the Bot using browser_activate()", show=False)
        sys.exit()
    except AttributeError:
        print("Invalid Input. Please check the given input.")
        text_to_speech("Invalid Input. Please check the given input.", show=False)
        sys.exit()
    except TimeoutException:
        print("Element not found. Please check the given input or change browser_set_waiting_time().")
        text_to_speech("Element not found. Please check the given input or change browser_set_waiting_time().", show=False)
        sys.exit()
    except IndexError:
        print("Please check the input values, and try again.")
        text_to_speech("Please check the input values, and try again.", show=False)
        sys.exit()
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in browser_write_h = " + str(ex))
    finally:
        return status

def browser_mouse_click_h(User_Visible_Text_Element="", element="", double_click=False, right_click=False):
    """Click on the given element.

    Args:
        User_Visible_Text_Element (str, optional): The element which is visible(Like : Sign in). 
        Defaults: "".
        element (str, optional): Use locate_element to get element and use to click. 
        Defaults: "".
        double_click (bool, optional): True to perform a Double click. 
        Defaults: False.
        right_click (bool, optional): True to perform a Right click. 
        Defaults: False.

    Returns:
        bool: Whether the function is successful or failed.
    """
    status = False
    try:
        if not User_Visible_Text_Element and not element:
            User_Visible_Text_Element = gui_get_any_input_from_user("Give visible text element (button/link/checkbox/radio etc) to Click")
        if not double_click and not right_click:
            if User_Visible_Text_Element and not element:
                browser.click(User_Visible_Text_Element)
            if not User_Visible_Text_Element and element:
                browser.click(element)
            status = True
        if double_click and not right_click:
            if User_Visible_Text_Element and not element:
                browser.doubleclick(User_Visible_Text_Element)
            if not User_Visible_Text_Element and element:
                browser.doubleclick(element)
            status = True
        if right_click and not double_click:
            if User_Visible_Text_Element and not element:
                browser.rightclick(User_Visible_Text_Element)
            if not User_Visible_Text_Element and element:
                browser.rightclick(element)
            status = True
    except TimeoutException:
        print("Element not found. Please check the given input or change browser_set_waiting_time().")
        text_to_speech("Element not found. Please check the given input or change browser_set_waiting_time().", show=False)
        sys.exit()
    except LookupError:
        print("Element not found. Please check the given input.")
        text_to_speech("Element not found. Please check the given input.", show=False)
        sys.exit()
    except WebDriverException:
        print("Chrome instance not found. Please restart the Bot using browser_activate()")
        text_to_speech("Chrome instance not found. Please restart the Bot using browser_activate()", show=False)
        sys.exit()
    except NoSuchWindowException:
        print("Chrome instance not found. Please restart the Bot using browser_activate()")
        text_to_speech("Chrome instance not found. Please restart the Bot using browser_activate()", show=False)
        sys.exit()
    except AttributeError:
        print("Invalid Input. Please check the given input.")
        text_to_speech("Invalid Input. Please check the given input.", show=False)
        sys.exit()
    except RuntimeError:
        print("Chrome instance not found. Please restart the Bot using browser_activate()")
        text_to_speech("Chrome instance not found. Please restart the Bot using browser_activate()", show=False)
        sys.exit()
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in browser_mouse_click_h = " + str(ex))
    finally:
        return status  

def browser_locate_element_h(selector="", get_text=False, multiple_elements=False):
    """Find the element by Xpath, id or css selection.

    Args:
        selector (str, optional): Give Xpath or CSS selector. Defaults to "".
        get_text (bool, optional): Give the text of the element. Defaults to False.
        multiple_elements (bool, optional): True if you want to get all the similar elements with matching selector as list. Defaults to False.

    Returns:
        element         : If only one element
        list of elements: If multiple_elements is True
    """
    try:
        if not selector:
            selector = gui_get_any_input_from_user('Browser element to locate (Helium)')
        if not multiple_elements:
            if get_text:
                return browser.S(selector).web_element.text
            return browser.S(selector)
        if multiple_elements:
            if get_text:
                return browser.find_all(browser.S(selector).web_element.text)
            return browser.find_all(browser.S(selector))
    except TimeoutException:
        print("Element not found. Please check the given input or change browser_set_waiting_time().")
        text_to_speech("Element not found. Please check the given input or change browser_set_waiting_time().", show=False)
        sys.exit()
    except LookupError:
        print("Element not found. Please check the given input.")
        text_to_speech("Element not found. Please check the given input.", show=False)
        sys.exit()
    except WebDriverException:
        print("Chrome instance not found. Please restart the Bot using browser_activate()")
        text_to_speech("Chrome instance not found. Please restart the Bot using browser_activate()", show=False)
        sys.exit()
    except NoSuchWindowException:
        print("Chrome instance not found. Please restart the Bot using browser_activate()")
        text_to_speech("Chrome instance not found. Please restart the Bot using browser_activate()", show=False)
        sys.exit()
    except InvalidSelectorException:
        print("Invalid selector. Please check the given input.")
        text_to_speech("Invalid selector. Please check the given input.", show=False)
        sys.exit()
    except AttributeError:
        print("Invalid Input. Please check the given input.")
        text_to_speech("Invalid Input. Please check the given input.", show=False)
        sys.exit()
    except RuntimeError:
        print("Chrome instance not found. Please restart the Bot using browser_activate()")
        text_to_speech("Chrome instance not found. Please restart the Bot using browser_activate()", show=False)
        sys.exit()
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in browser_locate_element_h = " + str(ex))

def browser_wait_until_h(text="", element="t"):
    """Wait until a specific element is found.

    Args:
        text (str, optional): To wait until the string appears on the screen. 
        Eg: Export Successfull Completed. Defaults: ""
        element (str, optional): Type of Element Whether its a Text(t) or Button(b). 
        Defaults: "t - Text".

    Returns:
        bool: Whether the function is successful or failed.
    """
    status = False
    try:
        if not text:
            text = gui_get_any_input_from_user("visible text element to Search & Wait for")

        if element.lower() == "t":
            browser.wait_until(browser.Text(text).exists, 10) # text
        
        elif element.lower() == "b":
            browser.wait_until(browser.Button(text).exists, 10) # button
        status = True
    except TimeoutException:
        print("Element not found. Please check the given input or change browser_set_waiting_time().")
        text_to_speech("Element not found. Please check the given input or change browser_set_waiting_time().", show=False)
        sys.exit()
    except LookupError:
        print("Element not found. Please check the given input.")
        text_to_speech("Element not found. Please check the given input.", show=False)
        sys.exit()
    except WebDriverException:
        print("Chrome instance not found. Please restart the Bot using browser_activate()")
        text_to_speech("Chrome instance not found. Please restart the Bot using browser_activate()", show=False)
        sys.exit()
    except NoSuchWindowException:
        print("Chrome instance not found. Please restart the Bot using browser_activate()")
        text_to_speech("Chrome instance not found. Please restart the Bot using browser_activate()", show=False)
        sys.exit()
    except AttributeError:
        print("Invalid Input. Please check the given input.")
        text_to_speech("Invalid Input. Please check the given input.", show=False)
        sys.exit()
    except RuntimeError:
        print("Chrome instance not found. Please restart the Bot using browser_activate()")
        text_to_speech("Chrome instance not found. Please restart the Bot using browser_activate()", show=False)
        sys.exit()   
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in browser_wait_until_h = " + str(ex))
    finally:
        return status

def browser_refresh_page_h():
    """Refresh the current active browser page.

    Returns:
        bool: Whether the function is successful or failed.
    """
    status = False
    try:
        browser.refresh()
        status = True
    except WebDriverException:
        print("Chrome instance not found. Please restart the Bot using browser_activate()")
        text_to_speech("Chrome instance not found. Please restart the Bot using browser_activate()", show=False)
        sys.exit()
    except NoSuchWindowException:
        print("Chrome instance not found. Please restart the Bot using browser_activate()")
        text_to_speech("Chrome instance not found. Please restart the Bot using browser_activate()", show=False)
        sys.exit()
    except RuntimeError:
        print("Chrome instance not found. Please restart the Bot using browser_activate()")
        text_to_speech("Chrome instance not found. Please restart the Bot using browser_activate()", show=False)
        sys.exit()
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in browser_refresh_page_h = " + str(ex))
        sys.exit()
    finally:
        return status

def browser_hit_enter_h():
    """Hits enter KEY in Browser

    Returns:
        bool: Whether the function is successful or failed.
    """
    status = False
    try:
        browser.press(browser.ENTER)
        status = True
    except WebDriverException:
        print("Chrome instance not found. Please restart the Bot using browser_activate()")
        text_to_speech("Chrome instance not found. Please restart the Bot using browser_activate()", show=False)
        sys.exit()
    except NoSuchWindowException:
        print("Chrome instance not found. Please restart the Bot using browser_activate()")
        text_to_speech("Chrome instance not found. Please restart the Bot using browser_activate()", show=False)
        sys.exit()
    except ElementNotInteractableException:
        print("Element not found. Please check the given input.")
        text_to_speech("Element not found. Please check the given input.", show=False)
        sys.exit()
    except UnexpectedAlertPresentException:
        print("Please close the alert box. And try again.")
        text_to_speech("Please close the alert box. And try again.", show=False)
        sys.exit()
    except AttributeError:
        print("Invalid Input. Please check the given input.")
        text_to_speech("Invalid Input. Please check the given input.", show=False)
        sys.exit()
    except RuntimeError:
        print("Chrome instance not found. Please restart the Bot using browser_activate()")
        text_to_speech("Chrome instance not found. Please restart the Bot using browser_activate()", show=False)
        sys.exit()
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in browser_hit_enter_h=" + str(ex))
    finally:
        return status

def browser_key_press_h(key_1="", key_2=""):
    """Type text using Browser Helium Functions and press hot keys.

    Args:
        key_1 (str): Keys you want to simulate or string you want to press 
        Eg: "tab" or "Murali". Defaults: ""
        key_2 (str, optional): Key you want to simulate with combination to key_1. 
        Eg: "shift" or "escape". Defaults: ""
    
    Returns:
        bool: Whether the function is successful or failed.
    """
    status = False
    from helium import ENTER, SHIFT, TAB, ALT, ESCAPE, CONTROL, press
    hot_keys = ["enter", "shift", "tab", "alt", "escape", "esc", "ctrl", "control"]
    browser_keys = [ENTER, SHIFT, TAB, ALT, ESCAPE, ESCAPE, CONTROL, CONTROL]
    try:
        if not key_1:
            key_list = gui_get_any_input_from_user('keys combination using + as delimeter. Ex: ctrl or tab')
            
            if key_list.count("+") == 0:
                key_1 = key_list[0]
            elif key_list.count("+") == 1:
                key_1, key_2 = key_list.split("+")
            else:
                print("Invalid keys combination")
                return False

        if key_1 and not key_2:
            if key_1.lower() in hot_keys:
                key_1 = browser_keys[hot_keys.index(key_1.lower())]
            press(key_1)
        if key_1 and key_2:
            if key_1.lower() in hot_keys and key_2.lower() in hot_keys:
                key_1 = browser_keys[hot_keys.index(key_1.lower())]
                key_2 = browser_keys[hot_keys.index(key_2.lower())]
            if key_1.lower() in hot_keys and key_2.lower() not in hot_keys:
                key_1 = browser_keys[hot_keys.index(key_1.lower())]
            press(key_1 + key_2)
        status = True
    except TimeoutException:
        print("Element not found. Please check the given input or change browser_set_waiting_time().")
        text_to_speech("Element not found. Please check the given input or change browser_set_waiting_time().", show=False)
        sys.exit()
    except WebDriverException:
        print("Chrome instance not found. Please restart the Bot using browser_activate()")
        text_to_speech("Chrome instance not found. Please restart the Bot using browser_activate()", show=False)
        sys.exit()
    except NoSuchWindowException:
        print("Chrome instance not found. Please restart the Bot using browser_activate()")
        text_to_speech("Chrome instance not found. Please restart the Bot using browser_activate()", show=False)
        sys.exit()
    except AttributeError:
        print("Invalid Input. Please check the given input.")
        text_to_speech("Invalid Input. Please check the given input.", show=False)
        sys.exit()
    except RuntimeError:
        print("Chrome instance not found. Please restart the Bot using browser_activate()")
        text_to_speech("Chrome instance not found. Please restart the Bot using browser_activate()", show=False)
        sys.exit()
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in browser_hit_enter_h=" + str(ex))
    finally:
        return status

def browser_mouse_hover_h(User_Visible_Text_Element=""):
    """Performs a Mouse Hover over the Given User Visible Text Element

    Args:
        User_Visible_Text_Element (str, optional): The element which is visible(Like : Sign in). 
        Defaults: "".

    Returns:
        bool: Whether the function is successful or failed.
    """
    status = False
    try:
        if not User_Visible_Text_Element:
            User_Visible_Text_Element = gui_get_any_input_from_user(
                'Visible Text/element to perform mouse hover on it ')
        browser.hover(User_Visible_Text_Element)
        status = True
    except TimeoutException:
        print("Element not found. Please check the given input or change browser_set_waiting_time().")
        text_to_speech("Element not found. Please check the given input or change browser_set_waiting_time().", show=False)
        sys.exit()
    except LookupError:
        print("Element not found. Please check the given input.")
        text_to_speech("Element not found. Please check the given input.", show=False)
        sys.exit()
    except WebDriverException:
        print("Chrome instance not found. Please restart the Bot using browser_activate()")
        text_to_speech("Chrome instance not found. Please restart the Bot using browser_activate()", show=False)
        sys.exit()
    except NoSuchWindowException:
        print("Chrome instance not found. Please restart the Bot using browser_activate()")
        text_to_speech("Chrome instance not found. Please restart the Bot using browser_activate()", show=False)
        sys.exit()
    except AttributeError:
        print("Invalid Input. Please check the given input.")
        text_to_speech("Invalid Input. Please check the given input.", show=False)
        sys.exit()
    except RuntimeError:
        print("Chrome instance not found. Please restart the Bot using browser_activate()")
        text_to_speech("Chrome instance not found. Please restart the Bot using browser_activate()", show=False)
        sys.exit()
    except Exception as e:
        print('Error in browser_mouse_hover_h = ', str(e))
    finally:
        return status

def browser_set_waiting_time(time=10):
    """
    Set the waiting time for the browser. If element is not found in the given time, it will raise an exception.

    Args:
        time ([int]): The time in seconds to wait for the element to be found.
    """
    try:
        browser.Config.implicit_wait_secs = int(time)
    except WebDriverException:
        print("Chrome instance not found. Please restart the Bot using browser_activate()")
        text_to_speech("Chrome instance not found. Please restart the Bot using browser_activate()", show=False)
        sys.exit()
    except NoSuchWindowException:
        print("Chrome instance not found. Please restart the Bot using browser_activate()")
        text_to_speech("Chrome instance not found. Please restart the Bot using browser_activate()", show=False)
        sys.exit()
    except AttributeError:
        print("Invalid Input. Please check the given input.")
        text_to_speech("Invalid Input. Please check the given input.", show=False)
        sys.exit()
    except RuntimeError:
        print("Chrome instance not found. Please restart the Bot using browser_activate()")
        text_to_speech("Chrome instance not found. Please restart the Bot using browser_activate()", show=False)
        sys.exit()
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in browser_set_waiting_time = " + str(ex))

def browser_quit_h():
    """Close the Browser or Browser Automation Session.

    Returns:
        bool: Whether the function is successful or failed.
    """
    status = False
    try:
        browser.kill_browser()
        status = True
    except WebDriverException:
        print("Chrome instance not found. Please restart the Bot using browser_activate()")
        text_to_speech("Chrome instance not found. Please restart the Bot using browser_activate()", show=False)
        sys.exit()
    except NoSuchWindowException:
        print("Chrome instance not found. Please restart the Bot using browser_activate()")
        text_to_speech("Chrome instance not found. Please restart the Bot using browser_activate()", show=False)
        sys.exit()
    except RuntimeError:
        print("Chrome instance not found. Please restart the Bot using browser_activate()")
        text_to_speech("Chrome instance not found. Please restart the Bot using browser_activate()", show=False)
        sys.exit()
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in browser_quit_h = " + str(ex))
    finally:
        return status

# ---------  Browser Functions Ends --------- 


# ---------  Folder Functions ---------

def folder_read_text_file(txt_file_path=""):
    """
    Reads from a given text file and returns entire contents as a single list
    """
    try:
        if not txt_file_path:
            txt_file_path = gui_get_any_file_from_user('the text file to READ from',"txt")

        with open(txt_file_path) as f:
            file_contents = f.readlines()
        return file_contents
    except:
        return None

def folder_write_text_file(txt_file_path="",contents=""):
    """
    Writes given contents to a text file
    """
    try:
        
        if not txt_file_path:
            txt_file_path = gui_get_any_file_from_user('the text file to WRITE to',"txt")
            

        if not contents:
            contents = gui_get_any_input_from_user('text file contents')

        f = open(txt_file_path,'w',encoding="utf-8")
        f.writelines(str(contents))
        f.close()
        
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in folder_write_text_file="+str(ex))

def folder_create(strFolderPath=""):
    """
    while making leaf directory if any intermediate-level directory is missing,
    folder_create() method will create them all.

    Parameters:
        folderPath (str) : path to the folder where the folder is to be created.

    For example consider the following path:
    
    """
    try:
        if not strFolderPath:
            strFolderPath = gui_get_any_input_from_user('folder path to Create folder')

        if not os.path.exists(strFolderPath):
            os.makedirs(strFolderPath, exist_ok=True)

    except OSError:
        print("Invalid Folder path, Please check the input and try again.")
        text_to_speech("Invalid Folder path, Please check the input and try again.", show=False)
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in folder_create="+str(ex))

def folder_create_text_file(textFolderPath="",txtFileName="", custom=False):
    """
    Creates Text file in the given path.
    Internally this uses folder_create() method to create folders if the folder/s does not exist.
    automatically adds txt extension if not given in textFilePath.

    Parameters:
        textFilePath (str) : Complete path to the folder with double slashes.
    """
    try:

        if not textFolderPath:
            textFolderPath = gui_get_folder_path_from_user('the folder to create text file')
        
        if not txtFileName:
            txtFileName = gui_get_any_input_from_user("text file name")
            txtFileName = txtFileName 

        if not custom:
            if ".txt" not in txtFileName:
                txtFileName = txtFileName + ".txt"
        
            if not os.path.exists(textFolderPath):
                folder_create(textFolderPath)
        
        file_path = os.path.join(textFolderPath, txtFileName)
        file_path = Path(file_path)
        
        if os_name == windows_os:
            f = open(file_path, 'w',encoding="utf-8")
            f.close()
        if os_name == linux_os:
            if not file_path.exists():
                file_path.touch()
        return file_path
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in folder_create_text_file="+str(ex))

def folder_get_all_filenames_as_list(strFolderPath="",extension='all'):
    """
    Get all the files of the given folder in a list.

    Parameters:
        strFolderPath  (str) : Location of the folder.
        extension      (str) : extention of the file. by default all the files will be listed regarless of the extension.
    
    returns:
        allFilesOfaFolderAsLst (List) : all the file names as a list.
    """
    try:
        if not strFolderPath:
            strFolderPath = gui_get_folder_path_from_user('a folder to get all its filenames')

        if extension == "all":
            allFilesOfaFolderAsLst = [ f for f in os.listdir(strFolderPath)]
        else:
            allFilesOfaFolderAsLst = [ f for f in os.listdir(strFolderPath) if f.endswith(extension) ]

        return allFilesOfaFolderAsLst
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in folder_get_all_filenames_as_list="+str(ex))

def folder_delete_all_files(fullPathOfTheFolder="",file_extension_without_dot="all"):  
    """
    Deletes all the files of the given folder

    Parameters:
        fullPathOfTheFolder  (str) : Location of the folder.
        extension            (str) : extention of the file. by default all the files will be deleted inside the given folder 
                                    regarless of the extension.
    returns:
        count (int) : number of files deleted.
    """ 
    file_extension_with_dot = ''
    try:
        if not fullPathOfTheFolder:
            fullPathOfTheFolder = gui_get_folder_path_from_user('a folder to delete all its files')

        count = 0 
        if "." not in file_extension_without_dot :
            file_extension_with_dot = "." + file_extension_without_dot

        if file_extension_with_dot.lower() == ".all":
            filelist = [ f for f in os.listdir(fullPathOfTheFolder) ]
        else:
            filelist = [ f for f in os.listdir(fullPathOfTheFolder) if f.endswith(file_extension_with_dot) ]
        print(filelist)
        for f in filelist:
            try:
                file_path = os.path.join(fullPathOfTheFolder, f)
                file_path = Path(file_path)
                os.remove(file_path)
                count +=1 
            except Exception as ex:
                selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
                print("Error in folder_delete_all_files = " + str(ex))
        
        return count
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in folder_delete_all_files="+str(ex)) 
        return -1

def file_rename(old_file_path='',new_file_name='',ext=False):
    '''
    Renames the given file name to new file name with same extension
    '''
    try:
        if not old_file_path:
            old_file_path = gui_get_any_input_from_user('Pass the complete file path to be renamed')
        
        if not new_file_name:
            if ext:
                new_file_name = gui_get_any_input_from_user('New file name without extension')
            else:
                new_file_name = gui_get_any_input_from_user('New file name with extension') 

        if os.path.exists(old_file_path):
            if new_file_name:
                if ext:
                    path_of_new_file = os.path.join('\\'.join(old_file_path.split('\\')[:-1]), new_file_name)
                
                else:
                    ext = old_file_path.split('\\')[-1].split('.')[-1]
                    path_of_new_file = os.path.join('\\'.join(old_file_path.split('\\')[:-1]) , '.'.join([new_file_name,ext]))
            
                os.rename(src=Path(old_file_path),dst=Path(path_of_new_file))
                print(path_of_new_file)
            else:
                raise Exception('new_file_name can\'t be empty.')
        else:
            raise Exception('Old_file_path is invalid. Please pass a valid path.')
     
    except Exception as e:
        print('Error in file_rename = ',str(e))

def file_get_json_details(path_of_json_file='',section=''):
    '''
    Returns all the details of the given section in a dictionary 
    '''
    try:
        if not path_of_json_file:
            path_of_json_file = gui_get_any_input_from_user('Pass the complete path of JSON file')
        
        if not section:
            section = gui_get_any_input_from_user('Pass the section to get all the details in it')
        
        # import json

        with open(path_of_json_file,'r') as fp:
            data = json.load(fp)
        fp.close()

        if section in list(data.keys()):
            return data.get(section)
        else:
            raise Exception('Section can\'t be find in given json file.')

    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        print(f'Error in file_get_json_details = {ex}')

# ---------  Folder Functions Ends ---------


# ---------  Window Operations Functions --------- 

def window_show_desktop():
    """
    Minimizes all the applications and shows Desktop.
    """
    try:
        time.sleep(0.5)
        pg.hotkey("win","d")
        time.sleep(0.5)
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in window_show_desktop="+str(ex))

def window_get_active_window():
    """
    Returns the title of the active window.
    """
    try:
        return win32gui.GetWindowText(win32gui.GetForegroundWindow())
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in window_get_active_window="+str(ex))

def window_activate_window(window_title=''):
    """
    Activates the given window.
    """
    try:
        if not window_title:
            open_win_list = window_get_all_opened_titles_windows()
            window_title = gui_get_dropdownlist_values_from_user("window titles to Activate & Maximize",dropdown_list=open_win_list,multi_select=False)[0]
        
        item,window_found = _window_find_exact_name(window_title)
        if window_found:
            windw = gw.getWindowsWithTitle(item)[0]

            try:
                windw.activate()
            except:
                windw.minimize()
                windw.maximize()
            time.sleep(1)
            
        else:
            print("No window OPEN by name="+str(window_title))
    except AttributeError:
        print("Invalid Input. Please check the given input.")
        text_to_speech("Invalid Input. Please check the given input.", show=False)
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in window_activate_window="+str(ex))

def window_get_all_opened_titles_windows():
    """
    Gives the title of all the existing (open) windows.

    Returns:
        allTitles_lst  (list) : returns all the titles of the window as list.
    """
    try:
        allTitles_lst = []
        lst = gw.getAllTitles()
        for item in lst:
            if str(item).strip() != "" and str(item).strip() not in allTitles_lst:
                allTitles_lst.append(str(item).strip())
        return allTitles_lst
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in window_get_all_opened_titles="+str(ex))

def window_activate_and_maximize_windows(windowName=""):
    """
    Activates and maximizes the desired window.

    Parameters:
        windowName  (str) : Name of the window to maximize.
    """
    try:
        if not windowName:
            open_win_list = window_get_all_opened_titles_windows()
            windowName = gui_get_dropdownlist_values_from_user("window titles to Activate & Maximize",dropdown_list=open_win_list,multi_select=False)[0]
       
        item,window_found = _window_find_exact_name(windowName)
        if window_found:
            windw = gw.getWindowsWithTitle(item)[0]

            try:
                windw.activate()
                windw.maximize()
            except:
                windw.minimize()
                time.sleep(1)
                windw.maximize()
            time.sleep(1)
            
        else:
            print("No window OPEN by name="+str(windowName))
    except gw.PyGetWindowException:
        print("Invalid Input. Please check the given input.")
        text_to_speech("Invalid Input. Please check the given input.", show=False)
    except AttributeError:
        print("Invalid Input. Please check the given input.")
        text_to_speech("Invalid Input. Please check the given input.", show=False)
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in window_activate_and_maximize="+str(ex))

def window_minimize_windows(windowName=""):
    """
    Activates and minimizes the desired window.

    Parameters:
        windowName  (str) : Name of the window to miniimize.
    """
    try:
        if not windowName:
            open_win_list = window_get_all_opened_titles_windows()
            windowName = gui_get_dropdownlist_values_from_user("window titles to Minimize",dropdown_list=open_win_list,multi_select=False)[0]
            
        item,window_found = _window_find_exact_name(windowName)
        if window_found:
            windw = gw.getWindowsWithTitle(item)[0]
            windw.minimize()
            time.sleep(1)
        else:
            print("No window available to minimize by name="+str(windowName))
    except AttributeError:
        print("Invalid Input. Please check the given input.")
        text_to_speech("Invalid Input. Please check the given input.", show=False)
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in window_minimize="+str(ex))

def window_close_windows(windowName=""):
    """
    Close the desired window.

    Parameters:
        windowName  (str) : Name of the window to close.
    """
    try:
        if not windowName:
            open_win_list = window_get_all_opened_titles_windows()
            windowName = gui_get_dropdownlist_values_from_user("window titles to Close",dropdown_list=open_win_list,multi_select=False)[0]
            
        item,window_found = _window_find_exact_name(windowName)
        if window_found:
            windw = gw.getWindowsWithTitle(item)[0]
            windw.close()
            time.sleep(1)
        else:
            print("No window available to close, by name="+str(windowName))
    except AttributeError:
        print("Invalid Input. Please check the given input.")
        text_to_speech("Invalid Input. Please check the given input.", show=False)
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in window_close="+str(ex))

def launch_any_exe_bat_application(pathOfExeFile=""):
    """Launches any exe or batch file or excel file etc.

    Args:
        pathOfExeFile (str, optional): Location of the file with extension 
        Eg: Notepad, TextEdit. Defaults to "".
    """
    status = False
    try:
        if not pathOfExeFile:
            pathOfExeFile = gui_get_any_file_from_user('EXE or BAT file')

        if os_name == windows_os:
            import win32gui, win32con
            try:
                subprocess.Popen(pathOfExeFile)
                time.sleep(2)
                hwnd = win32gui.GetForegroundWindow()
                win32gui.ShowWindow(hwnd, win32con.SW_MAXIMIZE)
                status = True
            except Exception as ex1:
                try:
                    os.startfile(pathOfExeFile)
                except Exception as ex2:
                    print("launch_any_exe_bat_application"+str(ex2))
            
        elif os_name == linux_os:
            try:
                subprocess.Popen(pathOfExeFile)
                time.sleep(2)
                status = True
            except Exception as ex:
                selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
                print("launch_any_exe_bat_application"+str(ex))
        
        elif os_name == mac_os:
            try:
                subprocess.Popen(f'open -a "{pathOfExeFile}"')
                status = True
            except Exception as ex:
                selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
                print("launch_any_exe_bat_application"+str(ex))
                
        time.sleep(1) 
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in launch_any_exe_bat_application="+str(ex))
    finally:
        return status

# ---------  Window Operations Functions Ends --------- 


# ---------  String Functions --------- 

def string_extract_only_alphabets(inputString=""):
    """
    Returns only alphabets from given input string
    """
    if not inputString:
        inputString = gui_get_any_input_from_user("input string to get only Alphabets")

    outputStr = ''.join(e for e in inputString if e.isalpha())
    return outputStr 

def string_extract_only_numbers(inputString=""):
    """
    Returns only numbers from given input string
    """
    if not inputString:
        inputString = gui_get_any_input_from_user("input string to get only Numbers")

    outputStr = ''.join(e for e in inputString if e.isnumeric())
    return outputStr       

def string_remove_special_characters(inputStr=""):
    """
    Removes all the special character.

    Parameters:
        inputStr  (str) : string for removing all the special character in it.

    Returns :
        outputStr (str) : returns the alphanumeric string.
    """

    if not inputStr:
        inputStr = gui_get_any_input_from_user('input string to remove Special characters')

    if inputStr:
        outputStr = ''.join(e for e in inputStr if e.isalnum())
        return outputStr  

def string_regex(inputStr="",strExpAfter="",strExpBefore="",intIndex=0):
    """
    Regex API service call, to search within a given string data
    """
    regex_url = "https://api.clointfusion.com/str_regex"
    try:
        resp = requests.post(regex_url, data={'str_input':str(inputStr),'before':strExpBefore, 'after':strExpAfter,'indx':intIndex})
        return resp.text
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in string_regex = " + str(ex))

# ---------  String Functions Ends --------- 
 

# ---------  Excel Functions --------- 

def excel_get_row_column_count(excel_path="", sheet_name="Sheet1", header=0):
    """
    Gets the row and coloumn count of the provided excel sheet.

    Parameters:
        excel_path  (str) : Full path to the excel file with slashes.
        sheet_name           (str) : by default it is Sheet1.

    Returns:
        row (int) : number of rows
        col (int) : number of coloumns
    """
    try:
        if not excel_path:
            excel_path, sheet_name, header = gui_get_excel_sheet_header_from_user("to get row/column count")
            
        df = pd.read_excel(excel_path, sheet_name=sheet_name, header=header,engine='openpyxl')
        row, col = df.shape
        row = row + 1
        return row, col
    except FileNotFoundError:
        print("File not found. Please check the given path.")
        text_to_speech("File not found. Please check the given path.", show=False)
        sys.exit()
    except ValueError:
        print("Input is not supported, could not convert string to float. Please check the inputs, and try again.")
        text_to_speech("Input is not supported, could not convert string to float. Please check the inputs, and try again.", show=False)
    except PermissionError:
        print("Please close the excel file, and try again.")
        text_to_speech("Please close the excel file, and try again.", show=False)
    except op.utils.exceptions.InvalidFileException:
        print("We currently support only : xlsx,.xlsm,.xltx,.xltm files. Please try again with one of those file formats.")
        text_to_speech("we currently support only : .xlsx,.xlsm,.xltx,.xltm files. Please try again with one of those file formats.", show=False)
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in excel_get_row_column_count="+str(ex))

def excel_copy_range_from_sheet(excel_path="", sheet_name='Sheet1', startCol=0, startRow=0, endCol=0, endRow=0): #*
    """
    Copies the specific range from the provided excel sheet and returns copied data as a list
    Parameters:
        excel_path :"Full path of the excel file with double slashes"
        sheet_name     :"Source sheet name from where contents are to be copied"
        startCol          :"Starting column number (index starts from 1) from where copying starts"
        startRow          :"Starting row number (index starts from 1) from where copying starts"
        endCol            :"Ending column number ex:4 upto where cells to be copied"
        endRow            :"Ending column number ex:5 upto where cells to be copied"

    Returns:
    rangeSelected        : the copied range data
    """
    try:
        if not excel_path:
            excel_path, sheet_name, _ = gui_get_excel_sheet_header_from_user('to copy range from')
            
        if startCol == 0 and startRow ==0 and endCol == 0 and endRow == 0:
            sRow_sCol_eRow_Col = gui_get_any_input_from_user('startRow , startCol, endRow, endCol (comma separated, index from 1)')    

            if sRow_sCol_eRow_Col:
                startRow , startCol, endRow, endCol = str(sRow_sCol_eRow_Col).split(",")
                startRow = int(startRow)
                startCol = int(startCol)
                endRow = int(endRow)
                endCol = int(endCol)

        from_wb = load_workbook(filename = excel_path)
        try:
            fromSheet = from_wb[sheet_name]
        except:
            fromSheet = from_wb.worksheets[0]
        rangeSelected = []

        if endRow < startRow:
            endRow = startRow

        #Loops through selected Rows
        for i in range(startRow,endRow + 1,1):
            #Appends the row to a RowSelected list
            rowSelected = []
            for j in range(startCol,endCol+1,1):
                rowSelected.append(fromSheet.cell(row = i, column = j).value)
            #Adds the RowSelected List and nests inside the rangeSelected
            rangeSelected.append(rowSelected)
    
        return rangeSelected
    except ValueError:
        print("Input is not supported, could not convert string to float. Please check the inputs, and try again.")
        text_to_speech("Input is not supported, could not convert string to float. Please check the inputs, and try again.", show=False)
    except PermissionError:
        print("Please close the excel file, and try again.")
        text_to_speech("Please close the excel file, and try again.", show=False)
    except op.utils.exceptions.InvalidFileException:
        print("We currently support only : xlsx,.xlsm,.xltx,.xltm files. Please try again with one of those file formats.")
        text_to_speech("we currently support only : .xlsx,.xlsm,.xltx,.xltm files. Please try again with one of those file formats.", show=False)
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in copy_range_from_excel_sheet="+str(ex))

def excel_copy_paste_range_from_to_sheet(excel_path="", sheet_name='Sheet1', startCol=0, startRow=0, endCol=0, endRow=0, copiedData=""):#*
    """
    Pastes the copied data in specific range of the given excel sheet.
    """
    try:
        try:
            if not copiedData:
                copiedData = excel_copy_range_from_sheet()

            if not excel_path:
                excel_path, sheet_name, _ = gui_get_excel_sheet_header_from_user('to paste range into')
                
            if startCol == 0 and startRow ==0 and endCol == 0 and endRow == 0:
                sRow_sCol_eRow_Col = gui_get_any_input_from_user('startRow , startCol, endRow, endCol (comma separated, index from 1)')    

                if sRow_sCol_eRow_Col:
                    startRow , startCol, endRow, endCol = str(sRow_sCol_eRow_Col).split(",")
                    startRow = int(startRow)
                    startCol = int(startCol)
                    endRow = int(endRow)
                    endCol = int(endCol)

            to_wb = load_workbook(filename = excel_path)
            toSheet = to_wb[sheet_name]

        except:
            try:
                excel_create_excel_file_in_given_folder((str(excel_path[:(str(excel_path).rindex("\\"))])),(str(excel_path[str(excel_path).rindex("\\")+1:excel_path.find(".")])),sheet_name)
            except:
                excel_create_excel_file_in_given_folder((str(excel_path[:(str(excel_path).rindex("/"))])),(str(excel_path[str(excel_path).rindex("/")+1:excel_path.find(".")])),sheet_name)

            to_wb = load_workbook(filename = excel_path)

            toSheet = to_wb[sheet_name]

        if endRow < startRow:
            endRow = startRow

        countRow = 0
        for i in range(startRow,endRow+1,1):
            countCol = 0
            for j in range(startCol,endCol+1,1):
                toSheet.cell(row = i, column = j).value = copiedData[countRow][countCol]
                countCol += 1
            countRow += 1
        to_wb.save(excel_path)
        return countRow-1
    except OSError:
        print("Please close the excel file, and try again.")
        text_to_speech("Please close the excel file, and try again.", show=False)
        sys.exit()
    except IndexError:
        print("Input is not supported, could not convert string to float. Please check the inputs, and try again.")
        text_to_speech("Input is not supported, could not convert string to float. Please check the inputs, and try again.", show=False)
        sys.exit()
    except TypeError:
        print("Input is not supported, could not convert string to float. Please check the inputs, and try again.")
        text_to_speech("Input is not supported, could not convert string to float. Please check the inputs, and try again.", show=False)
        sys.exit()
    except ValueError:
        print("Input is not supported, could not convert string to float. Please check the inputs, and try again.")
        text_to_speech("Input is not supported, could not convert string to float. Please check the inputs, and try again.", show=False)
    except PermissionError:
        print("Please close the excel file, and try again.")
        text_to_speech("Please close the excel file, and try again.", show=False)
    except op.utils.exceptions.InvalidFileException:
        print("We currently support only : xlsx,.xlsm,.xltx,.xltm files. Please try again with one of those file formats.")
        text_to_speech("we currently support only : .xlsx,.xlsm,.xltx,.xltm files. Please try again with one of those file formats.", show=False)
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in excel_copy_paste_range_from_to_sheet="+str(ex))

def excel_split_by_column(excel_path="",sheet_name='Sheet1',header=0,columnName=""):#*
    """
    Splits the excel file by Column Name
    """
    try:
        if not excel_path:
            excel_path, sheet_name, header = gui_get_excel_sheet_header_from_user('to split by column')
            
        if not columnName:
            col_lst = excel_get_all_header_columns(excel_path, sheet_name, header)
            columnName = gui_get_dropdownlist_values_from_user('this list of Columns (to split)',col_lst)

        data_df = pd.read_excel(excel_path,sheet_name=sheet_name,header=header,dtype=str,engine='openpyxl')
        
        grouped_df = data_df.groupby(columnName)
        
        for data in  grouped_df:  
            file_path = os.path.join(output_folder_path,str(data[0]) + ".xlsx")
            file_path = Path(file_path)
            grouped_df.get_group(data[0]).to_excel(file_path, index=False)

        # message_toast("Excel splitting done", file_folder_path=file_path)
            
    except ValueError:
        print("Input is not supported, could not convert string to float. Please check the inputs, and try again.")
        text_to_speech("Input is not supported, could not convert string to float. Please check the inputs, and try again.", show=False)
    except PermissionError:
        print("Please close the excel file, and try again.")
        text_to_speech("Please close the excel file, and try again.", show=False)
    except op.utils.exceptions.InvalidFileException:
        print("We currently support only : xlsx,.xlsm,.xltx,.xltm files. Please try again with one of those file formats.")
        text_to_speech("we currently support only : .xlsx,.xlsm,.xltx,.xltm files. Please try again with one of those file formats.", show=False)
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in excel_split_by_column="+str(ex))

def excel_split_the_file_on_row_count(excel_path="", sheet_name = 'Sheet1', rowSplitLimit="", outputFolderPath="", outputTemplateFileName ="Split"):#*
    """
    Splits the excel file as per given row limit
    """
    try:
        if not excel_path:
            excel_path, sheet_name, _ = gui_get_excel_sheet_header_from_user('to split on row count')
            
        if not rowSplitLimit:
            rowSplitLimit = int(gui_get_any_input_from_user("row split Count/Limit Ex: 20"))

        if not outputFolderPath:
            outputFolderPath = gui_get_folder_path_from_user('output folder to Save split excel files')

        src_wb = op.load_workbook(excel_path)
        src_ws = src_wb.worksheets[0] 

        src_ws_max_rows = src_ws.max_row
        src_ws_max_cols= src_ws.max_column 

        i = 1
        start_row = 2

        while start_row <= src_ws_max_rows:
            
            dest_wb = Workbook()
            dest_ws = dest_wb.active
            dest_ws.title = sheet_name

            #Copy ROW-1 (Header) from SOURCE to Each DESTINATION file
            selectedRange = _excel_copy_range(1,1,src_ws_max_cols,1,src_ws) #startCol, startRow, endCol, endRow, sheet
            _ =_excel_paste_range(1,1,src_ws_max_cols,1,dest_ws,selectedRange) #startCol, startRow, endCol, endRow, sheetReceiving,copiedData
            
            selectedRange = ""
            selectedRange = _excel_copy_range(1,start_row,src_ws_max_cols,start_row + rowSplitLimit - 1,src_ws) #startCol, startRow, endCol, endRow, sheet   
            _ =_excel_paste_range(1,2,src_ws_max_cols,rowSplitLimit + 1,dest_ws,selectedRange) #startCol, startRow, endCol, endRow, sheetReceiving,copiedData

            start_row = start_row + rowSplitLimit

            try:
                dest_file_name = str(outputFolderPath) + "\\" + outputTemplateFileName + "-" + str(i) + ".xlsx"
            except:
                dest_file_name = str(outputFolderPath) + "/" + outputTemplateFileName + "-" + str(i) + ".xlsx"

            dest_file_name = Path(dest_file_name)
            dest_wb.save(dest_file_name)
            
            i = i + 1
        return True
    except ValueError:
        print("Input is not supported, could not convert string to float. Please check the inputs, and try again.")
        text_to_speech("Input is not supported, could not convert string to float. Please check the inputs, and try again.", show=False)
    except PermissionError:
        print("Please close the excel file, and try again.")
        text_to_speech("Please close the excel file, and try again.", show=False)
    except op.utils.exceptions.InvalidFileException:
        print("We currently support only : xlsx,.xlsm,.xltx,.xltm files. Please try again with one of those file formats.")
        text_to_speech("we currently support only : .xlsx,.xlsm,.xltx,.xltm files. Please try again with one of those file formats.", show=False)
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in excel_split_the_file_on_row_count="+str(ex))

def excel_merge_all_files(input_folder_path="",output_folder_path=""):
    """
    Merges all the excel files in the given folder
    """
    try:
        if not input_folder_path:
            input_folder_path = gui_get_folder_path_from_user('input folder to MERGE files from')

        if not output_folder_path:
            output_folder_path = gui_get_folder_path_from_user('output folder to store Final merged file')
        
        filelist = [ f for f in os.listdir(input_folder_path) if f.endswith(".xlsx") ]
        all_excel_file_lst = []
        for file1 in filelist:
            file_path = os.path.join(input_folder_path,file1)
            file_path = Path(file_path)
            
            all_excel_file = pd.read_excel(file_path,dtype=str,engine='openpyxl')
            all_excel_file_lst.append(all_excel_file)

        appended_df = pd.concat(all_excel_file_lst)
        time_stamp_now=datetime.datetime.now().strftime("%m-%d-%Y")
        final_path = os.path.join(output_folder_path, "Final-" + time_stamp_now + ".xlsx")
        final_path= Path(final_path)
        appended_df.to_excel(final_path, index=False)

        # message_toast("Excel merging completed", file_folder_path=final_path)
        
        return True
    except ValueError:
        print("Input is not supported, could not convert string to float. Please check the inputs, and try again.")
        text_to_speech("Input is not supported, could not convert string to float. Please check the inputs, and try again.", show=False)
    except PermissionError:
        print("Please close the excel file, and try again.")
        text_to_speech("Please close the excel file, and try again.", show=False)
    except op.utils.exceptions.InvalidFileException:
        print("We currently support only : xlsx,.xlsm,.xltx,.xltm files. Please try again with one of those file formats.")
        text_to_speech("we currently support only : .xlsx,.xlsm,.xltx,.xltm files. Please try again with one of those file formats.", show=False)
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in excel_merge_all_files="+str(ex))

def excel_drop_columns(excel_path="", sheet_name='Sheet1', header=0, columnsToBeDropped = ""):
    """
    Drops the desired column from the given excel file
    """
    from matplotlib.pyplot import axis
    
    try:
        if not excel_path:
            excel_path, sheet_name, header = gui_get_excel_sheet_header_from_user('input excel to Drop the columns from')

        if not columnsToBeDropped:
            col_lst = excel_get_all_header_columns(excel_path, sheet_name, header)
            columnsToBeDropped = gui_get_dropdownlist_values_from_user('columns list to drop',col_lst) 

        df=pd.read_excel(excel_path,sheet_name=sheet_name, header=header,engine='openpyxl') 

        if isinstance(columnsToBeDropped, list):
            df.drop(columnsToBeDropped, axis = 1, inplace = True) 
        else:
            df.drop([columnsToBeDropped], axis = 1, inplace = True) 


            
        with pd.ExcelWriter(excel_path) as writer: # pylint: disable=abstract-class-instantiated
            df.to_excel(writer, sheet_name=sheet_name,index=False) 

    except ValueError:
        print("Input is not supported, could not convert string to float. Please check the inputs, and try again.")
        text_to_speech("Input is not supported, could not convert string to float. Please check the inputs, and try again.", show=False)
    except PermissionError:
        print("Please close the excel file, and try again.")
        text_to_speech("Please close the excel file, and try again.", show=False)
    except op.utils.exceptions.InvalidFileException:
        print("We currently support only : xlsx,.xlsm,.xltx,.xltm files. Please try again with one of those file formats.")
        text_to_speech("we currently support only : .xlsx,.xlsm,.xltx,.xltm files. Please try again with one of those file formats.", show=False)
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in excel_drop_columns="+str(ex))

def excel_sort_columns(excel_path="",sheet_name='Sheet1',header=0,firstColumnToBeSorted=None,secondColumnToBeSorted=None,thirdColumnToBeSorted=None,firstColumnSortType=True,secondColumnSortType=True,thirdColumnSortType=True, view_excel=False):#*
    """
    A function which takes excel full path to excel and column names on which sort is to be performed

    """
    try:
        if not excel_path:
            excel_path, sheet_name, header = gui_get_excel_sheet_header_from_user('to sort the column')

        if not firstColumnToBeSorted:
            col_lst = excel_get_all_header_columns(excel_path, sheet_name, header)
            usecols = gui_get_dropdownlist_values_from_user('minimum 1 and maximum 3 columns to sort',col_lst)
            
            if len(usecols) == 3:
                firstColumnToBeSorted , secondColumnToBeSorted , thirdColumnToBeSorted = usecols
            elif len(usecols) == 2:
                firstColumnToBeSorted , secondColumnToBeSorted = usecols
            elif len(usecols) == 1:
                firstColumnToBeSorted = usecols[0]
        df=pd.read_excel(excel_path,sheet_name=sheet_name, header=header,engine='openpyxl')

        if view_excel:
            show(df)

        if thirdColumnToBeSorted is not None and secondColumnToBeSorted is not None and firstColumnToBeSorted is not None:
            df=df.sort_values([firstColumnToBeSorted,secondColumnToBeSorted,thirdColumnToBeSorted],ascending=[firstColumnSortType,secondColumnSortType,thirdColumnSortType])
        
        elif secondColumnToBeSorted is not None and firstColumnToBeSorted is not None:
            df=df.sort_values([firstColumnToBeSorted,secondColumnToBeSorted],ascending=[firstColumnSortType,secondColumnSortType])
        
        elif firstColumnToBeSorted is not None:
            df=df.sort_values([firstColumnToBeSorted],ascending=[firstColumnSortType])

        book = load_workbook(excel_path)
        writer = pd.ExcelWriter(excel_path, engine='openpyxl') # pylint: disable=abstract-class-instantiated
        
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
    
        df.to_excel(writer,sheet_name=sheet_name,index=False)

        writer.save()
        
        return True
    except ValueError:
        print("Input is not supported, could not convert string to float. Please check the inputs, and try again.")
        text_to_speech("Input is not supported, could not convert string to float. Please check the inputs, and try again.", show=False)
    except PermissionError:
        print("Please close the excel file, and try again.")
        text_to_speech("Please close the excel file, and try again.", show=False)
    except op.utils.exceptions.InvalidFileException:
        print("We currently support only : xlsx,.xlsm,.xltx,.xltm files. Please try again with one of those file formats.")
        text_to_speech("we currently support only : .xlsx,.xlsm,.xltx,.xltm files. Please try again with one of those file formats.", show=False)
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in excel_sort_columns="+str(ex))        

def excel_clear_sheet(excel_path="",sheet_name="Sheet1", header=0):
    """
    Clears the contents of given excel files keeping header row intact
    """
    try:
        
        if not excel_path:
            excel_path, sheet_name, header = gui_get_excel_sheet_header_from_user('to clear the sheet')

        df = pd.read_excel(excel_path,sheet_name=sheet_name,header=header,engine='openpyxl') 
        df = df.head(0)

        with pd.ExcelWriter(excel_path) as writer: # pylint: disable=abstract-class-instantiated
            df.to_excel(writer,sheet_name=sheet_name, index=False)

    except ValueError:
        print("Input is not supported, could not convert string to float. Please check the inputs, and try again.")
        text_to_speech("Input is not supported, could not convert string to float. Please check the inputs, and try again.", show=False)
    except PermissionError:
        print("Please close the excel file, and try again.")
        text_to_speech("Please close the excel file, and try again.", show=False)
    except op.utils.exceptions.InvalidFileException:
        print("We currently support only : xlsx,.xlsm,.xltx,.xltm files. Please try again with one of those file formats.")
        text_to_speech("we currently support only : .xlsx,.xlsm,.xltx,.xltm files. Please try again with one of those file formats.", show=False)
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in excel_clear_sheet="+str(ex))

def excel_set_single_cell(excel_path="", sheet_name="Sheet1", header=0, columnName="", cellNumber=0, setText=""): #*
    """
    Writes the given text to the desired column/cell number for the given excel file
    """
    try:
        if not excel_path:
            excel_path, sheet_name, header = gui_get_excel_sheet_header_from_user('to set cell')
            
        if not columnName:
            col_lst = excel_get_all_header_columns(excel_path, sheet_name, header)
            columnName = gui_get_dropdownlist_values_from_user('list of columns to set vlaue',col_lst,multi_select=False)   

        if not setText:
            setText = gui_get_any_input_from_user("text value to set the cell")

        df = pd.read_excel(excel_path,sheet_name=sheet_name,header=header,engine='openpyxl')
        
        df.at[cellNumber,columnName] = setText
        append_df_to_excel(excel_path, df, index=False,startrow=0)
        
        return True

    except ValueError:
        print("Input is not supported, could not convert string to float. Please check the inputs, and try again.")
        text_to_speech("Input is not supported, could not convert string to float. Please check the inputs, and try again.", show=False)
    except PermissionError:
        print("Please close the excel file, and try again.")
        text_to_speech("Please close the excel file, and try again.", show=False)
    except op.utils.exceptions.InvalidFileException:
        print("We currently support only : xlsx,.xlsm,.xltx,.xltm files. Please try again with one of those file formats.")
        text_to_speech("we currently support only : .xlsx,.xlsm,.xltx,.xltm files. Please try again with one of those file formats.", show=False)
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in excel_set_single_cell="+str(ex))

def excel_get_single_cell(excel_path="",sheet_name="Sheet1",header=0, columnName="",cellNumber=0): #*
    """
    Gets the text from the desired column/cell number of the given excel file
    """
    try:
        if not excel_path:
            excel_path, sheet_name, header = gui_get_excel_sheet_header_from_user('to get cell')
            
        if not columnName:
            col_lst = excel_get_all_header_columns(excel_path, sheet_name, header)
            columnName = gui_get_dropdownlist_values_from_user('list of columns to get vlaue',col_lst,multi_select=False)   

        if not isinstance(columnName, list):
            columnName = [columnName]       

        df = pd.read_excel(excel_path,sheet_name=sheet_name,header=header,usecols={columnName[0]},engine='openpyxl')
        cellValue = df.at[cellNumber,columnName[0]]
        return cellValue
    except ValueError:
        print("Input is not supported, could not convert string to float. Please check the inputs, and try again.")
        text_to_speech("Input is not supported, could not convert string to float. Please check the inputs, and try again.", show=False)
    except PermissionError:
        print("Please close the excel file, and try again.")
        text_to_speech("Please close the excel file, and try again.", show=False)
    except op.utils.exceptions.InvalidFileException:
        print("We currently support only : xlsx,.xlsm,.xltx,.xltm files. Please try again with one of those file formats.")
        text_to_speech("we currently support only : .xlsx,.xlsm,.xltx,.xltm files. Please try again with one of those file formats.", show=False)
    except KeyError:
        print("Column name not found in the excel file. Please check the column name and try again.")
        text_to_speech("Column name not found in the excel file. Please check the column name and try again.", show=False)
        sys.exit()
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in excel_get_single_cell="+str(ex))

def excel_remove_duplicates(excel_path="",sheet_name="Sheet1", header=0, columnName="", saveResultsInSameExcel=True, which_one_to_keep="first"): #*
    """
    Drops the duplicates from the desired Column of the given excel file
    """
    try:
        if not excel_path:
            excel_path, sheet_name, header = gui_get_excel_sheet_header_from_user('to remove duplicates')
            
        if not columnName:
            col_lst = excel_get_all_header_columns(excel_path, sheet_name, header)
            columnName = gui_get_dropdownlist_values_from_user('list of columns to remove duplicates',col_lst)  
    
        df = pd.read_excel(excel_path, sheet_name=sheet_name,header=header,engine='openpyxl') 

        count = 0 
        if saveResultsInSameExcel:
            df.drop_duplicates(subset=columnName, keep=which_one_to_keep, inplace=True)
            with pd.ExcelWriter(excel_path) as writer: # pylint: disable=abstract-class-instantiated
                df.to_excel(writer,sheet_name=sheet_name,index=False)

            count = df.shape[0]
        else:
            df1 = df.drop_duplicates(subset=columnName, keep=which_one_to_keep, inplace=False)
            excel_path = str(excel_path).replace(".","_DupDropped.")
            with pd.ExcelWriter(excel_path) as writer: # pylint: disable=abstract-class-instantiated
                df1.to_excel(writer,sheet_name=sheet_name,index=False)
            count = df1.shape[0]

        return count
    except ValueError:
        print("Input is not supported, could not convert string to float. Please check the inputs, and try again.")
        text_to_speech("Input is not supported, could not convert string to float. Please check the inputs, and try again.", show=False)
    except PermissionError:
        print("Please close the excel file, and try again.")
        text_to_speech("Please close the excel file, and try again.", show=False)
    except op.utils.exceptions.InvalidFileException:
        print("We currently support only : xlsx,.xlsm,.xltx,.xltm files. Please try again with one of those file formats.")
        text_to_speech("we currently support only : .xlsx,.xlsm,.xltx,.xltm files. Please try again with one of those file formats.", show=False)
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in excel_remove_duplicates="+str(ex))

def excel_vlook_up(filepath_1="", sheet_name_1 = 'Sheet1', header_1 = 0, filepath_2="", sheet_name_2 = 'Sheet1', header_2 = 0, Output_path="", OutputExcelFileName="", match_column_name="",how='left', view_excel=False):#*
    """
    Performs excel_vlook_up on the given excel files for the desired columns. Possible values for how are "inner","left", "right", "outer"
    """
    try:
        if not filepath_1:
            filepath_1, sheet_name_1, header_1 = gui_get_excel_sheet_header_from_user('(Vlookup) first excel')
             
        if not filepath_2:
            filepath_2, sheet_name_2, header_2 = gui_get_excel_sheet_header_from_user('(Vlookup) second excel')
            
        if not match_column_name:
            col_lst = excel_get_all_header_columns(filepath_1, sheet_name_1, header_1)
            match_column_name = gui_get_dropdownlist_values_from_user('Vlookup column name to be matched',col_lst,multi_select=False) 
            match_column_name = match_column_name[0]
        df1 = pd.read_excel(filepath_1, sheet_name = sheet_name_1, header = header_1,engine='openpyxl')
        df2 = pd.read_excel(filepath_2, sheet_name = sheet_name_2, header = header_2,engine='openpyxl')

        df = pd.merge(df1, df2, on= match_column_name, how = how)

        if view_excel:
            show(df)

        output_file_path = ""
        if str(OutputExcelFileName).endswith(".*"):
            OutputExcelFileName = OutputExcelFileName.split(".")[0]
        
        if Output_path and OutputExcelFileName:
            if ".xlsx" in OutputExcelFileName:
                output_file_path = os.path.join(Output_path, OutputExcelFileName)
            else:
                output_file_path = os.path.join(Output_path, OutputExcelFileName  + ".xlsx")

        else:
            output_file_path = filepath_1

        output_file_path = Path(output_file_path)
        with pd.ExcelWriter(output_file_path) as writer: # pylint: disable=abstract-class-instantiated
            df.to_excel(writer, index=False)
            
        return True
    
    except ValueError:
        print("Input is not supported, could not convert string to float. Please check the inputs, and try again.")
        text_to_speech("Input is not supported, could not convert string to float. Please check the inputs, and try again.", show=False)
    except PermissionError:
        print("Please close the excel file, and try again.")
        text_to_speech("Please close the excel file, and try again.", show=False)
    except op.utils.exceptions.InvalidFileException:
        print("We currently support only : xlsx,.xlsm,.xltx,.xltm files. Please try again with one of those file formats.")
        text_to_speech("we currently support only : .xlsx,.xlsm,.xltx,.xltm files. Please try again with one of those file formats.", show=False)
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in excel_vlook_up="+str(ex))

def excel_change_corrupt_xls_to_xlsx(xls_file ='',xlsx_file = '', xls_sheet_name=''): 
    '''
        Repair corrupt file to regular file and then convert it to xlsx.
        status : Done.
    '''
    try :
        # Used to save the file as excel workbook
        # Need to install this library
        from xlwt import Workbook
        import io
        from xls2xlsx import XLS2XLSX
        # Opening the file 
        file1 = io.open(xls_file, "r")
        data = file1.readlines()

        # Creating a workbook object
        xldoc = Workbook()
        # Adding a sheet to the workbook object
        sheet = xldoc.add_sheet(xls_sheet_name, cell_overwrite_ok=True)
        # Iterating and saving the data to sheet
        for i, row in enumerate(data):
            # Two things are done here
            # Removing the '\n' which comes while reading the file using io.open
            # Getting the values after splitting using '\t'
            for j, val in enumerate(row.replace('\n', '').split('\t')):
                sheet.write(i, j, val)

        # Saving the file as a normal xls excel file
        xldoc.save(xls_file)

        # checking the downloaded file is present or not 
        if os.path.exists(xls_file):
            # converting xls to xlsx
            x2x = XLS2XLSX(xls_file)
            x2x.to_xlsx(xlsx_file)
        return True   
    except ValueError:
        print("Input is not supported, could not convert string to float. Please check the inputs, and try again.")
        text_to_speech("Input is not supported, could not convert string to float. Please check the inputs, and try again.", show=False)
    except PermissionError:
        print("Please close the excel file, and try again.")
        text_to_speech("Please close the excel file, and try again.", show=False)
    except op.utils.exceptions.InvalidFileException:
        print("We currently support only : xlsx,.xlsm,.xltx,.xltm files. Please try again with one of those file formats.")
        text_to_speech("we currently support only : .xlsx,.xlsm,.xltx,.xltm files. Please try again with one of those file formats.", show=False)
    except Exception as e:
        exception_msg = f"Error in excel_change_corrupt_xls_to_xlsx : {str(e)}"
        return exception_msg

def excel_convert_xls_to_xlsx(xls_file_path='',xlsx_file_path=''):
    """
    Converts given XLS file to XLSX
    """
    try:
        # Checking the path and then converting it to xlsx file
        from xls2xlsx import XLS2XLSX
        if os.path.exists(xls_file_path):
            # converting xls to xlsx
            x2x = XLS2XLSX(xls_file_path)
            x2x.to_xlsx(xlsx_file_path)
        return True
    except ValueError:
        print("Input is not supported, could not convert string to float. Please check the inputs, and try again.")
        text_to_speech("Input is not supported, could not convert string to float. Please check the inputs, and try again.", show=False)
    except PermissionError:
        print("Please close the excel file, and try again.")
        text_to_speech("Please close the excel file, and try again.", show=False)
    except op.utils.exceptions.InvalidFileException:
        print("We currently support only : xlsx,.xlsm,.xltx,.xltm files. Please try again with one of those file formats.")
        text_to_speech("we currently support only : .xlsx,.xlsm,.xltx,.xltm files. Please try again with one of those file formats.", show=False)
    except Exception as e:
        errMsg = f"Error in converting file to xlsx format : {str(e)}"
        return errMsg

def excel_apply_format_as_table(excel_file_path,table_style="TableStyleMedium21",sheet_name='Sheet1'): # range : "A1:AA"
    '''
        Applies table format to the used range of the given excel.
        Just it takes an path and converts it to table here you can change the table style below.
        if you want to change the table style just change the styles by refering excel
    '''
    import win32com.client 
    excel_instance = win32com.client.gencache.EnsureDispatch("Excel.Application")
    excel_instance.Visible = False
    excel_instance.DisplayAlerts = False
    
    exc_workbook = excel_instance.Workbooks.Open(Filename=excel_file_path.replace("/", "\\")) # .Sheets.Item[sheet_name]
    try :
        exc_workbook.Worksheets(sheet_name).Select()
        excel_instance.ActiveSheet.UsedRange.Select()
        excel_instance.Selection.Columns.AutoFit()
        excel_instance.ActiveSheet.ListObjects.Add().TableStyle = table_style
        exc_workbook.Close(SaveChanges=1)
        excel_instance.Quit()
    except ValueError:
        print("Input is not supported, could not convert string to float. Please check the inputs, and try again.")
        text_to_speech("Input is not supported, could not convert string to float. Please check the inputs, and try again.", show=False)
    except PermissionError:
        print("Please close the excel file, and try again.")
        text_to_speech("Please close the excel file, and try again.", show=False)
    except op.utils.exceptions.InvalidFileException:
        print("We currently support only : xlsx,.xlsm,.xltx,.xltm files. Please try again with one of those file formats.")
        text_to_speech("we currently support only : .xlsx,.xlsm,.xltx,.xltm files. Please try again with one of those file formats.", show=False)
    except:
        exc_workbook.Close()
        excel_instance.Quit()
        raise Exception("Given Excel already has a table")

def excel_split_on_user_defined_conditions(excel_file_path,sheet_name='Sheet1',column_name='',condition_strings=None,output_dir='',view_excel=False):
    '''
        Splits the excel based on user defined row/column conditions
        Just give the column name and row condition which you want split your  excel.
        Give one string or if more conditions the  pass as list it will split the excel based on those conditions and save  them 
        in the given output directory.
        Here if output dir is not there it will create output dir in current directory and save all excels there. 
        If you want unique rows data in different excel files simply don't pass any thing in condition strings
    '''
    try:
        if not os.path.exists(output_dir):
            folder_create(output_dir)
        df = pd.read_excel(excel_file_path,sheet_name=sheet_name)

        if view_excel:
            show(df)

        if condition_strings == None:
            
            condition_strings = df[column_name].unique()
            for condition_str in condition_strings:
                df_new = df.loc[df[column_name] == condition_str]
                excel_newfile_path = output_dir + "\\" + column_name + '-'+condition_str   + '.xlsx'
                df_new.to_excel(excel_newfile_path, index=False)
        else:  
            if type(condition_strings) == str:
                condition_strings = [condition_strings]
            for condition_str in condition_strings:
                df_new = df.loc[df[column_name] == condition_str]
                excel_newfile_path = output_dir + "\\" + column_name + '-'+condition_str   + '.xlsx'
                df_new.to_excel(excel_newfile_path, index=False)
    except ValueError:
        print("Input is not supported, could not convert string to float. Please check the inputs, and try again.")
        text_to_speech("Input is not supported, could not convert string to float. Please check the inputs, and try again.", show=False)
    except PermissionError:
        print("Please close the excel file, and try again.")
        text_to_speech("Please close the excel file, and try again.", show=False)
    except op.utils.exceptions.InvalidFileException:
        print("We currently support only : xlsx,.xlsm,.xltx,.xltm files. Please try again with one of those file formats.")
        text_to_speech("we currently support only : .xlsx,.xlsm,.xltx,.xltm files. Please try again with one of those file formats.", show=False)
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        errMsg = f"Error in excel_split_on_user_defined_conditions: {str(ex)}"
        return errMsg

def excel_convert_to_image(excel_file_path=""):
    """
    Returns an Image (PNG) of given Excel
    """
    try:
        if os_name == windows_os:
            from PIL import ImageGrab

            import win32com.client 
            excel = win32com.client.gencache.EnsureDispatch('Excel.Application')

            image_path = str(excel_file_path).replace(".xlsx",".PNG")
            wb = excel.Workbooks.Open(str(excel_file_path))
            
            wb.Windows(1).Visible = False
            ws = wb.Worksheets(1)
                        
            df_row_cnt = pd.read_excel(excel_file_path,engine="openpyxl")
            row_cnt, _ = df_row_cnt.shape
            df_row_cnt = ''
            
            win32c = win32com.client.constants
            ws.Range("A1:E{}".format(row_cnt+1)).CopyPicture(Format=win32c.xlBitmap)

            img = ImageGrab.grabclipboard()
            image_path = ''.join(e for e in image_path if e.isalnum())
            image_path = img_folder_path / "Excel_Image_{}.PNG".format(image_path)
            img.save(image_path)

            wb.Close(True)
            excel.Quit()
            del excel
            del excel_file_path

            return image_path
        else:
            print("This feature is available only on Windows OS")
    except ValueError:
        print("Input is not supported, could not convert string to float. Please check the inputs, and try again.")
        text_to_speech("Input is not supported, could not convert string to float. Please check the inputs, and try again.", show=False)
    except PermissionError:
        print("Please close the excel file, and try again.")
        text_to_speech("Please close the excel file, and try again.", show=False)
    except op.utils.exceptions.InvalidFileException:
        print("We currently support only : xlsx,.xlsm,.xltx,.xltm files. Please try again with one of those file formats.")
        text_to_speech("we currently support only : .xlsx,.xlsm,.xltx,.xltm files. Please try again with one of those file formats.", show=False)
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in excel_convert_to_image="+str(ex))

def excel_create_excel_file_in_given_folder(fullPathToTheFolder="",excelFileName="",sheet_name="Sheet1"):
    """
    Creates an excel file in the desired folder with desired filename

    Internally this uses folder_create() method to create folders if the folder/s does not exist.

    Parameters:
        fullPathToTheFolder (str) : Complete path to the folder with double slashes.
        excelFileName       (str) : File Name of the excel to be created (.xlsx extension will be added automatically.
        sheet_name           (str) : By default it will be "Sheet1".
    
    Returns:
        returns boolean TRUE if the excel file is created
    """
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title =sheet_name

        if not fullPathToTheFolder:
            fullPathToTheFolder = gui_get_folder_path_from_user('the folder to create excel file')
            
        if not excelFileName:
            excelFileName = gui_get_any_input_from_user("excel file name (without extension)")
        
        folder_create(fullPathToTheFolder)

        if ".xlsx" in excelFileName:
            excel_path = os.path.join(fullPathToTheFolder,excelFileName)
        else:
            excel_path = os.path.join(fullPathToTheFolder,excelFileName + ".xlsx")
            
        excel_path = Path(excel_path)

        wb.save(filename = excel_path)
        
        return True
    except ValueError:
        print("Input is not supported, could not convert string to float. Please check the inputs, and try again.")
        text_to_speech("Input is not supported, could not convert string to float. Please check the inputs, and try again.", show=False)
    except PermissionError:
        print("Please close the excel file, and try again.")
        text_to_speech("Please close the excel file, and try again.", show=False)
    except op.utils.exceptions.InvalidFileException:
        print("We currently support only : xlsx,.xlsm,.xltx,.xltm files. Please try again with one of those file formats.")
        text_to_speech("we currently support only : .xlsx,.xlsm,.xltx,.xltm files. Please try again with one of those file formats.", show=False)
    except TypeError:
        print("Invalid Argument or input. Please check the inputs, and try again.")
        text_to_speech("Invalid Argument or input. Please check the inputs, and try again.", show=False)
    except OSError:
        print("Invalid Argument or input. Please check the inputs, and try again.")
        text_to_speech("Invalid Argument or input. Please check the inputs, and try again.", show=False)
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in excel_create_excel_file_in_given_folder="+str(ex))

def excel_if_value_exists(excel_path="",sheet_name='Sheet1',header=0,usecols="",value=""):
    """
    Check if a given value exists in given excel. Returns True / False
    """
    try:
        if not excel_path:
            excel_path, sheet_name, header = gui_get_excel_sheet_header_from_user('to search the VALUE')

        if not value:
            value = gui_get_any_input_from_user('VALUE to be searched')
        
        if usecols:
            df = pd.read_excel(excel_path, sheet_name=sheet_name, header=header, usecols=usecols,engine='openpyxl')
        else:
            df = pd.read_excel(excel_path, sheet_name=sheet_name, header=header,engine='openpyxl')
        
        if value in df.values:
            df = ''
            return True
        else:
            df = ''
            return False

    except ValueError:
        print("Input is not supported, could not convert string to float. Please check the inputs, and try again.")
        text_to_speech("Input is not supported, could not convert string to float. Please check the inputs, and try again.", show=False)
    except PermissionError:
        print("Please close the excel file, and try again.")
        text_to_speech("Please close the excel file, and try again.", show=False)
    except op.utils.exceptions.InvalidFileException:
        print("We currently support only : xlsx,.xlsm,.xltx,.xltm files. Please try again with one of those file formats.")
        text_to_speech("we currently support only : .xlsx,.xlsm,.xltx,.xltm files. Please try again with one of those file formats.", show=False)
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in excel_if_value_exists="+str(ex))

def excel_create_file(fullPathToTheFile="",fileName="",sheet_name="Sheet1"):
    """
        Create a Excel file in fullPathToTheFile with filename.
    """
    try:
        if not fullPathToTheFile:
            fullPathToTheFile = gui_get_any_input_from_user('folder path to create excel')

        if not fileName:
            fileName = gui_get_any_input_from_user("Excel File Name (without extension)")

        if not os.path.exists(fullPathToTheFile):
            os.makedirs(fullPathToTheFile)
        if ".xlsx" not in fileName:
            fileName = fileName + ".xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title =sheet_name

        fileName = os.path.join(fullPathToTheFile,fileName)
        
        fileName = Path(fileName)

        wb.save(filename = fileName)
        
        return True
    except ValueError:
        print("Input is not supported, could not convert string to float. Please check the inputs, and try again.")
        text_to_speech("Input is not supported, could not convert string to float. Please check the inputs, and try again.", show=False)
    except PermissionError:
        print("Please close the excel file, and try again.")
        text_to_speech("Please close the excel file, and try again.", show=False)
    except op.utils.exceptions.InvalidFileException:
        print("We currently support only : xlsx,.xlsm,.xltx,.xltm files. Please try again with one of those file formats.")
        text_to_speech("we currently support only : .xlsx,.xlsm,.xltx,.xltm files. Please try again with one of those file formats.", show=False)
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in excel_create_file="+str(ex))

def excel_to_colored_html(formatted_excel_path=""):
    """
    Converts given Excel to HTML preserving the Excel format and saves in same folder as .html
    """
    try:
        from xlsx2html import xlsx2html

        if not formatted_excel_path:
            formatted_excel_path = gui_get_any_file_from_user('Excel file to convert to HTML','xlsx')        

        formatted_html_path = str(formatted_excel_path).replace(".xlsx",".html")
        xlsx2html(formatted_excel_path, formatted_html_path)
        return formatted_html_path
    except ValueError:
        print("Input is not supported, could not convert string to float. Please check the inputs, and try again.")
        text_to_speech("Input is not supported, could not convert string to float. Please check the inputs, and try again.", show=False)
    except PermissionError:
        print("Please close the excel file, and try again.")
        text_to_speech("Please close the excel file, and try again.", show=False)
    except op.utils.exceptions.InvalidFileException:
        print("We currently support only : xlsx,.xlsm,.xltx,.xltm files. Please try again with one of those file formats.")
        text_to_speech("we currently support only : .xlsx,.xlsm,.xltx,.xltm files. Please try again with one of those file formats.", show=False)
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in excel_to_colored_html="+str(ex))

def excel_get_all_sheet_names(excelFilePath=""):
    """
    Gives you all names of the sheets in the given excel sheet.

    Parameters:
        excelFilePath  (str) : Full path to the excel file with slashes.
    
    returns : 
        all the names of the excelsheets as a LIST.
    """
    try:
        if not excelFilePath:
            excelFilePath = gui_get_any_file_from_user("xlsx")

        wb = load_workbook(excelFilePath)
        return wb.sheetnames
    except ValueError:
        print("Input is not supported, could not convert string to float. Please check the inputs, and try again.")
        text_to_speech("Input is not supported, could not convert string to float. Please check the inputs, and try again.", show=False)
    except PermissionError:
        print("Please close the excel file, and try again.")
        text_to_speech("Please close the excel file, and try again.", show=False)
    except op.utils.exceptions.InvalidFileException:
        print("We currently support only : xlsx,.xlsm,.xltx,.xltm files. Please try again with one of those file formats.")
        text_to_speech("we currently support only : .xlsx,.xlsm,.xltx,.xltm files. Please try again with one of those file formats.", show=False)
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in excel_get_all_sheet_names="+str(ex))

def excel_get_all_header_columns(excel_path="",sheet_name="Sheet1",header=0):
    """
    Gives you all column header names of the given excel sheet.
    """
    col_lst = []
    try:
        if not excel_path:
            excel_path,sheet_name,header = gui_get_excel_sheet_header_from_user('to all header columns as a list')

        col_lst = pd.read_excel(excel_path,sheet_name=sheet_name,header=header,nrows=1,dtype=str,engine='openpyxl').columns.tolist()
        return col_lst
    except ValueError:
        print("Input is not supported, could not convert string to float. Please check the inputs, and try again.")
        text_to_speech("Input is not supported, could not convert string to float. Please check the inputs, and try again.", show=False)
    except PermissionError:
        print("Please close the excel file, and try again.")
        text_to_speech("Please close the excel file, and try again.", show=False)
    except op.utils.exceptions.InvalidFileException:
        print("We currently support only : xlsx,.xlsm,.xltx,.xltm files. Please try again with one of those file formats.")
        text_to_speech("we currently support only : .xlsx,.xlsm,.xltx,.xltm files. Please try again with one of those file formats.", show=False)
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in excel_get_all_header_columns="+str(ex))

def excel_describe_data(excel_path="",sheet_name='Sheet1',header=0,view_excel=False):
    """
    Describe statistical data for the given excel
    """
    try:
        if not excel_path:
            excel_path, sheet_name, header = gui_get_excel_sheet_header_from_user("to Statistically Describe excel data")
            
        df = pd.read_excel(excel_path, sheet_name=sheet_name, header=header,engine='openpyxl')

        if view_excel:
            show(df)
        #user_option_lst = ['Numerical','String','Both']

        #user_choice = gui_get_dropdownlist_values_from_user("list of datatypes",user_option_lst)

        #if user_choice == 'Numerical':
        #    return df.describe(include = [np.number])
        #elif user_choice == 'String':
        #    return df.describe(include = ['O'])
        #else:
        #    return df.describe(include='all')
        return df.describe()

    except ValueError:
        print("Input is not supported, could not convert string to float. Please check the inputs, and try again.")
        text_to_speech("Input is not supported, could not convert string to float. Please check the inputs, and try again.", show=False)
    except PermissionError:
        print("Please close the excel file, and try again.")
        text_to_speech("Please close the excel file, and try again.", show=False)
    except op.utils.exceptions.InvalidFileException:
        print("We currently support only : xlsx,.xlsm,.xltx,.xltm files. Please try again with one of those file formats.")
        text_to_speech("we currently support only : .xlsx,.xlsm,.xltx,.xltm files. Please try again with one of those file formats.", show=False)
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in excel_describe_data="+str(ex))

def excel_sub_routines():
    """
    Excel VBA Macros called from ClointFusion
    """
    try:
        if os_name == windows_os:
            import xlwings as xw
            cf_excel_rountine_file_path = os.path.join(clointfusion_directory,"Misc","CF_Excel_Routines.xlsb")

            import win32com.client
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            
            user_choice = gui_get_dropdownlist_values_from_user("list of Macros", ['Greet Me','Get Row/Col Count','List all Sheets','Save Worksheets as PDF','Send Outlook Email'],multi_select=False)[0]
            
            if user_choice == 'Greet Me':
                excel.Workbooks.Open(Filename=cf_excel_rountine_file_path, ReadOnly=1)
                user_data = gui_get_any_input_from_user('your name')
                excel.Application.Run("'CF_Excel_Routines.xlsb'!btnHello_World_Click", user_data)

            else:
                user_excel_path, _, _ = gui_get_excel_sheet_header_from_user('to {}'.format(user_choice))    

                user_excel_path_with_sr = str(user_excel_path).replace(".xlsx",".xlsb")

                try:
                    os.remove(user_excel_path_with_sr)
                except Exception as ex:
                    selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
                    print("Error in excel_sub_routines = " + str(ex))

                shutil.copy2(cf_excel_rountine_file_path,user_excel_path_with_sr)

                
                wb1 = xw.Book(user_excel_path)
                wb2 = xw.Book(user_excel_path_with_sr)

                ws1 = wb1.sheets(1)

                ws1.api.Copy(Before=wb2.sheets(1).api)
                try:
                    wb2.save(user_excel_path_with_sr)
                except Exception as ex:
                    selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
                    print("Error in excel_sub_routines = " + str(ex))

                wb1.close()
                wb2.close()

                try:
                    wb1.app.quit()
                    wb2.app.quit()
                except Exception as ex:
                    selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
                    print("Error in excel_sub_routines = " + str(ex))

                excel.Workbooks.Open(Filename=user_excel_path_with_sr, ReadOnly=1)
                file_name = str(Path(user_excel_path_with_sr).stem) + ".xlsb"

                if user_choice == 'Get Row/Col Count':
                    excel.Application.Run("'{}'!getRowColCount".format(file_name))

                elif user_choice == "List all Sheets":
                    excel.Application.Run("'{}'!getAllSheetsAsList".format(file_name))

                elif user_choice == "Save Worksheets as PDF":
                    excel.Application.Run("'{}'!SaveWorksheetAsPDF".format(file_name),str(output_folder_path))

                elif user_choice == "Send Outlook Email":
                    
                    toAddress = gui_get_any_input_from_user("To Email Address")
                    subject = gui_get_any_input_from_user("Email Subject")
                    EmailBody = gui_get_any_input_from_user("Email Body", multi_line=True)

                    excel.Application.Run("'{}'!Send_Mail_Outlook".format(file_name),toAddress,subject,EmailBody)

            excel.Workbooks.Close()
            excel.Application.Quit() 
            del excel

            try:
                ew = gw.getWindowsWithTitle('Excel')[0]
                ew.close()
            except Exception as ex:
                selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
                print("Error in excel_sub_routines = " + str(ex))
        else:
            print("This feature is available only on Windows OS")

    except ValueError:
        print("Input is not supported, could not convert string to float. Please check the inputs, and try again.")
        text_to_speech("Input is not supported, could not convert string to float. Please check the inputs, and try again.", show=False)
    except PermissionError:
        print("Please close the excel file, and try again.")
        text_to_speech("Please close the excel file, and try again.", show=False)
    except op.utils.exceptions.InvalidFileException:
        print("We currently support only : xlsx,.xlsm,.xltx,.xltm files. Please try again with one of those file formats.")
        text_to_speech("we currently support only : .xlsx,.xlsm,.xltx,.xltm files. Please try again with one of those file formats.", show=False)
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in excel_sub_routines="+str(ex))

def convert_csv_to_excel(csv_path="",sep=""):
    """
    Function to convert CSV to Excel 

    Ex: convert_csv_to_excel()
    """
    try:
        if not csv_path:
            csv_path = gui_get_any_file_from_user("CSV to convert to EXCEL","csv")

        if not sep:
            sep = gui_get_any_input_from_user("Delimeter Ex: |")

        csv_file_name = _extract_filename_from_filepath(csv_path)
        excel_file_name = csv_file_name + ".xlsx"        

        excel_file_path = os.path.join(output_folder_path,excel_file_name)
        excel_file_path = Path(excel_file_path)
        writer = pd.ExcelWriter(excel_file_path) # pylint: disable=abstract-class-instantiated

        df=pd.read_csv(csv_path,sep=sep)
        df.to_excel(writer, sheet_name='Sheet1', index=False)

        writer.save()
        
        print("Excel file saved : "+str(excel_file_path))

        # message_toast("CSV to excel conversion done", file_folder_path=excel_file_path)

    except ValueError:
        print("Input is not supported, could not convert string to float. Please check the inputs, and try again.")
        text_to_speech("Input is not supported, could not convert string to float. Please check the inputs, and try again.", show=False)
    except PermissionError:
        print("Please close the excel file, and try again.")
        text_to_speech("Please close the excel file, and try again.", show=False)
    except op.utils.exceptions.InvalidFileException:
        print("We currently support only : xlsx,.xlsm,.xltx,.xltm files. Please try again with one of those file formats.")
        text_to_speech("we currently support only : .xlsx,.xlsm,.xltx,.xltm files. Please try again with one of those file formats.", show=False)
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in convert_csv_to_excel="+str(ex))

def isNaN(value):
    """
    Returns TRUE if a given value is NaN False otherwise
    """
    try:
        import math
        return math.isnan(float(value))
    except ValueError:
        print("Input is not supported, could not convert string to float. Please check the inputs, and try again.")
        text_to_speech("Input is not supported, could not convert string to float. Please check the inputs, and try again.", show=False)
    except PermissionError:
        print("Please close the excel file, and try again.")
        text_to_speech("Please close the excel file, and try again.", show=False)
    except op.utils.exceptions.InvalidFileException:
        print("We currently support only : xlsx,.xlsm,.xltx,.xltm files. Please try again with one of those file formats.")
        text_to_speech("we currently support only : .xlsx,.xlsm,.xltx,.xltm files. Please try again with one of those file formats.", show=False)
    except:
        return False

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None, startcol=None,
    truncate_sheet=False, resizeColumns=True, na_rep = 'NA', **to_excel_kwargs):
    
    try:
        from string import ascii_uppercase
        from openpyxl.utils import get_column_letter
        
        # ignore [engine] parameter if it was passed
        if 'engine' in to_excel_kwargs:
            to_excel_kwargs.pop('engine')

        try:
            f = open(filename)
            # Do something with the file
        except IOError:
            # print("File not accessible")
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            wb.save(filename)

        writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

        try:
            # try to open an existing workbook
            writer.book = load_workbook(filename)

            # get the last row in the existing Excel sheet
            # if it was not specified explicitly
            if startrow is None and sheet_name in writer.book.sheetnames:
                startrow = writer.book[sheet_name].max_row

            # truncate sheet
            if truncate_sheet and sheet_name in writer.book.sheetnames:
                # index of [sheet_name] sheet
                idx = writer.book.sheetnames.index(sheet_name)
                # remove [sheet_name]
                writer.book.remove(writer.book.worksheets[idx])
                # create an empty sheet [sheet_name] using old index
                writer.book.create_sheet(sheet_name, idx)

            # copy existing sheets
            writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
        except FileNotFoundError:
            # file does not exist yet, we will create it
            pass

        if startrow is None:
            # startrow = -1
            startrow = 0

        if startcol is None:
            startcol = 0

        # write out the new sheet
        df.to_excel(writer, sheet_name, startrow=startrow, startcol=startcol, na_rep=na_rep, **to_excel_kwargs)

        if resizeColumns:

            ws = writer.book[sheet_name]

            def auto_format_cell_width(ws):
                for letter in range(1,ws.max_column):
                    maximum_value = 0
                    for cell in ws[get_column_letter(letter)]:
                        val_to_check = len(str(cell.value))
                        if val_to_check > maximum_value:
                            maximum_value = val_to_check
                    ws.column_dimensions[get_column_letter(letter)].width = maximum_value + 2

            auto_format_cell_width(ws)

        # save the workbook
        writer.save()
    except ValueError:
        print("Input is not supported, could not convert string to float. Please check the inputs, and try again.")
        text_to_speech("Input is not supported, could not convert string to float. Please check the inputs, and try again.", show=False)
    except PermissionError:
        print("Please close the excel file, and try again.")
        text_to_speech("Please close the excel file, and try again.", show=False)
    except op.utils.exceptions.InvalidFileException:
        print("We currently support only : xlsx,.xlsm,.xltx,.xltm files. Please try again with one of those file formats.")
        text_to_speech("we currently support only : .xlsx,.xlsm,.xltx,.xltm files. Please try again with one of those file formats.", show=False)
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in append_df_to_excel="+str(ex))

# ---------  Excel Functions Ends --------- 


# --------- Windows Objects Functions ---------
@lru_cache(None)
def win_obj_open_app(title,program_path_with_name,file_path_with_name="",backend='uia'):  
    from pywinauto import Desktop, Application

    """
    Open any windows application
    Parameters : 
        Title - Title of the application window.
        program_path_with_name - The full path of the application
        file_path_with_name - The full path to the file (only if required ex: to open an already saved excel file)
    """
    if os_name == windows_os:
        try:  
            if file_path_with_name:
                app = Application(backend=backend).start(r'{} "{}"'.format(program_path_with_name, file_path_with_name))
            else:
                app = Application(backend=backend).start(program_path_with_name)
                
            if title.lower() == "calculator":
                main_dlg = Desktop(backend=backend).Calculator
            else:
                main_dlg = app.window(title_re='.*?' + title + '.*?')
            time.sleep(1)
            return app, main_dlg
        except Exception as ex:
            selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
            print("Exception in win_obj_open_app : " + str(ex))
    else:
        print("Works only on windows OS")

def win_obj_get_all_objects(main_dlg,save=False,file_name_with_path=""):
    from pywinauto import Desktop, Application

    """
    Print or Save all the windows object elements of an application.
    Parameters : 
        main_dlg - Main Dialogue Handle returned from OpenApp_w() function.
        save - True if you want to save.
        file_name_with_path - new txt file name with path if you want to save)
    """
    if os_name == windows_os:
        try:
            if save and file_name_with_path:
                main_dlg.print_control_identifiers(filename=file_name_with_path)
                print("File Saved...")
            else:
                main_dlg.print_control_identifiers()
        except Exception as ex:
            selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
            print("Exception in win_obj_get_all_objects : " + str(ex))
    else:
        print("Works only on windows OS")

def win_obj_mouse_click(main_dlg,title="", auto_id="", control_type=""):
    from pywinauto import Desktop, Application

    """
    Simulate high level mouse clicks on windows object elements.
    Parameters : 
        main_dlg - Main Dialogue Handle returned from OpenApp_w() function.
        title - Title of the application window.
        auto_id - Automation ID of the windows object element.
        control_type - Control type of the windows object element.
    """
    if os_name == windows_os:
        try:
            main_dlg.set_focus()
            if title:
                main_dlg.child_window(title=title).invoke()
            elif auto_id and control_type:
                main_dlg.child_window(auto_id=auto_id, control_type='Button').invoke()
            elif auto_id:
                main_dlg.child_window(auto_id=auto_id).invoke()
            else:
                print("Need \'title\' or \'auto_id\' Parameter for Mouse Click to work")
                exit()
        except Exception as ex:
            selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
            print("Exception in win_obj_mouse_click : " + str(ex))
    else:
        print("Works only on windows OS")

def win_obj_key_press(main_dlg,write,title="", auto_id="", control_type=""):
    from pywinauto import Desktop, Application

    """
    Simulate high level Keypress on windows object elements.
    Parameters : 
        main_dlg - Main Dialogue Handle returned from OpenApp_w() function.
        write - text to write.
        title - Title of the application window.
        auto_id - Automation ID of the windows object element.
        control_type - Control type of the windows object element.
    """
    if os_name == windows_os:
        try:
            main_dlg.set_focus()
            if title:
                main_dlg.child_window(title=title).type_keys(write, with_spaces=True)
            elif auto_id and control_type:
                main_dlg.child_window(auto_id=auto_id, control_type='Text').type_keys(write, with_spaces=True)
            elif auto_id:
                main_dlg.child_window(auto_id=auto_id).type_keys(write, with_spaces=True)
            else:
                main_dlg.type_keys(write, with_spaces=True)
        except Exception as ex:
            selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
            print("Exception in win_obj_key_press : " + str(ex))
    else:
        print("Works only on windows OS")

def win_obj_get_text(main_dlg,title="", auto_id="", control_type="", value = False):
    from pywinauto import Desktop, Application

    """
    Read text from windows object element.
    Parameters : 
        main_dlg - Main Dialogue Handle returned from OpenApp_w() function.
        title - Title of the application window.
        auto_id - Automation ID of the windows object element.
        control_type - Control type of the windows object element.
        Value - True to read  a set of text and false to read another set of text for the same windows object element.
    """
    if os_name == windows_os:
        try:
            main_dlg.set_focus()
            if title:
                if value:
                    read = main_dlg.child_window(title=title)
                    read = read.legacy_properties()['Value']
                else:
                    read = main_dlg.child_window(title=title).window_text()
                return read
            elif auto_id and control_type:
                if value:
                    read = main_dlg.child_window(auto_id=auto_id, control_type='Text')
                    read = read.legacy_properties()['Value']
                else:
                    read = main_dlg.child_window(auto_id=auto_id, control_type='Text').window_text()
                return read
            elif auto_id:
                if value:
                    read = main_dlg.child_window(auto_id=auto_id)
                    read = read.legacy_properties()['Value']
                else:
                    read = main_dlg.child_window(auto_id=auto_id).window_text()
                return read
            else:
                if value:
                    read = main_dlg.legacy_properties()['Value']
                else:
                    read = main_dlg.window_text()
                return read
        except Exception as ex:
            selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
            print("Exception in win_obj_get_text : " + str(ex))
    else:
        print("Works only on windows OS")

# --------- Windows Objects Functions Ends ---------


# --------- Screenscraping Functions ---------

def scrape_save_contents_to_notepad(folderPathToSaveTheNotepad="",switch_to_window="",X=0,Y=0): 
    #"Full path to the folder (with double slashes) where notepad is to be stored"
    """
    Copy pastes all the available text on the screen to notepad and saves it.
    """
    try:
        if not folderPathToSaveTheNotepad:
            folderPathToSaveTheNotepad = gui_get_folder_path_from_user('folder to save notepad contents')

        message_counter_down_timer("Screen scraping in (seconds)",3)
        
        if switch_to_window:
            if os_name == windows_os:
                window_activate_and_maximize_windows(switch_to_window)
            elif os_name == linux_os:
                mouse_click(100,100)
        
        time.sleep(1)
        if X == 0 and Y == 0:
            X = pg.size()[0]/2
            Y = pg.size()[1]/2
        pg.click(X,Y)
        time.sleep(0.5)

        pg.hotkey("ctrl", "a")
        time.sleep(1)
        
        pg.hotkey("ctrl", "c")
        time.sleep(1)
        
        clipboard_data = clipboard.paste()
        time.sleep(2)
        
        screen_clear_search()

        notepad_file_path = Path(folderPathToSaveTheNotepad)    
        notepad_file_path = notepad_file_path / 'notepad-contents.txt'

        f = open(notepad_file_path, "w", encoding="utf-8")
        f.write(clipboard_data)
        time.sleep(10)
        f.close()

        clipboard_data = ''
        return "Saved the contents at " + str(notepad_file_path)
    except OSError:
        print('Please check the folder path or close the notepad file')
        text_to_speech('Please check the folder path or close the notepad file', show=False)
        sys.exit()
    except FileNotFoundError:
        print("Please check the filepath, and try again.")
        text_to_speech("Please check the filepath, and try again.", show=False)
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in scrape_save_contents_to_notepad = "+str(ex))
    
def scrape_get_contents_by_search_copy_paste(highlightText=""):
    """
    Gets the focus on the screen by searching given text using crtl+f and performs copy/paste of all data. Useful in Citrix applications
    This is useful in Citrix applications
    """
    output_lst_newline_removed = []
    try:
        if not highlightText:
            highlightText = gui_get_any_input_from_user("text to be searched in Citrix environment")

        time.sleep(1)
        
        pg.hotkey("ctrl", "f")
        time.sleep(1)
        pg.typewrite(highlightText)
        time.sleep(1)
        
        pg.hotkey("enter")
        time.sleep(1)
        
        pg.hotkey("esc")
        time.sleep(2)

        pg.PAUSE = 2
        
        pg.hotkey("ctrl", "a")
        time.sleep(2)
        
        pg.hotkey("ctrl", "c")
        time.sleep(2)
        
        clipboard_data = clipboard.paste()
        time.sleep(2)
        
        screen_clear_search()

        entire_data_as_list= clipboard_data.splitlines()
        for line in entire_data_as_list:
            if line.strip():
                output_lst_newline_removed.append(line.strip())

        clipboard_data = ''
        return output_lst_newline_removed
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in scrape_get_contents_by_search_copy_paste="+str(ex))

def screen_clear_search(delay=0.2):
    """
    Clears previously found text (crtl+f highlight)
    """
    try:
        pg.hotkey("ctrl","f")
        time.sleep(delay)
        pg.typewrite("^%#")
        time.sleep(delay)
        pg.hotkey("esc")
        time.sleep(delay)
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in screen_clear_search="+str(ex))

def search_highlight_tab_enter_open(searchText="",hitEnterKey="Yes",shift_tab='No'):
    """
    Searches for a text on screen using crtl+f and hits enter.
    This function is useful in Citrix environment
    """
    try:
        if not searchText:
            searchText = gui_get_any_input_from_user("Search Text to Highlight (in Citrix Environment)")

        time.sleep(0.5)
        
        pg.hotkey("ctrl", "f")
        time.sleep(0.5)
        
        pg.write(searchText)
        time.sleep(0.5)
        
        pg.hotkey("enter")
        time.sleep(0.5)
        
        pg.hotkey("esc")
        time.sleep(0.2)
        if hitEnterKey.lower() == "yes" and shift_tab.lower() == "yes":
            
            pg.hotkey("tab")
            time.sleep(0.3)
            
            pg.hotkey("shift", "tab")
            time.sleep(0.3)
            
            pg.hotkey("enter")
            time.sleep(2)
        elif hitEnterKey.lower() == "yes" and shift_tab.lower() == "no":
            
            pg.hotkey("enter")
            time.sleep(2)
        return True

    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in search_highlight_tab_enter_open="+str(ex))

def find_text_on_screen(searchText="",delay=0.1, occurance=1,isSearchToBeCleared=False):
    """
    Clears previous search and finds the provided text on screen.
    """
    screen_clear_search() #default

    if not searchText:
        searchText = gui_get_any_input_from_user("search text to Find on screen")

    time.sleep(delay)
    pg.hotkey("ctrl","f")
    time.sleep(delay)
    pg.typewrite(searchText)
    time.sleep(delay)

    for i in range(occurance-1):
        pg.hotkey("enter")
        time.sleep(delay)

    pg.hotkey("esc")
    time.sleep(delay)

    if isSearchToBeCleared:
        screen_clear_search()

# --------- Screenscraping Functions Ends ---------


# --------- Schedule Functions ---------

def schedule_create_task_windows(Weekly_Daily="D",week_day="Sun",start_time_hh_mm_24_hr_frmt="11:00"):
    """
    Schedules (weekly & daily options as of now) the current BOT (.bat) using Windows Task Scheduler. Please call create_batch_file() function before using this function to convert .pyw file to .bat
    """
    global batch_file_path
    try:

        str_cmd = ""

        if not batch_file_path:
            batch_file_path = gui_get_any_file_from_user('BATCH file to Schedule. Please call create_batch_file() to create one')

        if Weekly_Daily == "D":
            str_cmd = r"powershell.exe Start-Process schtasks '/create  /SC DAILY /tn ClointFusion\{} /tr {} /st {}' ".format(bot_name,batch_file_path,start_time_hh_mm_24_hr_frmt)
        elif Weekly_Daily == "W":
            str_cmd = r"powershell.exe Start-Process schtasks '/create  /SC WEEKLY /D {} /tn ClointFusion\{} /tr {} /st {}' ".format(week_day,bot_name,batch_file_path,start_time_hh_mm_24_hr_frmt)

        subprocess.call(str_cmd)
        print("Task Scheduled")
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in schedule_create_task_windows="+str(ex))

def schedule_delete_task_windows():
    """
    Deletes already scheduled task. Asks user to supply task_name used during scheduling the task. You can also perform this action from Windows Task Scheduler.
    """
    try:
        str_cmd = r"powershell.exe Start-Process schtasks '/delete /tn ClointFusion\{} ' ".format(bot_name)
        
        subprocess.call(str_cmd)
        print("Task {} Deleted".format(bot_name))

    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in schedule_delete_task="+str(ex))

# --------- Schedule Functions Ends ---------


# --------- Email Functions ---------

def email_send_via_desktop_outlook(toAddress="",ccAddress="",subject="",htmlBody="",embedImgPath="",attachmentFilePath=""):
    """
    Send email using Outlook from Desktop email application
    """
    try:
        if os_name == windows_os:
            if toAddress and subject and htmlBody:
                import win32com.client 
                outlook = win32com.client.Dispatch('outlook.application')
                mail = outlook.CreateItem(0)

                if type(toAddress) is list:
                    for m in toAddress:
                        mail.Recipients.Add(m)
                else:        
                    mail.To = toAddress

                mail.CC = ccAddress

                mail.Subject = subject

                mail.HTMLBody = f"<body><html> {htmlBody} <br> <img src="" cid:MyId1""> </body></html>"

                if embedImgPath:
                    attachment = mail.Attachments.Add(embedImgPath)
                    attachment.PropertyAccessor.SetProperty("http://schemas.microsoft.com/mapi/proptag/0x3712001F", "MyId1")
                    
                if attachmentFilePath:
                    mail.Attachments.Add(attachmentFilePath)

                mail.Send()

                print(f"Mail sent to {toAddress}")
        else:
            print("This feature is available only on Windows OS")

    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in email_send_via_desktop_outlook="+str(ex))

# --------- Email Functions Ends ---------


# --------- Utility Functions ---------

def ocr_now(img_path=""):
    """
    Recognize and “read” the text embedded in images using Google's Tesseract-OCR
    """
    try:
        import base64
        
        ocr_url = "https://api.clointfusion.com/ocr_now"
        
        with open(img_path, "rb") as image2string:
            converted_string = base64.b64encode(image2string.read())
        
            response = requests.post(ocr_url,data={'img_as_base64':converted_string})
        
        return response.text
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        return "Error in ocr_now" + str(ex)

def find(function_partial_name=""):
    # Find and inspect python functions
    try:
        if function_partial_name:
            import ClointFusion.cce as cf
            pi.search(cf, name=function_partial_name)
        
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in find="+str(ex))

def pause_program(seconds="5"):
    """
    Stops the program for given seconds
    """
    try:
        seconds = int(seconds)
        time.sleep(seconds)
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in pause_program="+str(ex))

def create_batch_file(application_exe_pyw_file_path=""):
    """
    Creates .bat file for the given application / exe or even .pyw BOT developed by you. This is required in Task Scheduler.
    """

    global batch_file_path
    try:
        if not application_exe_pyw_file_path:
            
            application_exe_pyw_file_path = gui_get_any_file_from_user('.pyw/.exe file for which .bat is to be made')

            while not (str(application_exe_pyw_file_path).endswith(".exe") or str(application_exe_pyw_file_path).endswith(".pyw")):
                print("Please choose the file ending with .pyw or .exe")
                application_exe_pyw_file_path = gui_get_any_file_from_user('.pyw/.exe file for which .bat is to be made')
            
        application_name= ""

        if str(application_exe_pyw_file_path).endswith(".exe"):
            application_name = _extract_filename_from_filepath(application_exe_pyw_file_path) + ".exe"
        else:
            application_name = _extract_filename_from_filepath(application_exe_pyw_file_path) + ".pyw"

        cmd = ""

        if "exe" in application_name:
            application_name = str(application_name).replace("exe","bat")
            cmd = "start \"\" " + '"' + application_exe_pyw_file_path + '" /popup\n'

        elif "pyw" in application_name: 
            application_name = str(application_name).replace("pyw","bat")
            cmd = "start \"\" " + '"' + sys.executable + '" ' + '"' + application_exe_pyw_file_path + '" /popup\n'

        batch_file_path = os.path.join(batch_file_path,application_name)
        batch_file_path = Path(batch_file_path)
        
        if not os.path.exists(batch_file_path):
            
            f = open(batch_file_path, 'w',encoding="utf-8")
            f.write("@ECHO OFF\n")
            f.write("timeout 5 > nul\n")
            f.write(cmd) 
            f.write("exit")    
            f.close()

        print("Batch file saved in " + str(batch_file_path))
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in create_batch_file="+str(ex))

    finally:
        return batch_file_path

def dismantle_code(strFunctionName=""):
    """
    This functions dis-assembles given function and shows you column-by-column summary to explain the output of disassembled bytecode.

    Ex: dismantle_code(show_emoji)
    """
    try:
        import dis

        if not strFunctionName:
            strFunctionName = gui_get_any_input_from_user('Exact function name to dis-assemble. Ex: show_emoji')
            print("Code dismantling {}".format(strFunctionName))
            return dis.dis(strFunctionName) 
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in dismantle_code="+str(ex)) 

def download_this_file(url=""):
    """
    Downloads a given url file to BOT output folder or Browser's Download folder
    """
    try:
        if not url:
            url = gui_get_any_input_from_user('URL to Download')

        if "export" in url:
            webbrowser.open_new(url)

        else:
            extension = str(url).rsplit("." ,1)[1]
            r = requests.get(url) 
            download_file_path = output_folder_path / "downloaded_cf.{}".format(extension)

            with open(download_file_path,'wb') as f: 
                f.write(r.content) 
                
            message_toast("File downloaded", file_folder_path=download_file_path)                
            return download_file_path

    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in download_this_file="+str(ex))

def clear_screen():
    """
    Clears Python Interpreter Terminal Window Screen
    """
    try:
      command = 'clear'
      if os.name in ('nt', 'dos'):  # If Machine is running on Windows, use cls
        command = 'cls'
      os.system(command)
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in clear_screen = " + str(ex))

def speech_to_text():
    """
    Speech to Text using Google's Generic API
    """
    global engine
    bol_url = "https://api.clointfusion.com/update_bol"
    system_uuid = selft.get_uuid()

    while True:
        with sr.Microphone() as source:
            r.dynamic_energy_threshold = True
            if r.energy_threshold in energy_threshold or r.energy_threshold <= sorted(energy_threshold)[-1]:
                r.energy_threshold = sorted(energy_threshold)[-1]
            else:
                energy_threshold.append(r.energy_threshold)

            r.pause_threshold = 0.8

            r.adjust_for_ambient_noise(source)

            audio=r.listen(source)

            try:
                query = r.recognize_google(audio)

                try:
                    resp = requests.post(bol_url, data={'user_uuid':str(system_uuid),'user_cmd':str(query)})
                    print(resp.text)
                except Exception as ex:
                    selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
                    print(str(ex))

                print(f"You Said : {query}")
                
                clear_screen()
                print("ClointFusion Bol is here to help you")
                return query
                break
            except AttributeError:
                print('Could not find PyAudio or no Microphone input device found. It may be being used by another application.')
                text_to_speech("Could not find PyAudio or no Microphone input device found. It may be being used by another application.")
                sys.exit()
            except sr.UnknownValueError:
                pass
            except sr.RequestError as e:
                print("Try Again")    

# --------- Utility Functions Ends ---------


# --------- Self-Test and ClointFusion Related Functions ---------

def _init_cf_quick_test_log_file(log_path_arg):
    """
    Internal function to generates the log and saves it to the file in the given base directory. 
    """
    global log_file_path
    
    try:
        
        dt_tm= str(datetime.datetime.now())    
        dt_tm = dt_tm.replace(" ","_")
        dt_tm = dt_tm.replace(":","-")
        dt_tm = dt_tm.split(".")[0]

        log_file_path = Path(os.path.join(log_path_arg, str(dt_tm) + ".txt"))
        
        for handler in logging.root.handlers[:]:
            logging.root.removeHandler(handler)
            
        logging.basicConfig(filename=log_file_path, level=logging.INFO, format='%(asctime)s  :  %(message)s',datefmt='%Y-%m-%d %H:%M:%S')           
    
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in _init_log_file="+str(ex))
    finally:
        host_ip = socket.gethostbyname(socket.gethostname()) 
        logging.info("{} ClointFusion Self Testing initiated for version {}".format(str(os_name).title(), s_version))
        logging.info("{}/{}".format(host_ip,str(get_public_ip())))

def _rerun_clointfusion_first_run(ex):
    try:
        pg.alert("Please Re-run..." + str(ex))
    except:
        put_text("Please Re-run..." + str(ex)).show()

def call_social_media():
    #opens all social media links of ClointFusion
    try:
        webbrowser.open_new_tab("https://www.facebook.com/ClointFusion")
    except Exception as ex:
        print("Error in call_social_media = " + str(ex))
 
    try:
        webbrowser.open_new_tab("https://twitter.com/ClointFusion")
    except Exception as ex:
        print("Error in call_social_media = " + str(ex))

    try:
        webbrowser.open_new_tab("https://www.youtube.com/channel/UCIygBtp1y_XEnC71znWEW2w")
    except Exception as ex:
        print("Error in call_social_media = " + str(ex))

    try:
        webbrowser.open_new_tab("https://www.linkedin.com/showcase/clointfusion_official")
    except Exception as ex:
        print("Error in call_social_media = " + str(ex))

    try:
        webbrowser.open_new_tab("https://www.reddit.com/user/Cloint-Fusion")
    except Exception as ex:
        print("Error in call_social_media = " + str(ex))

    try:
        webbrowser.open_new_tab("https://www.instagram.com/clointfusion")
    except Exception as ex:
        print("Error in call_social_media = " + str(ex))

    try:
        webbrowser.open_new_tab("https://www.kooapp.com/profile/ClointFusion")
    except Exception as ex:
        print("Error in call_social_media = " + str(ex))

    try:
        webbrowser.open_new_tab("https://discord.com/invite/tsMBN4PXKH")
    except Exception as ex:
        print("Error in call_social_media = " + str(ex))

    try:
        webbrowser.open_new_tab("https://www.eventbrite.com/e/2-days-event-on-software-bot-rpa-development-with-no-coding-tickets-183070046437")
    except Exception as ex:
        print("Error in call_social_media = " + str(ex))

    try:
        webbrowser.open_new_tab("https://internshala.com/internship/detail/python-rpa-automation-software-bot-development-work-from-home-job-internship-at-clointfusion1631715670")
    except Exception as ex:
        print("Error in call_social_media = " + str(ex))

def clointfusion_self_test_cases(temp_current_working_dir, start_time, console_window_name):
    """
    Main function for Self Test, which is called by GUI
    """
    global enable_semi_automatic_mode, log_file_path

    TEST_CASES_STATUS_MESSAGE = ""
    SUCCESS = False

    red_close_PNG_1 = temp_current_working_dir / "RED-Close_1.PNG"

    if not os.path.exists(red_close_PNG_1):
        urllib.request.urlretrieve('https://raw.githubusercontent.com/ClointFusion/Image_ICONS_GIFs/main/RED-Close_1.PNG',red_close_PNG_1)

    test_folder_path = Path(os.path.join(temp_current_working_dir,"ClointFusion_Self_Tests"))
    test_run_excel_path = Path(os.path.join(test_folder_path,'Quick_Self_Test_Excel.xlsx'))
    user_chosen_test_folder = Path(temp_current_working_dir)
    test_folder_path = Path(test_folder_path)  
    test_run_excel_path = Path(test_run_excel_path)

    enable_semi_automatic_mode = True

    try: # Test Self Test Cases
        text_to_speech("I will let you know the progress of the self-test, and, explain what is being done on your computer.", show=False)
        text_to_speech("Sit back and relax, while i check your machine for compatibility of Clointfusion package...", show=False)
        
        message_pop_up('Importing ClointFusion', delay=2)
        print('Importing ClointFusion')
        print()
        print('ClointFusion imported successfully '+ show_emoji())
        print("\n____________________________________________________________\n")
        print()
        logging.info('ClointFusion imported successfully')
        
        try: # Creating temp folder for ClointFusion Self Test
            folder_create(user_chosen_test_folder) 
            print('Test folder location {}'.format(user_chosen_test_folder))
            logging.info('Test folder location {}'.format(user_chosen_test_folder))
            
            tmp_img_folder_path =  os.path.join(user_chosen_test_folder, "Images")
            tmp_batch_file_path = os.path.join(user_chosen_test_folder, "Batch_File")
            tmp_config_folder_path = os.path.join(user_chosen_test_folder, "Config_Files")
            tmp_output_folder_path = os.path.join(user_chosen_test_folder, "Output")
            tmp_error_screen_shots_path = os.path.join(user_chosen_test_folder, "Error_Screenshots")
            
            try:
                print('Creating sub folders viz. img/batch/config/output/error_screen_shot at {}'.format(temp_current_working_dir))
                folder_create(tmp_img_folder_path)
                folder_create(tmp_batch_file_path)
                folder_create(tmp_config_folder_path)
                folder_create(tmp_error_screen_shots_path)
                folder_create(tmp_output_folder_path)
            except Exception as ex:
                selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
                print('Unable to create basic sub-folders for img/batch/config/output/error_screen_shot=' + str(ex))
                logging.info('Unable to create basic sub-folders for img/batch/config/output/error_screen_shot=' + str(ex))
                TEST_CASES_STATUS_MESSAGE  += 'Unable to create basic sub-folders for img/batch/config/output/error_screen_shot=' + str(ex)

            print()
            print("\n____________________________________________________________\n")
            text_to_speech("ClointFusion Self Testing Initiated...", show=False)
            print('ClointFusion Self Testing Initiated '+show_emoji())
            logging.info('ClointFusion Self Testing Initiated')
            
        except Exception as ex:
            selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
            print('Error while creating sub-folders='+str(ex))
            logging.info('Error while creating sub-folders='+str(ex))
            TEST_CASES_STATUS_MESSAGE  += 'Error while creating sub-folders='+str(ex)

        try: # Folder Operations
            print()
            print("\n____________________________________________________________\n")
            text_to_speech("Let me test folder operations...", show=False)
            print('Testing folder operations')
            
            
            folder_create(Path(os.path.join(test_folder_path,"My Test Folder")))
            folder_create_text_file(test_folder_path, "My Text File")
            excel_create_excel_file_in_given_folder(test_folder_path,'Quick_Self_Test_Excel')
            excel_create_excel_file_in_given_folder(test_folder_path,'My Excel-1')
            excel_create_excel_file_in_given_folder(test_folder_path,'My Excel-2')

            try:
                excel_create_excel_file_in_given_folder(os.path.join(test_folder_path,"Delete Excel"),'Delete-Excel-1')
                excel_create_excel_file_in_given_folder(os.path.join(test_folder_path,"Delete Excel"),'Delete-Excel-2')
                folder_delete_all_files(os.path.join(test_folder_path,'Delete Excel'), "xlsx")
            except Exception as ex:
                selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
                print('Unable to delete files in test folder='+str(ex))
                logging.info('Unable to delete files in test folder='+str(ex))
                TEST_CASES_STATUS_MESSAGE  += 'Unable to delete files in test folder='+str(ex)

            folder_create(Path(test_folder_path / 'Split_Merge'))
            print(folder_get_all_filenames_as_list(test_folder_path))
            print(folder_get_all_filenames_as_list(test_folder_path, extension="xlsx"))
            
            text_to_speech("Folder operations test is successful", show=False)
            print('Folder operations tested successfully '+show_emoji())
            print("\n____________________________________________________________\n")
            logging.info('Folder operations tested successfully')
        except Exception as ex:
            selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
            print('Error while testing Folder operations='+str(ex))
            logging.info('Error while testing Folder operations='+str(ex))
            TEST_CASES_STATUS_MESSAGE  += 'Error while testing Folder operations='+str(ex)

        if os_name == windows_os: # Window operations
            try:
                print()
                print("\n____________________________________________________________\n")
                text_to_speech("This is Windows PC, I love Windows, let me test the windows specific functions.", show=False)
                print('Started testing window based operations...')
                
                window_show_desktop()
                launch_any_exe_bat_application(test_run_excel_path)
                time.sleep(5)
                window_minimize_windows('Quick_Self_Test_Excel')
                time.sleep(5)
                window_activate_and_maximize_windows('Quick_Self_Test_Excel')
                time.sleep(5)
                print(window_get_all_opened_titles_windows())
                window_close_windows('Quick_Self_Test_Excel')
                
                print('Window based operations tested successfully '+show_emoji())
                print("\n____________________________________________________________\n")
                text_to_speech("Great, Windows functions work flawlessly on your PC. Cannot wait to see how you use these functions.", show=False)
                logging.info('Window based operations tested successfully')
            except Exception as ex:
                selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
                print('Error while testing window based operations='+str(ex))
                logging.info('Error while testing window based operations='+str(ex))
                TEST_CASES_STATUS_MESSAGE  += 'Error while testing window based operations='+str(ex)
        else:
            text_to_speech("Skipping window operations as it is Windows OS specific.", show=False)
            print('Skipping window operations as it is Windows OS specific')
            logging.info('Skipping window operations as it is Windows OS specific')
            TEST_CASES_STATUS_MESSAGE  += 'Skipping window operations as it is Windows OS specific'
        
        try: # Excel operations
            print()
            print("\n____________________________________________________________\n")
            text_to_speech("Let me test the excel compatability.", show=False)
            print('Started testing excel operations...')

            excel_create_excel_file_in_given_folder(test_folder_path, "Test_Excel_File", "Test_Sheet")
            print(excel_get_row_column_count(test_run_excel_path))
            excel_create_excel_file_in_given_folder(test_folder_path,excelFileName="Excel_Test_Data")
            test_excel_path = test_folder_path / "Excel_Test_Data.xlsx"
            excel_set_single_cell(test_excel_path,columnName="Name",cellNumber=0,setText="A")
            excel_set_single_cell(test_excel_path,columnName="Name",cellNumber=1,setText="B")
            excel_set_single_cell(test_excel_path,columnName="Name",cellNumber=2,setText="C")
            excel_set_single_cell(test_excel_path,columnName="Name",cellNumber=3,setText="D")
            excel_set_single_cell(test_excel_path,columnName="Name",cellNumber=4,setText="E")
            excel_set_single_cell(test_excel_path,columnName="Age",cellNumber=0,setText="1")
            excel_set_single_cell(test_excel_path,columnName="Age",cellNumber=1,setText="2")
            excel_set_single_cell(test_excel_path,columnName="Age",cellNumber=2,setText="4")
            excel_set_single_cell(test_excel_path,columnName="Age",cellNumber=3,setText="3")
            excel_set_single_cell(test_excel_path,columnName="Age",cellNumber=4,setText="5")
            print(excel_get_single_cell(test_excel_path,sheet_name='Sheet1',columnName='Name'))
            excel_create_file(test_folder_path,"My New Paste Excel")
            excel_create_excel_file_in_given_folder(test_folder_path,'My Excel-3','CF-Sheet-1')
            excel_file_path = test_folder_path / 'My Excel-3.xlsx'
            print(excel_get_all_sheet_names(excel_file_path))
            print(excel_get_all_sheet_names(test_run_excel_path))
            excel_copied_Data=excel_copy_range_from_sheet(test_excel_path, sheet_name="Sheet1", startCol=1, startRow=1, endCol=2, endRow=6)
            print(excel_copied_Data)
            excel_copy_paste_range_from_to_sheet(Path(os.path.join(test_folder_path,"My New Paste Excel.xlsx")), sheet_name="Sheet1", startCol=1, startRow=1, endCol=2, endRow=6, copiedData=excel_copied_Data)
            excel_split_by_column(excel_path=Path(os.path.join(test_folder_path,"My New Paste Excel.xlsx")), sheet_name="Sheet1", header=0, columnName="Name")
            folder_create(Path(test_folder_path / 'Split_Merge'))
            excel_split_the_file_on_row_count(excel_path=Path(test_folder_path / "My New Paste Excel.xlsx"), sheet_name="Sheet1", rowSplitLimit=1, outputFolderPath=os.path.join(test_folder_path,'Split_Merge'), outputTemplateFileName="Split")
            excel_merge_all_files(input_folder_path=test_folder_path / "Split_Merge", output_folder_path=Path(test_folder_path,'Split_Merge'))
            excel_drop_columns(Path(test_folder_path / "My New Paste Excel.xlsx"), columnsToBeDropped ="Age")
            excel_sort_columns(excel_path=test_excel_path, sheet_name="Sheet1", header=0, firstColumnToBeSorted="Age", secondColumnToBeSorted="Name")
            excel_clear_sheet(Path(test_folder_path / "My New Paste Excel.xlsx"), sheet_name="Sheet1", header=0)
            excel_set_single_cell(test_excel_path,columnName="Name",cellNumber=5,setText="E")
            excel_set_single_cell(test_excel_path,columnName="Age",cellNumber=5,setText="5")
            excel_remove_duplicates(excel_path=test_excel_path, sheet_name="Sheet1", header=0,columnName="Name", which_one_to_keep="first")
            excel_create_file(test_folder_path,"My VLookUp Excel")
            excel_set_single_cell(Path(test_folder_path,"My VLookUp Excel.xlsx"),columnName="Name",cellNumber=0,setText="A")
            excel_set_single_cell(Path(test_folder_path,"My VLookUp Excel.xlsx"),columnName="Name",cellNumber=1,setText="B")
            excel_set_single_cell(Path(test_folder_path,"My VLookUp Excel.xlsx"),columnName="Name",cellNumber=2,setText="C")
            excel_set_single_cell(Path(test_folder_path,"My VLookUp Excel.xlsx"),columnName="Name",cellNumber=3,setText="D")
            excel_set_single_cell(Path(test_folder_path,"My VLookUp Excel.xlsx"),columnName="Name",cellNumber=4,setText="E")
            excel_set_single_cell(Path(test_folder_path,"My VLookUp Excel.xlsx"),columnName="Salary",cellNumber=0,setText="1")
            excel_set_single_cell(Path(test_folder_path,"My VLookUp Excel.xlsx"),columnName="Salary",cellNumber=1,setText="2")
            excel_set_single_cell(Path(test_folder_path,"My VLookUp Excel.xlsx"),columnName="Salary",cellNumber=2,setText="4")
            excel_set_single_cell(Path(test_folder_path,"My VLookUp Excel.xlsx"),columnName="Salary",cellNumber=3,setText="3")
            excel_set_single_cell(Path(test_folder_path,"My VLookUp Excel.xlsx"),columnName="Salary",cellNumber=4,setText="5")
            excel_vlook_up(filepath_1=test_excel_path,filepath_2=Path(test_folder_path,"My VLookUp Excel.xlsx"),match_column_name="Name")
            
            print('Excel operations tested successfully '+show_emoji())
            print("\n____________________________________________________________\n")
            text_to_speech("Excel operations tested successfully.", show=False)
            logging.info('Excel operations tested successfully')
            
            text_to_speech("Isn't amazingly quick, But yeah, more than 28 functions, and, more than 50 excel operations, have been tested just now.", show=False)
            text_to_speech("Banking, finance, from data collection, to data cleaning, and sending reports, everything can be done, with excel functions and Clointfusion. Give it a try today.", show=False)
        except Exception as ex:
            selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
            print("Error while testing Excel Operations="+str(ex))
            logging.info("Error while testing Excel Operations="+str(ex))
            TEST_CASES_STATUS_MESSAGE  += "Error while testing Excel Operations="+str(ex)
            
        try: # String operations
            print()
            print("\n____________________________________________________________\n")
            text_to_speech("Testing String manipulation functions.", show=False)
            print('Started testing String Operations...')
            
            print(string_remove_special_characters("C!@loin#$tFu*(sion"))
            print(string_extract_only_alphabets(inputString="C1l2o#%^int&*Fus12i5on"))
            print(string_extract_only_numbers("C1l2o3i4n5t6F7u8i9o0n"))
            
            print('String operations tested successfully '+show_emoji())
            print("\n____________________________________________________________\n")
            text_to_speech("String operations tested successfully.", show=False)
            logging.info('String operations tested successfully')
        except Exception as ex:
            selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
            print('Error while testing string operations='+str(ex))
            logging.info('Error while testing string operations='+str(ex))
            TEST_CASES_STATUS_MESSAGE  += "Error while testing string operations="+str(ex)
            
        try: # Keyboard operations
            print()
            print("\n____________________________________________________________\n")
            text_to_speech("Let me test keyboard functions.", show=False)
            print('Started testing keyboard operations...')
            
            add_msg = f"Hi {user_name},\nClointFusion is celebrating One Year of Hackathon ! Motivate your friends to register Now : https://tinyurl.com/ClointFusion" #"Performing ClointFusion Self Test for Notepad"

            if os_name == windows_os:
                launch_any_exe_bat_application("notepad") # Windows
                key_write_enter(write_to_window="notepad",text_to_write=add_msg)
                text_to_speech("By the way, keep your data a secret, not me, Let your friends know, how cool ClointFusion is ..!", show=False)
                key_hit_enter(write_to_window="notepad")
                key_press(key_1="alt", key_2="f4", write_to_window="notepad")
                key_press("right")
                key_hit_enter()
            elif os_name == linux_os:
                launch_any_exe_bat_application("gedit") # Ubuntu
                key_write_enter(text_to_write=add_msg)
                text_to_speech("By the way, keep your data a secret, not me, Let your friends know, how cool ClointFusion is", show=False)
                key_hit_enter()
                key_press(key_1="alt", key_2="f4")
                subprocess.Popen(f"killall -9 gedit", shell=True,
                           stdout=subprocess.PIPE, 
                           stderr=subprocess.PIPE)
            elif os_name == mac_os:
                try:
                    launch_any_exe_bat_application("TextEdit") # macOS
                    key_write_enter(text_to_write=add_msg)
                    text_to_speech("By the way, keep your data a secret, not me, Let your friends know, how cool ClointFusion is", show=False)
                    key_hit_enter()
                    key_press(key_1="command", key_2="f4")
                    subprocess.Popen('pkill -9 "TextEdit"', shell=True,
                           stdout=subprocess.PIPE, 
                           stderr=subprocess.PIPE)
                except:
                    print("Currently Not Supported.")
                    logging.info('Keyboard operations Skipped.')
                    TEST_CASES_STATUS_MESSAGE  += "Keyboard operations Skipped for MAC OS"
            
            print('Keyboard operations tested successfully '+show_emoji())
            print("\n____________________________________________________________\n")


            text_to_speech("Keyboard operations tested successfully.", show=False)
            logging.info('Keyboard operations tested successfully')
        except Exception as ex:
            selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
            print('Error in keyboard operations='+str(ex))
            logging.info('Error in keyboard operations='+str(ex))
            try:
                key_press(key_1="alt", key_2="f4")
            except:
                pg.hotkey("alt", "f4")

        try: # Screen scraping operations
            print()
            print("\n____________________________________________________________\n")
            text_to_speech("Testing screen-scraping functions..", show=False)
            print('Started testing screen-scraping operations...')
            
            browser_activate('https://google.com')
            text_to_speech("Thanks to internet, we are indexed, on all search engines, just type clointfusion hackathon", show=False)
            browser_write_h('clointfusion hackathon')
            time.sleep(5)
            browser_hit_enter_h()
            browser_mouse_click_h("Python based RPA Development Platform")
            folder_create(os.path.join(test_folder_path,'Screen_scrape'))
            scrape_save_contents_to_notepad(test_folder_path / 'Screen_scrape', switch_to_window="Python based")
            time.sleep(3)
            browser_navigate_h("https://sites.google.com/view/clointfusion-hackathon/date-with-clointfusion")
            text_to_speech("Date with ClointFusion, is an initiative, for fast track entry, into our growing workforce.", show=False)
            browser.scroll_down(2000)
            time.sleep(10)
            
            
            print('Screen-scraping operations tested successfully '+show_emoji())
            print("\n____________________________________________________________\n")
            text_to_speech("Screen-scraping functions tested successfully..", show=False)
            logging.info("Screen-scraping functions tested successfully")
        except Exception as ex:
            selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
            print('Error while testing screenscraping functions='+str(ex))
            logging.info('Error while testing screenscraping functions='+str(ex))
            TEST_CASES_STATUS_MESSAGE  += 'Error while testing screenscraping functions='+str(ex)

        try: # Mouse operations
            print()
            print("\n____________________________________________________________\n")
            text_to_speech("Testing mouse operations", show=False)
            print('Started testing mouse operations...')
            
            browser_navigate_h("https://sites.google.com/view/clointfusion-hackathon")
            mouse_move(850,500)
            time.sleep(2)
            mouse_drag_from_to(600,450,1150,680)
            search_highlight_tab_enter_open("chat.whatsapp")
            text_to_speech("Please join our WhatsApp community group to keep yourself updated on our progress.", show=False)
            mouse_click(int(pg.size()[0]/2),int(pg.size()[1]/2)) #Click at center of the screen
            browser_quit_h()

            print('Mouse operations tested successfully ' + show_emoji())
            print("\n____________________________________________________________\n")
            text_to_speech("Mouse operations tested successfully", show=False)
            logging.info('Mouse operations tested successfully')
        except Exception as ex:
            selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
            print('Error in mouse operations='+str(ex))
            logging.info('Error in mouse operations='+str(ex))
            key_press(key_1="ctrl", key_2="w")
            TEST_CASES_STATUS_MESSAGE  += 'Error in mouse operations='+str(ex)
        

        try: # Browser operations
            print()
            print("\n____________________________________________________________\n")
            text_to_speech("Great going, almost done, let me quickly test the browser functions.", show=False)
            print("Started testing Browser's Operations...")
            
            if browser_activate("https://pypi.org"):
                browser_write_h("ClointFusion",User_Visible_Text_Element="Search projects")
                browser_hit_enter_h()
                time.sleep(5)
                browser_navigate_h('https://pypi.org/project/ClointFusion/')
                text_to_speech("This is our pypi page, you can read, our detailed documentation, and view, our well, explained gifs.", show=False)
                browser.scroll_down(2000)
                time.sleep(10)
                browser_mouse_click_h(element="RPA",double_click=True)
                browser_mouse_click_h(element=browser_locate_element_h('//*[@id="description"]/div/h2[2]/a'))
                text_to_speech("By the way, this is your DOST, powered by Clointfusion, Made automation easy, just drag, and drop, and automate.", show=False)
                text_to_speech("Type D O S T, in terminal to launch.", show=False)
                pause_program(10)
                browser_quit_h()
                
                print("Tested Browser's Helium functions successfully " + show_emoji())
                print("\n____________________________________________________________\n")
                text_to_speech("Browser functions tested successfully..", show=False)
                logging.info("Tested Browser's Helium functions successfully")
            else:
                TEST_CASES_STATUS_MESSAGE  += "Helium package's Compatible Chrome is missing"
        except Exception as ex:
            selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
            print("Error while Testing Browser Helium functions="+str(ex))
            logging.info("Error while Testing Browser Helium functions="+str(ex))
            key_press(key_1="ctrl", key_2="w") #to close any open browser
            TEST_CASES_STATUS_MESSAGE  += "Error while Testing Browser Helium functions="+str(ex)

        try: # CLI operations
            print()
            print("\n____________________________________________________________\n")
            text_to_speech("Let me show you, some terminal commands,", show=False)
            print('Started testing CLIs Operations...')
            
            text_to_speech("Press CF in terminal to get all the available commands", show=False)
            print("Press 'cf' in terminal for this function.\n")
            #Test work
            cli_cf_test()
            
            text_to_speech("Want to quickly test your internet speed, type CF underscore ST.", show=False)
            print("Type 'cf_st' in terminal for this function")
            text_to_speech("Internet speed is being tested in the terminal.", show=False)
            
            # Test Speed Test
            window_activate_and_maximize_windows(console_window_name)
            cli_speed_test_test()
            window_minimize_windows(console_window_name)
            window_activate_window("Welcome to ClointFusion - Made in India with LOVE")
            print("I am back !!!")
            
            text_to_speech("Want to see, how much time you spent, on what, and which application. type work in terminal. But First, let me tell you this, your data is stored locally and only belongs to you.", show=False)
            print("Type 'work' in terminal for this function.")
            # Test BRE WHM
            cli_bre_whm_test()
            text_to_speech("This Report looks colorful and neat in a terminal, give it a try, after self-test", show=False)
            
            print("CLI's functions tested successfully"+show_emoji())
            print("\n____________________________________________________________\n")
            text_to_speech("CLI's functions tested successfully", show=False)
            logging.info("CLI functions tested successfully")    
        except Exception as ex:
            selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
            print("Error while Testing CLIs")
            logging.info("Error while Testing CLIs")
            TEST_CASES_STATUS_MESSAGE  += "Error while Testing CLIs"
        
        try: # Message operations
            print()
            print("\n____________________________________________________________\n")
            text_to_speech("Ok, Let me show you some different message pop ups.", show=False)
            print("Started testing message functions...")
            
            text_to_speech("Message with timer...", show=False)
            message_counter_down_timer("Message with timer (in seconds)",6)
            
            text_to_speech("Popup Message...", show=False)
            message_pop_up("Testing popup message.")
            
            text_to_speech("Flash Message...", show=False)
            message_flash("Testing flash message.")
            
            print('Message operations tested successfully '+show_emoji())
            text_to_speech("Message functions tested successfully", show=False)
            print("\n____________________________________________________________\n")
            logging.info('Message functions tested successfully')
        except Exception as ex:
            selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
            print("Error while testing Flash message="+str(ex))
            logging.info("Error while testing Flash message="+str(ex))
            TEST_CASES_STATUS_MESSAGE  += "Error while testing Flash message="+str(ex)
            
        try:
            file_contents = ''
            try:
                with open(log_file_path, encoding="utf-8") as f:
                    file_contents = f.readlines()
            except Exception as ex:
                selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
                file_contents = 'Unable to read the file' + str(ex)
                TEST_CASES_STATUS_MESSAGE += "Unable to read log file" + str(ex)
            
            try:
                text_to_speech("Let me quickly, do the neccessary registration, for you. So you can, get started with automation", show=False)
                time_taken= timedelta(seconds=time.monotonic()  - start_time)
                os_hn_ip = "OS:{}".format(os_name) + "HN:{}".format(socket.gethostname()) + ",IP:" + str(socket.gethostbyname(socket.gethostname())) + "/" + str(get_public_ip())
            except Exception as ex:
                selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
                TEST_CASES_STATUS_MESSAGE += str(ex)
            
            try:
                driver = selft.gf(os_hn_ip, time_taken, file_contents)

                cursr.execute("UPDATE CF_IMP_VALUES set SELF_TEST = 'True' where ID = 1")
                connct.commit()
                clear_screen()
                time.sleep(5)
                text_to_speech("Closing the browser now.", show=False)
                browser.set_driver(driver)
                browser_quit_h()
                time.sleep(2)
                selft.ast()
                SUCCESS = True
                time.sleep(5)
            except Exception as ex:
                selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
                TEST_CASES_STATUS_MESSAGE += str(ex)
            
            if str(TEST_CASES_STATUS_MESSAGE).strip()  == "":
                print("")
                logging.info("ClointFusion Self Testing Completed")
                print("Congratulations - ClointFusion is compatible with your computer " + show_emoji('clap') + show_emoji('clap'))
                message_pop_up("Congratulations !!!\n\nClointFusion is compatible with your computer settings")
                print("\n____________________________________________________________\n")
                text_to_speech("Self Test is Completed. And i am happy to say, your PC is compatible with ClointFusion", show=False)
                
        except Exception as ex:
            selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
            print("ClointFusion Automated Testing Failed "+str(ex))
            logging.info("ClointFusion Automated Testing Failed "+str(ex))
            TEST_CASES_STATUS_MESSAGE  += "ClointFusion Automated Testing Failed "+str(ex)
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        print("ClointFusion Automated Testing Failed "+str(ex))
        logging.info("ClointFusion Automated Testing Failed "+str(ex))
        TEST_CASES_STATUS_MESSAGE  += "ClointFusion Automated Testing Failed "+str(ex)
        
    finally:
        enable_semi_automatic_mode = False
        
        if str(TEST_CASES_STATUS_MESSAGE).strip()  == "":
            text_to_speech("All set and ready to go. Wanna to call me again, type B O L, in terminal", show=False)
            print("Type 'bol'")
            text_to_speech("Always remember, Think what to automate, not how to, because, Clointfusion is there to do it.", show=False)
            message_toast("ClointFusion is compatible with your computer's settings !", website_url="https://tinyurl.com/ClointFusion")
        else:
            print("ClointFusion Self Testing has Failed for few Functions")
            print(TEST_CASES_STATUS_MESSAGE)
            logging.info("ClointFusion Self Testing has Failed for few Functions")
            logging.info(TEST_CASES_STATUS_MESSAGE)
            SUCCESS = False

        if os_name == windows_os:
            try:
                window_close_windows('Welcome to ClointFusion - Made in India with LOVE')
            except:
                pos = mouse_search_snip_return_coordinates_x_y(str(red_close_PNG_1),wait=5)
                mouse_click(pos[0], pos[1])
        else:
            print("Please click red 'Close' button")
        return TEST_CASES_STATUS_MESSAGE, SUCCESS

def clointfusion_self_demo_tour(temp_current_working_dir, start_time, console_window_name, tour=False):
    TEST_CASES_STATUS_MESSAGE = ""
    SUCCESS = False

    test_folder_path = Path(os.path.join(temp_current_working_dir,"ClointFusion_Self_Tests"))
    test_run_excel_path = Path(os.path.join(test_folder_path,'Quick_Self_Test_Excel.xlsx'))
    user_chosen_test_folder = Path(temp_current_working_dir)
    test_folder_path = Path(test_folder_path)  
    test_run_excel_path = Path(test_run_excel_path)

    enable_semi_automatic_mode = True

# SELF TEST AND TOUR STARTS HERE
    tour_comp = False
    try:

# Keyboard and Mouse Operations
        try: 
            # Description
            if not tour:
                print()
                print("\n____________________________________________________________\n")
                print('Started testing Keyboard and Mouse operations...')
                text_to_speech("First and foremost, there is no automation, unless we can control, the mouse and keyboard.", show=False)
                text_to_speech("ClointFusion currently has three, keyboard functions, to control keyboard", show=False)
                message_pop_up("\nKeyboard\n```````````\n1)  key_press\n2)  key_write_enter\n3)  key_hit_enter\n\n", delay=5)
                text_to_speech("and we have four, mouse functions, to control mouse pointer", show=False)
                message_pop_up("\nMouse\n```````````\n1)  mouse_click\n2)  mouse_move\n3)  mouse_drag_from_to\n4)  mouse_search_snip_return_coordinates_x_y\n\n", delay=5)
                text_to_speech("with many more on the way.", show=False)
                text_to_speech("Now, let me put these functions to the test and show you how to automate them.", show=False)
            else:
                print("\n____________________________________________________________\n")
                text_to_speech("ClointFusion currently has three, keyboard functions, to control keyboard")
                message_pop_up("\nKeyboard\n```````````\n1)  key_press\n2)  key_write_enter\n3)  key_hit_enter\n\n", delay=5)
                print("\nKeyboard\n```````````\n1)  key_press\n2)  key_write_enter\n3)  key_hit_enter\n\n")
                text_to_speech("Four, mouse functions, to control mouse pointer")
                message_pop_up("\nMouse\n```````````\n1)  mouse_click\n2)  mouse_move\n3)  mouse_drag_from_to\n4)  mouse_search_snip_return_coordinates_x_y\n\n", delay=5)
                print("\nMouse\n```````````\n1)  mouse_click\n2)  mouse_move\n3)  mouse_drag_from_to\n4)  mouse_search_snip_return_coordinates_x_y\n\n")
                text_to_speech("Now let me use them to show you what all you can do.", show=False)
            
            # Actions
            
            add_msg = f"Hi {user_name},\nClointFusion is very happy and lucky to collaborate with IIT Dharwad for hackathon 14 !!! \nMotivate your friends to register Now : https://tinyurl.com/ClointFusion" #"Performing ClointFusion Self Test for Notepad"

            if os_name == windows_os:
                try:
                    if tour:
                        print("""\nCommands I will be using in the next scenario:\n~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~\nimport ClointFusion as cf\nscreen_width = cf.pg.size().width\nscreen_height = cf.pg.size().height\ncf.launch_any_exe_bat_application("mspaint")\ncf.pause_program(3)\ncf.key_press("ctrl", "pageup")\ncf.mouse_move(screen_width/2, screen_height/2)\ncf.key_press("b")\ncf.mouse_click()\ncf.key_press("e")\ncf.mouse_drag_from_to(screen_width/3, screen_height/3, screen_width/3-200, screen_height/3)\ncf.mouse_drag_from_to(screen_width/3-200, screen_height/3, screen_width/3-200, screen_height/3+200)\ncf.mouse_drag_from_to(screen_width/3-200, screen_height/3+200, screen_width/3, screen_height/3+200,)\ncf.mouse_drag_from_to(screen_width/3+30, screen_height/3, screen_width/3+30, screen_height/3+220)\ncf.mouse_drag_from_to(screen_width/3+30, screen_height/3, screen_width/3+230, screen_height/3)\ncf.mouse_drag_from_to(screen_width/3+30, screen_height/3+80, screen_width/3+230, screen_height/3+80)\ncf.key_press("t")\ncf.mouse_click()\ncf.key_write_enter("ClointFusion is awesome, right?")\ncf.key_press("ctrl", "a")\ncf.pause_program(3)\ncf.text_to_speech("ClointFusion is awesome, right?", show=False)\ncf.key_press("alt", "f4")\ncf.key_press("right")\ncf.key_hit_enter()""")
                    text_to_speech("I'll use MS Paint as a demonstration tool.", show=False)
                    screen_width = pg.size().width
                    screen_height = pg.size().height
                    launch_any_exe_bat_application("mspaint")
                    pause_program(3)
                    key_press("ctrl", "pageup")
                    if screen_width > 1920 or screen_height > 1080:
                        key_press("ctrl", "pageup")
                    mouse_move(screen_width/2, screen_height/2)
                    key_press("b")
                    mouse_click()
                    key_press("e")
                    text_to_speech("C", show=False)
                    mouse_drag_from_to(screen_width/3, screen_height/3, screen_width/3-200, screen_height/3)
                    mouse_drag_from_to(screen_width/3-200, screen_height/3, screen_width/3-200, screen_height/3+200)
                    mouse_drag_from_to(screen_width/3-200, screen_height/3+200, screen_width/3, screen_height/3+200,)
                    text_to_speech("F", show=False)
                    mouse_drag_from_to(screen_width/3+30, screen_height/3, screen_width/3+30, screen_height/3+220)
                    mouse_drag_from_to(screen_width/3+30, screen_height/3, screen_width/3+230, screen_height/3)
                    mouse_drag_from_to(screen_width/3+30, screen_height/3+80, screen_width/3+180, screen_height/3+80)
                    key_press("t")
                    mouse_click()
                    key_write_enter("ClointFusion is awesome, isn't?")
                    key_press("ctrl", "a")
                    pause_program(3)
                    text_to_speech("ClointFusion is awesome, isn't it?", show=False)
                    key_press("alt", "f4")
                    key_press("right")
                    key_hit_enter()
                    text_to_speech("But let me tell you something, the only limit to automation is your creativity.", show=False)  
                    launch_any_exe_bat_application("notepad") # Windows
                    key_write_enter(write_to_window="notepad",text_to_write=add_msg)
                    text_to_speech("By the way, keep your data a secret, not me, Let your friends know, how cool ClointFusion is", show=False)
                    key_hit_enter(write_to_window="notepad")
                    key_press(key_1="alt", key_2="f4", write_to_window="notepad")
                    key_press("right")
                    key_hit_enter()
                    window_show_desktop()
                except Exception as ex:
                    selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
                    print("ClointFusion Automated Testing Failed for Keyboard and mouse Functions | Windows "+str(ex))
                    logging.info("ClointFusion Automated Testing Failed for Keyboard and mouse Functions | Windows "+str(ex))
                    TEST_CASES_STATUS_MESSAGE  += "ClointFusion Automated Testing Failed for Keyboard and mouse Functions | Windows "+str(ex)
            elif os_name == linux_os:
                try:
                    launch_any_exe_bat_application("gedit") # Ubuntu
                    key_write_enter(text_to_write=add_msg)
                    text_to_speech("By the way, keep your data a secret, not me, Let your friends know, how cool ClointFusion is", show=False)
                    key_hit_enter()
                    key_press(key_1="alt", key_2="f4")
                    subprocess.Popen(f"killall -9 gedit", shell=True,
                            stdout=subprocess.PIPE, 
                            stderr=subprocess.PIPE)
                except Exception as ex:
                    selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
                    print("ClointFusion Automated Testing Failed for Keyboard and mouse Functions | Linux "+str(ex))
                    logging.info("ClointFusion Automated Testing Failed for Keyboard and mouse Functions | Linux "+str(ex))
                    TEST_CASES_STATUS_MESSAGE  += "ClointFusion Automated Testing Failed for Keyboard and mouse Functions | Linux "+str(ex)
            elif os_name == mac_os:
                try:
                    launch_any_exe_bat_application("TextEdit") # macOS
                    key_write_enter(text_to_write=add_msg)
                    text_to_speech("By the way, keep your data a secret, not me, Let your friends know, how cool ClointFusion is", show=False)
                    key_hit_enter()
                    key_press(key_1="command", key_2="f4")
                    subprocess.Popen('pkill -9 "TextEdit"', shell=True,
                           stdout=subprocess.PIPE, 
                           stderr=subprocess.PIPE)
                except Exception as ex:
                    selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
                    print("Currently Not Supported."+str(ex))
                    logging.info('Keyboard operations Skipped.')
                    TEST_CASES_STATUS_MESSAGE  += "Keyboard operations Skipped for MAC OS"
            if not tour:
                print()
                print('Finished testing Keyboard and Mouse operations. '+show_emoji())
                print("\n____________________________________________________________\n")
                logging.info('Keyboard and mouse operations tested successfully')
            else:
                print("\n____________________________________________________________\n")
                text_to_speech("Let's go on to the next section.")
                print("\n____________________________________________________________\n")
        except Exception as ex:
            selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
            print("ClointFusion Automated Testing Failed for Keyboard and mouse Functions"+str(ex))
            logging.info(' Error Keyboard and mouse operations tested successfully')
            TEST_CASES_STATUS_MESSAGE  += "Error while Testing Keyboard and mouse operations"

# Browser Operations
        try: 
            # Description
            if not tour:
                print()
                print("\n____________________________________________________________\n")
                print('Started testing Browser operations...')
                text_to_speech("With the pandemic, working from home, as well as everything of our employment and education, has gone entirely online. ", show=False)
                text_to_speech("If we cannot automate the web, there's no point in automation. ClointFusion includes 11 browser automation functions.", show=False)
                message_pop_up("\nBrowser\n```````````\n1)  browser_activate\n2)  browser_navigate_h\n3)  browser_write_h\n4)  browser_mouse_click_h\n5)  browser_locate_element_h\n6)  browser_wait_until_h\n7)  browser_refresh_page_h\n8)  browser_hit_enter_h\n9)  browser_key_press_h\n10)  browser_mouse_hover_h\n11)  browser_quit_h\n\n", delay=5)
                text_to_speech("With these functions, you can do whatever you want, and, however you like, in a web browser.", show=False)
                text_to_speech("Now, let me put these functions to the test.", show=False)
            else:
                print("\n____________________________________________________________\n")
                text_to_speech("ClointFusion currently has eleven, browser automation functions.")
                message_pop_up("\nBrowser\n```````````\n1)  browser_activate\n2)  browser_navigate_h\n3)  browser_write_h\n4)  browser_mouse_click_h\n5)  browser_locate_element_h\n6)  browser_wait_until_h\n7)  browser_refresh_page_h\n8)  browser_hit_enter_h\n9)  browser_key_press_h\n10)  browser_mouse_hover_h\n11)  browser_quit_h\n\n", delay=5)
                print("\nBrowser\n```````````\n1)  browser_activate\n2)  browser_navigate_h\n3)  browser_write_h\n4)  browser_mouse_click_h\n5)  browser_locate_element_h\n6)  browser_wait_until_h\n7)  browser_refresh_page_h\n8)  browser_hit_enter_h\n9)  browser_key_press_h\n10)  browser_mouse_hover_h\n11)  browser_quit_h\n\n")
                
            # Actions
            if tour:
                print("""\nCommands I will be using in the next scenario:\n~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~\nimport ClointFusion as cf\ncf.browser_activate('https://google.com')\ncf.browser_write_h('clointfusion hackathon')\ncf.pause_program(3)\ncf.browser_hit_enter_h()\ncf.browser_mouse_click_h("Python based RPA Development Platform")\ncf.folder_create(os.path.join(test_folder_path,'Screen_scrape'))\ncf.scrape_save_contents_to_notepad(test_folder_path / 'Screen_scrape', switch_to_window="Python based")\ncf.os.startfile(os.path.join(test_folder_path,'Screen_scrape'))\ncf.os.startfil(os.path.join(test_folder_path,'Screen_scrape','notepad-contents.txt'))\ncf.pg.scroll(10)\ncf.pause_program(3)\ncf.pg.scroll(10)\ncf.window_close_windows('Screen_scrape')\ncf.window_close_windows('notepad-contents')\ncf.window_activate_and_maximize_windows("Python based RPA Development")\ncf.browser_navigate_h("https://sites.google.com/view/clointfusion-hackathon/date-with-clointfusion")\ncf.browser.scroll_down(1000)\ncf.browser_navigate_h('https://pypi.org')\ncf.browser_write_h("ClointFusion",User_Visible_Text_Element="Search projects")\ncf.browser_hit_enter_h()\ncf.pause_program(5)\ncf.browser_navigate_h('https://pypi.org/project/ClointFusion/')\ncf.browser.scroll_down(2000)\ncf.browser_mouse_click_h(element="RPA",double_click=True)\ncf.browser_navigate_h('https://dost.clointfusion.com/')\ncf.pause_program(10)\ncf.browser_quit_h()""")
            
            if not tour: # Self test for browser
                if os_name == windows_os:
                    if browser_activate('https://google.com'):
                        text_to_speech("Thanks to internet, we are indexed, on all search engines, just type clointfusion hackathon", show=False)
                        browser_write_h('clointfusion hackathon')
                        pause_program(3)
                        browser_hit_enter_h()
                        browser_mouse_click_h("Python based RPA Development Platform")
                        text_to_speech("Ever wanted to simply copy whole page and save to notepad? You can simply do it with one scraping function. Let me show you", show=False)
                        folder_create(os.path.join(test_folder_path,'Screen_scrape'))
                        scrape_save_contents_to_notepad(test_folder_path / 'Screen_scrape', switch_to_window="Python based")
                        text_to_speech("Its done already,", show=False)
                        os.startfile(os.path.join(test_folder_path,'Screen_scrape'))
                        text_to_speech("Its in this folder and saved as notepad-contents, let me open the file.", show=False)
                        os.startfile(os.path.join(test_folder_path,'Screen_scrape','notepad-contents.txt'))
                        pg.scroll(10)
                        pause_program(3)
                        pg.scroll(10)
                        text_to_speech("Ok, Lets get back to the web.", show=False)
                        window_close_windows('Screen_scrape')
                        window_close_windows('notepad-contents')
                        window_activate_and_maximize_windows("Python based RPA Development")
                        text_to_speech("Do you love what we do, Interested in joining our workforce, lets, have a Date..", show=False)
                        browser_navigate_h("https://sites.google.com/view/clointfusion-hackathon/date-with-clointfusion")
                        text_to_speech("Date with ClointFusion, is an initiative, for fast track entry, into our growing workforce.", show=False)
                        browser.scroll_down(1000)
                        text_to_speech("Let me, take you to our pypi page.", show=False)
                        browser_navigate_h('https://pypi.org')
                        browser_write_h("ClointFusion",User_Visible_Text_Element="Search projects")
                        browser_hit_enter_h()
                        pause_program(3)
                        browser_navigate_h('https://pypi.org/project/ClointFusion/')
                        browser_mouse_click_h(element="RPA",double_click=True)
                        browser.scroll_down(2000)
                        text_to_speech("Here, You can read, our detailed documentation, and view, our well, explained gifs.", show=False)
                        text_to_speech("Are you worried, that you dont know how to code, so you cannot automate", show=False)
                        text_to_speech("Are you just bored, to copy paste the same syntax twice, and thrice", show=False)
                        text_to_speech("ClointFusion, got you covered.", show=False)
                        browser_navigate_h('https://dost.clointfusion.com/')
                        text_to_speech("DOST, a block based approach to automate, powered by Clointfusion, just drag, and drop, the functions, and automate.", show=False)
                        pause_program(10)
                        text_to_speech("Just give it a try, and let us know your opinion via email to Clointfusion@cloint.com", show=False)
                        browser_quit_h()
                        text_to_speech("Is'nt interesting so far. But we still have so much, sit back and relax.", show=False)
                    else:
                        TEST_CASES_STATUS_MESSAGE  += "Helium package's Compatible Chrome is missing"
                else:
                    if browser_activate('https://google.com'):
                        text_to_speech("Thanks to internet, we are indexed, on all search engines, just type clointfusion hackathon", show=False)
                        browser_write_h('clointfusion hackathon')
                        pause_program(3)
                        browser_hit_enter_h()
                        browser_mouse_click_h("Python based RPA Development Platform")
                        text_to_speech("Do you love what we do? Interested in joining our workforce, lets, have a Date", show=False)
                        browser_navigate_h("https://sites.google.com/view/clointfusion-hackathon/date-with-clointfusion")
                        text_to_speech("Date with ClointFusion, is an initiative, for fast track entry, into our growing workforce.", show=False)
                        browser.scroll_down(1000)
                        text_to_speech("Let me, take you to our pypi page.", show=False)
                        browser_navigate_h('https://pypi.org')
                        browser_write_h("ClointFusion",User_Visible_Text_Element="Search projects")
                        browser_hit_enter_h()
                        pause_program(3)
                        browser_navigate_h('https://pypi.org/project/ClointFusion/')
                        browser_mouse_click_h(element="RPA",double_click=True)
                        browser.scroll_down(2000)
                        text_to_speech("Here, You can read, our detailed documentation, and view, our well, explained gifs.", show=False)
                        text_to_speech("Are you worried, that you dont know how to code, so you cannot automate", show=False)
                        text_to_speech("Are you just bored, to copy paste the same syntax twice, and thrice", show=False)
                        text_to_speech("ClointFusion, got you covered.", show=False)
                        browser_navigate_h('https://dost.clointfusion.com/')
                        text_to_speech("DOST, a block based approach to automate, powered by Clointfusion, just drag, and drop, the functions, and automate.", show=False)
                        pause_program(10)
                        text_to_speech("Just give it a try, and let us know your opinion via email to Clointfusion@cloint.com", show=False)
                        browser_quit_h()
                        text_to_speech("Is'nt interesting so far. But we still have so much, sit back and relax.", show=False)
                    else:
                        TEST_CASES_STATUS_MESSAGE  += "Helium package's Compatible Chrome is missing"
                print()
                print('Finished testing Browser operations...'+show_emoji())
                print("\n____________________________________________________________\n")
                logging.info("Tested Browser's Helium functions successfully")
            else: # Guided Tour
                if os_name == windows_os:
                    if browser_activate('https://google.com'):
                        browser_write_h('clointfusion hackathon')
                        pause_program(3)
                        browser_hit_enter_h()
                        browser_mouse_click_h("Python based RPA Development Platform")
                        text_to_speech("I will do the screen scraping now.", show=False)
                        folder_create(os.path.join(test_folder_path,'Screen_scrape'))
                        scrape_save_contents_to_notepad(test_folder_path / 'Screen_scrape', switch_to_window="Python based")
                        os.startfile(os.path.join(test_folder_path,'Screen_scrape'))
                        text_to_speech("Its in this folder and saved as notepad-contents, let me open the file.", show=False)
                        os.startfile(os.path.join(test_folder_path,'Screen_scrape','notepad-contents.txt'))
                        pause_program(3)
                        text_to_speech("Ok, Lets get back to the web.", show=False)
                        window_close_windows('Screen_scrape')
                        window_close_windows('notepad-contents')
                        window_activate_and_maximize_windows("Python based RPA Development")
                        browser_navigate_h("https://sites.google.com/view/clointfusion-hackathon/date-with-clointfusion")
                        browser.scroll_down(1000)
                        browser_navigate_h('https://pypi.org')
                        browser_write_h("ClointFusion",User_Visible_Text_Element="Search projects")
                        browser_hit_enter_h()
                        pause_program(3)
                        browser_navigate_h('https://pypi.org/project/ClointFusion/')
                        browser_mouse_click_h(element="RPA",double_click=True)
                        browser.scroll_down(2000)
                        text_to_speech("Here, You can read, our detailed documentation, and view, our well, explained gifs.", show=False)
                        browser_navigate_h('https://dost.clointfusion.com/')
                        text_to_speech("DOST, a block based approach to automate, powered by Clointfusion, just drag, and drop, the functions, and automate.")
                        pause_program(10)
                        text_to_speech("Just give it a try, and let us know your opinion via email to Clointfusion@cloint.com")
                        browser_quit_h()
                        print("\n____________________________________________________________\n")
                        text_to_speech("Let's go on to the next section.")
                        print("\n____________________________________________________________\n")
                    else:
                        text_to_speech("Sorry, experiencing some trouble.")
                        text_to_speech("Try, again later.")
                        pass
                else:
                    if browser_activate('https://google.com'):
                        browser_write_h('clointfusion hackathon')
                        pause_program(3)
                        browser_hit_enter_h()
                        browser_mouse_click_h("Python based RPA Development Platform")
                        browser_navigate_h("https://sites.google.com/view/clointfusion-hackathon/date-with-clointfusion")
                        browser.scroll_down(1000)
                        browser_navigate_h('https://pypi.org')
                        browser_write_h("ClointFusion",User_Visible_Text_Element="Search projects")
                        browser_hit_enter_h()
                        pause_program(3)
                        browser_navigate_h('https://pypi.org/project/ClointFusion/')
                        browser_mouse_click_h(element="RPA",double_click=True)
                        browser.scroll_down(2000)
                        text_to_speech("Here, You can read, our detailed documentation, and view, our well, explained gifs.", show=False)
                        browser_navigate_h('https://dost.clointfusion.com/')
                        text_to_speech("DOST, a block based approach to automate, powered by Clointfusion, just drag, and drop, the functions, and automate.")
                        pause_program(10)
                        text_to_speech("Just give it a try, and let us know your opinion via email to Clointfusion@cloint.com")
                        browser_quit_h()
                        print("\n____________________________________________________________\n")
                        text_to_speech("Let's go on to the next section.")
                        print("\n____________________________________________________________\n")
                    else:
                        text_to_speech("Sorry, experiencing some trouble.")
                        text_to_speech("Try, again later.")
                        pass
      
        except Exception as ex:
            selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
            print("ClointFusion Automated Testing Failed for Browser Functions"+str(ex))
            logging.info("Error while Testing Browser Helium functions="+str(ex))
            key_press(key_1="ctrl", key_2="w") #to close any open browser
            TEST_CASES_STATUS_MESSAGE  += "Error while Testing Browser Helium functions="+str(ex)
     
# Windows Operations
        if os_name == windows_os:
            try:
                # Description
                if not tour:
                    print()
                    print("\n____________________________________________________________\n")
                    print('Started testing Windows operations...')
                    text_to_speech("This PC is powered by Windows. Windows PCs are the best for automation.", show=False)
                    text_to_speech("ClointFusion has six window-specific functions.", show=False)
                    message_pop_up("\nWindows\n```````````\n1)  window_show_desktop\n2)  window_get_all_opened_titles_windows\n3)  window_activate_and_maximize_windows\n4)  window_minimize_windows\n5)  window_close_windows\n6)  launch_any_exe_bat_application\n\n", delay=5)
                    text_to_speech("These functions help you to perform high-level automation in Windows. Let me show you while I test them.", show=False)
                else:
                    print("\n____________________________________________________________\n")
                    text_to_speech("ClointFusion currently has six, windows specific functions.")
                    message_pop_up("\nWindows\n```````````\n1)  window_show_desktop\n2)  window_get_all_opened_titles_windows\n3)  window_activate_and_maximize_windows\n4)  window_minimize_windows\n5)  window_close_windows\n6)  launch_any_exe_bat_application\n\n", delay=5)
                    print("\nWindows\n```````````\n1)  window_show_desktop\n2)  window_get_all_opened_titles_windows\n3)  window_activate_and_maximize_windows\n4)  window_minimize_windows\n5)  window_close_windows\n6)  launch_any_exe_bat_application\n\n")
                
                # Actions
                if tour:
                    print("""\nCommands I will be using in the next scenario:\n~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~\nimport ClointFusion as cf\ncf.launch_any_exe_bat_application("write")\ncf.pause_program(2)\ncf.window_minimize_windows('Document')\ncf.pause_program(2)\ncf.window_activate_and_maximize_windows('Document')\ncf.pause_program(2)\nwindow_names = cf.window_get_all_opened_titles_windows()\ncf.window_close_windows('Document')""")
                text_to_speech("Show Desktop", show=False)
                message_pop_up("\nCommand being used :\n```````````\ncf.window_show_desktop()\n\n", delay=3)
                window_show_desktop()
                text_to_speech("want to launch Launch any exe or application", show=False)
                message_pop_up("\nCommand being used :\n```````````\ncf.launch_any_exe_bat_application('write')\n\n", delay=3)
                launch_any_exe_bat_application("write")
                text_to_speech("Minimize window", show=False)
                pause_program(2)
                message_pop_up("\nCommand being used :\n`````````````\ncf.window_minimize_windows('Document')\n\n", delay=3)
                window_minimize_windows('Document')
                text_to_speech("Maximize window", show=False)
                pause_program(2)
                message_pop_up("\nCommand being used :\n```````````\ncf.window_activate_and_maximize_windows('Document')\n\n", delay=5)
                window_activate_and_maximize_windows('Document')
                pause_program(2)
                text_to_speech("Get all window names", show=False)
                message_pop_up("\nCommand being used :\n```````````\ncf.window_get_all_opened_titles_windows()\n\n", delay=3)
                window_names = window_get_all_opened_titles_windows()
                message_pop_up("Window Names\n``````````````" + str("\n"+"\n".join(window_names)))
                text_to_speech("Close window", show=False)
                message_pop_up("\nCommand being used :\n```````````\ncf.window_close_windows('Document')\n\n", delay=3)
                window_close_windows('Document')
                if not tour:
                    text_to_speech("See, its that easy, Automation is now right in your hands.", show=False)
                    print()
                    print('Finsihed testing Windows operations. '+show_emoji())
                    print("\n____________________________________________________________\n")
                    logging.info('Window based operations tested successfully')
                else:
                    print("\n____________________________________________________________\n")
                    text_to_speech("Let's go on to the next section.")
                    print("\n____________________________________________________________\n")

            except Exception as ex:
                selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
                print("ClointFusion Automated Testing Failed for Windows Functions"+str(ex))
                logging.info('Error while testing window based operations='+str(ex))
                TEST_CASES_STATUS_MESSAGE  += 'Error while testing window based operations='+str(ex)

# String operations
        try: 
            if not tour:
                print()
                print("\n____________________________________________________________\n")
                text_to_speech("Testing String manipulation functions.", show=False)
                print('Started testing String Operations...')
            else:
                print("\n____________________________________________________________\n")
                text_to_speech("ClointFusion currently has three, string manipulation functions.")
                message_pop_up("\nStrings\n```````````\n1)  string_extract_only_alphabets\n2)  string_extract_only_numbers\n3)  string_remove_special_characters\n\n", delay=5)
                print("\nStrings\n```````````\n1)  string_extract_only_alphabets\n2)  string_extract_only_numbers\n3)  string_remove_special_characters\n")
            
            if tour:
                print("""\nCommands I will be using in the next scenario:\n~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~\nimport ClointFusion as cf\ncf.string_remove_special_characters("C!@loin#$tFu*(sion"))\ncf.string_extract_only_alphabets(inputString="C1l2o#%^int&*Fus12i5on"))\ncf.string_extract_only_numbers("C1l2o3i4n5t6F7u8i9o0n"))""")
            print(string_remove_special_characters("C!@loin#$tFu*(sion"))
            print(string_extract_only_alphabets(inputString="C1l2o#%^int&*Fus12i5on"))
            print(string_extract_only_numbers("C1l2o3i4n5t6F7u8i9o0n"))
            if not tour:
                print('String operations tested successfully '+show_emoji())
                print("\n____________________________________________________________\n")
                text_to_speech("String operations tested successfully.", show=False)
                logging.info('String operations tested successfully')
            else:
                print("\n____________________________________________________________\n")
                text_to_speech("Let's go on to the next section.")
                print("\n____________________________________________________________\n")
        except Exception as ex:
            selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
            print('Error while testing string operations='+str(ex))
            logging.info('Error while testing string operations='+str(ex))
            TEST_CASES_STATUS_MESSAGE  += "Error while testing string operations="+str(ex)
        
# Folder Operations
        try:
            if not tour:
                print()
                print("\n____________________________________________________________\n")
                text_to_speech("Let me test folder operations", show=False)
                print('Testing folder operations')
            else:
                print("\n____________________________________________________________\n")
                text_to_speech("ClointFusion currently has eight, folder and file related functions.")
                message_pop_up("\nFolders\n```````````\n1)  folder_read_text_file\n2)  folder_write_text_file\n3)  folder_create\n4)  folder_create_text_file\n5)  folder_get_all_filenames_as_list\n6)  folder_delete_all_files\n7)  file_rename\n8)  file_get_json_details\n\n", delay=5)
                print("\nFolders\n```````````\n1)  folder_read_text_file\n2)  folder_write_text_file\n3)  folder_create\n4)  folder_create_text_file\n5)  folder_get_all_filenames_as_list\n6)  folder_delete_all_files\n7)  file_rename\n8)  file_get_json_details\n\n")
               
            if tour:
                print("""\nCommands I will be using in the next scenario:\n~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~\nimport ClointFusion as cf\ncf.folder_create(Path(cf.os.path.join(test_folder_path,"My Test Folder")))\ncf.os.startfile(cf.os.path.join(test_folder_path))\ncf.pause_program(1)\ncf.folder_create_text_file(test_folder_path, "My Text File")\ncf.pause_program(1)\ncf.excel_create_excel_file_in_given_folder(test_folder_path,'Quick_Self_Test_Excel')\ncf.pause_program(1)\ncf.excel_create_excel_file_in_given_folder(test_folder_path,'My Excel-1')\ncf.pause_program(1)\ncf.excel_create_excel_file_in_given_folder(test_folder_path,'My Excel-2')\ncf.excel_create_excel_file_in_given_folder(cf.os.path.join(test_folder_path,"Delete Excel"),'Delete-Excel-1')\ncf.pause_program(2)\ncf.excel_create_excel_file_in_given_folder(cf.os.path.join(test_folder_path,"Delete Excel"),'Delete-Excel-2')\ncf.pause_program(2)\ncf.folder_delete_all_files(cf.os.path.join(test_folder_path,'Delete Excel'), "xlsx")\ncf.folder_create(Path(test_folder_path / 'Split_Merge'))\ncf.pause_program(2)\ncf.print(folder_get_all_filenames_as_list(test_folder_path))\ncf.pause_program(2)\ncf.print(folder_get_all_filenames_as_list(test_folder_path, extension="xlsx"))\n""")
            if os_name == windows_os:
                folder_create(Path(os.path.join(test_folder_path,"My Test Folder")))
                os.startfile(os.path.join(test_folder_path))
                text_to_speech("Watch this folder carefully, i will create random folders, text files, excel files.", show=False)
                pause_program(1)
                folder_create_text_file(test_folder_path, "My Text File")
                pause_program(1)
                excel_create_excel_file_in_given_folder(test_folder_path,'Quick_Self_Test_Excel')
                pause_program(1)
                excel_create_excel_file_in_given_folder(test_folder_path,'My Excel-1')
                pause_program(1)
                excel_create_excel_file_in_given_folder(test_folder_path,'My Excel-2')
            else:
                folder_create(Path(os.path.join(test_folder_path,"My Test Folder")))
                folder_create_text_file(test_folder_path, "My Text File")
                pause_program(1)
                excel_create_excel_file_in_given_folder(test_folder_path,'Quick_Self_Test_Excel')
                pause_program(1)
                excel_create_excel_file_in_given_folder(test_folder_path,'My Excel-1')
                pause_program(1)
                excel_create_excel_file_in_given_folder(test_folder_path,'My Excel-2')

            try:
                excel_create_excel_file_in_given_folder(os.path.join(test_folder_path,"Delete Excel"),'Delete-Excel-1')
                pause_program(2)
                excel_create_excel_file_in_given_folder(os.path.join(test_folder_path,"Delete Excel"),'Delete-Excel-2')
                pause_program(2)
                folder_delete_all_files(os.path.join(test_folder_path,'Delete Excel'), "xlsx")
            except Exception as ex:
                selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
                print('Unable to delete files in test folder='+str(ex))
                logging.info('Unable to delete files in test folder='+str(ex))
                TEST_CASES_STATUS_MESSAGE  += 'Unable to delete files in test folder='+str(ex)

            folder_create(Path(test_folder_path / 'Split_Merge'))
            pause_program(2)
            print(folder_get_all_filenames_as_list(test_folder_path))
            pause_program(2)
            print(folder_get_all_filenames_as_list(test_folder_path, extension="xlsx"))
            if os_name == windows_os:
                window_close_windows("ClointFusion_Self_Tests")
            if not tour:
                text_to_speech("Folder operations test is successful", show=False)
                print("\n")
                print('Folder operations tested successfully. '+show_emoji())
                print("\n____________________________________________________________\n")
                logging.info('Folder operations tested successfully')
            else:
                print("\n____________________________________________________________\n")
                text_to_speech("Let's go on to the next section.")
                print("\n____________________________________________________________\n")
        except Exception as ex:
            selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
            print('Error while testing Folder operations='+str(ex))
            logging.info('Error while testing Folder operations='+str(ex))
            TEST_CASES_STATUS_MESSAGE  += 'Error while testing Folder operations='+str(ex)   

# Excel operations
        if not tour:     
            try: 
                print()
                print("\n____________________________________________________________\n")
                text_to_speech("Let me test the excel compatability.", show=False)
                print('Started testing excel operations...')

                excel_create_excel_file_in_given_folder(test_folder_path, "Test_Excel_File", "Test_Sheet")
                print(excel_get_row_column_count(test_run_excel_path))
                excel_create_excel_file_in_given_folder(test_folder_path,excelFileName="Excel_Test_Data")
                test_excel_path = test_folder_path / "Excel_Test_Data.xlsx"
                excel_set_single_cell(test_excel_path,columnName="Name",cellNumber=0,setText="A")
                excel_set_single_cell(test_excel_path,columnName="Name",cellNumber=1,setText="B")
                excel_set_single_cell(test_excel_path,columnName="Name",cellNumber=2,setText="C")
                excel_set_single_cell(test_excel_path,columnName="Name",cellNumber=3,setText="D")
                excel_set_single_cell(test_excel_path,columnName="Name",cellNumber=4,setText="E")
                excel_set_single_cell(test_excel_path,columnName="Age",cellNumber=0,setText="1")
                excel_set_single_cell(test_excel_path,columnName="Age",cellNumber=1,setText="2")
                excel_set_single_cell(test_excel_path,columnName="Age",cellNumber=2,setText="4")
                excel_set_single_cell(test_excel_path,columnName="Age",cellNumber=3,setText="3")
                excel_set_single_cell(test_excel_path,columnName="Age",cellNumber=4,setText="5")
                print(excel_get_single_cell(test_excel_path,sheet_name='Sheet1',columnName='Name'))
                excel_create_file(test_folder_path,"My New Paste Excel")
                excel_create_excel_file_in_given_folder(test_folder_path,'My Excel-3','CF-Sheet-1')
                excel_file_path = test_folder_path / 'My Excel-3.xlsx'
                print(excel_get_all_sheet_names(excel_file_path))
                print(excel_get_all_sheet_names(test_run_excel_path))
                excel_copied_Data=excel_copy_range_from_sheet(test_excel_path, sheet_name="Sheet1", startCol=1, startRow=1, endCol=2, endRow=6)
                print(excel_copied_Data)
                excel_copy_paste_range_from_to_sheet(Path(os.path.join(test_folder_path,"My New Paste Excel.xlsx")), sheet_name="Sheet1", startCol=1, startRow=1, endCol=2, endRow=6, copiedData=excel_copied_Data)
                excel_split_by_column(excel_path=Path(os.path.join(test_folder_path,"My New Paste Excel.xlsx")), sheet_name="Sheet1", header=0, columnName="Name")
                folder_create(Path(test_folder_path / 'Split_Merge'))
                excel_split_the_file_on_row_count(excel_path=Path(test_folder_path / "My New Paste Excel.xlsx"), sheet_name="Sheet1", rowSplitLimit=1, outputFolderPath=os.path.join(test_folder_path,'Split_Merge'), outputTemplateFileName="Split")
                
                excel_merge_all_files(input_folder_path=test_folder_path / "Split_Merge", output_folder_path=Path(test_folder_path,'Split_Merge'))
                
                excel_drop_columns(Path(test_folder_path / "My New Paste Excel.xlsx"), columnsToBeDropped ="Age")
                excel_sort_columns(excel_path=test_excel_path, sheet_name="Sheet1", header=0, firstColumnToBeSorted="Age", secondColumnToBeSorted="Name")
                excel_clear_sheet(Path(test_folder_path / "My New Paste Excel.xlsx"), sheet_name="Sheet1", header=0)
                excel_set_single_cell(test_excel_path,columnName="Name",cellNumber=5,setText="E")
                excel_set_single_cell(test_excel_path,columnName="Age",cellNumber=5,setText="5")
                excel_remove_duplicates(excel_path=test_excel_path, sheet_name="Sheet1", header=0,columnName="Name", which_one_to_keep="first")
                excel_create_file(test_folder_path,"My VLookUp Excel")
                excel_set_single_cell(Path(test_folder_path,"My VLookUp Excel.xlsx"),columnName="Name",cellNumber=0,setText="Avinash")
                excel_set_single_cell(Path(test_folder_path,"My VLookUp Excel.xlsx"),columnName="Name",cellNumber=1,setText="Anil")
                excel_set_single_cell(Path(test_folder_path,"My VLookUp Excel.xlsx"),columnName="Name",cellNumber=2,setText="Karthick")
                excel_set_single_cell(Path(test_folder_path,"My VLookUp Excel.xlsx"),columnName="Name",cellNumber=3,setText="Prashant")
                excel_set_single_cell(Path(test_folder_path,"My VLookUp Excel.xlsx"),columnName="Name",cellNumber=4,setText="Fharook")
                excel_set_single_cell(Path(test_folder_path,"My VLookUp Excel.xlsx"),columnName="Salary",cellNumber=0,setText="20000")
                excel_set_single_cell(Path(test_folder_path,"My VLookUp Excel.xlsx"),columnName="Salary",cellNumber=1,setText="50000")
                excel_set_single_cell(Path(test_folder_path,"My VLookUp Excel.xlsx"),columnName="Salary",cellNumber=2,setText="40000")
                excel_set_single_cell(Path(test_folder_path,"My VLookUp Excel.xlsx"),columnName="Salary",cellNumber=3,setText="30000")
                excel_set_single_cell(Path(test_folder_path,"My VLookUp Excel.xlsx"),columnName="Salary",cellNumber=4,setText="5000")
                excel_vlook_up(filepath_1=test_excel_path,filepath_2=Path(test_folder_path,"My VLookUp Excel.xlsx"),match_column_name="Name")
                if os_name == windows_os:
                    os.startfile(Path(test_folder_path,"My VLookUp Excel.xlsx"))
                    pause_program(10)
                    window_close_windows("My VLookUp Excel")
                    try:
                        window_close_windows("My VLookUp Excel.xlsx")
                    except:
                        pass

                print('Excel operations tested successfully. '+show_emoji())
                text_to_speech("Excel operations tested successfully.", show=False)
                print("\n____________________________________________________________\n")
                logging.info('Excel operations tested successfully')
                
                text_to_speech("Isn't amazingly quick, But yeah, more than 28 functions, and, more than 50 excel operations, have been tested just now.", show=False)
                text_to_speech("Banking, finance, from data collection, to data cleaning, and sending reports, everything can be done, with excel functions, and Clointfusion. Give it a try today.", show=False)
            except Exception as ex:
                selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
                print("Error while testing Excel Operations="+str(ex))
                logging.info("Error while testing Excel Operations="+str(ex))
                TEST_CASES_STATUS_MESSAGE  += "Error while testing Excel Operations="+str(ex)
        else:
            pass

# Message operations       
        try:
            if not tour:
                print()
                print("\n____________________________________________________________\n")
                text_to_speech("Ok, Let me show you some different message pop ups.", show=False)
                print("Started testing message functions...")
            else:
                print("\n____________________________________________________________\n")
                text_to_speech("ClointFusion currently has four, types of message pop ups.")
                message_pop_up("\nMessages\n```````````\n1)  message_counter_down_timer\n2)  message_pop_up\n3)  message_flash\n4)  message_toast\n\n", delay=5)
                print("\nMessages\n```````````\n1)  message_counter_down_timer\n2)  message_pop_up\n3)  message_flash\n4)  message_toast\n\n")
            
            if tour:
                print("""\nCommands I will be using in the next scenario:\n~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~\nimport ClointFusion as cf\ncf.message_counter_down_timer("Message with timer (in seconds)",6)\ncf.message_pop_up("Testing popup message.")\ncf.message_flash("Testing flash message.")""")

            text_to_speech("Message with timer...", show=False)
            message_counter_down_timer("Message with timer (in seconds)",6)
            text_to_speech("Popup Message...", show=False)
            message_pop_up("Testing popup message.")
            text_to_speech("Flash Message...", show=False)
            message_flash("Testing flash message.")

            if not tour:
                print()
                print('Message operations tested successfully. '+show_emoji())
                text_to_speech("Message functions tested successfully", show=False)
                print("\n____________________________________________________________\n")
                logging.info('Message functions tested successfully')
            else:
                print("\n____________________________________________________________\n")
                text_to_speech("Let's go on to the next section.")
                print("\n____________________________________________________________\n")
                
        except Exception as ex:
            selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
            print("Error while testing Flash message="+str(ex))
            logging.info("Error while testing Flash message="+str(ex))
            TEST_CASES_STATUS_MESSAGE  += "Error while testing Flash message="+str(ex)
   
# CLI operations
        try:
            if not tour:
                print()
                print("\n____________________________________________________________\n")
                text_to_speech("Nothing can beat a CLI command, in convenience, or effectivity.", show=False)
                text_to_speech("Let me show you, some terminal commands, offered by ClointFusion.", show=False)
                print('Started testing CLIs Operations...')
            else:
                print("\n____________________________________________________________\n")
                text_to_speech("ClointFusion currently has, more than ten, CLI commands.")
                message_pop_up("\nCLI\n```````````\n1)  cf\n2)  bol\n3)  dost\n4)  cf_tray\n5)  cf_st\n6)  cf_work\n7)  cf_wm\n8)  cf_sm\n9)  cf_like\n10)  cf_py\n11)  cf_tour\n12)  colab\n13)  cf_vlookup\n\n", delay=5)
                
            
            if tour:
                print("""\nCommands I will be using in the next scenario in terminal :\n~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~\n cf \n cf_st \n cf_work \n""")
            
            text_to_speech("Press CF in terminal to get all the available commands", show=False)
            print("\n_________________\n")
            print("Press 'cf' in terminal for this function.\n")
            #Test work
            cli_cf_test()
            
            try:
                text_to_speech("Want to quickly test your internet speed, type CF underscore ST.", show=False)
                print("\n_________________\n")
                print("Type 'cf_st' in terminal for this function")
                
                # Test Speed Test
                if os_name == windows_os:
                    text_to_speech("Internet speed is being tested in the terminal. Let me show you..", show=False)
                    window_activate_and_maximize_windows(console_window_name)
                    cli_speed_test_test()
                    pause_program(5)
                    window_minimize_windows(console_window_name)
                    window_activate_window("Welcome to ClointFusion - Made in India with LOVE")
                    print("\n_________________\n")
                    text_to_speech("I am back !!!", show=False)
                else:
                    cli_speed_test_test()
            except: # Dont want to cancel self-test if speed test fails
                pass
            
            text_to_speech("Want to see, how much time you spent, on what, and which application. type cf_work in terminal. But First, let me tell you this, your data is stored locally and only belongs to you.", show=False)
            print("\n_________________\n")
            print("Type 'cf_work' in terminal for this function.")
            # Test BRE WHM
            cli_bre_whm_test()
            text_to_speech("This Report looks colorful and neat in a terminal, give it a try, after self-test", show=False)
            
            if not tour:
                print()
                print("CLI's functions tested successfully. "+show_emoji())
                print("\n____________________________________________________________\n")
                text_to_speech("CLI's functions tested successfully", show=False)
                logging.info("CLI functions tested successfully")  
        except Exception as ex:
            selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
            print("Error while Testing CLIs")
            logging.info("Error while Testing CLIs")
            TEST_CASES_STATUS_MESSAGE  += "Error while Testing CLIs"

# Register User
        if not tour:
            try:
                file_contents = ''
                try:
                    with open(log_file_path, encoding="utf-8") as f:
                        file_contents = f.readlines()
                except Exception as ex:
                    selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
                    file_contents = 'Unable to read the file' + str(ex)
                    TEST_CASES_STATUS_MESSAGE += "Unable to read log file" + str(ex)
                
                try:
                    text_to_speech("Let me quickly, do the necessary registration, for you. So you can, get started with automation", show=False)
                    time_taken= timedelta(seconds=time.monotonic()  - start_time)
                    os_hn_ip = "OS:{}".format(os_name) + "HN:{}".format(socket.gethostname()) + ",IP:" + str(socket.gethostbyname(socket.gethostname())) + "/" + str(get_public_ip())
                except Exception as ex:
                    selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
                    TEST_CASES_STATUS_MESSAGE += str(ex)
                
                try:
                    driver = selft.gf(os_hn_ip, time_taken, file_contents)
                    cursr.execute("UPDATE CF_IMP_VALUES set SELF_TEST = 'True' where ID = 1")
                    connct.commit()
                    clear_screen()
                    time.sleep(5)
                    selft.ast()
                    text_to_speech("Closing the browser now.", show=False)
                    browser.set_driver(driver)
                    browser_quit_h()
                    time.sleep(2)
                    
                    SUCCESS = True
                    time.sleep(5)
                except Exception as ex:
                    selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
                    TEST_CASES_STATUS_MESSAGE += str(ex)
                
                if str(TEST_CASES_STATUS_MESSAGE).strip()  == "":
                    print("")
                    logging.info("ClointFusion Self Testing Completed")
                    print("Congratulations - ClointFusion is compatible with your computer " + show_emoji('clap') + show_emoji('clap'))
                    message_pop_up("Congratulations !!!\n\nClointFusion is compatible with your computer settings")
                    print("\n____________________________________________________________\n")
                    text_to_speech("Self Test is Completed. And i am happy to say, your PC is compatible with ClointFusion", show=False)
                    text_to_speech("An email with this self-test report is being sent to the registered email address.", show=False)
                    
            except Exception as ex:
                selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
                print("ClointFusion Automated Testing Failed "+str(ex))
                logging.info("ClointFusion Automated Testing Failed "+str(ex))
                TEST_CASES_STATUS_MESSAGE  += "ClointFusion Automated Testing Failed "+str(ex) 
        else:
            tour_comp = True
                
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        print("ClointFusion Automated Testing Failed "+str(ex))
        logging.info("ClointFusion Automated Testing Failed "+str(ex))
        TEST_CASES_STATUS_MESSAGE  += "ClointFusion Automated Testing Failed "+str(ex)

# End of Self Test
    finally:
        enable_semi_automatic_mode = False
        if not tour:
            if str(TEST_CASES_STATUS_MESSAGE).strip()  == "":
                text_to_speech("All set and ready to go. Wanna to call me again, type B O L, in terminal", show=False)
                print("Type 'bol'")
                text_to_speech("Always remember, Think what to automate, not how to, because, Clointfusion is there to do it.", show=False)
                message_toast("ClointFusion is compatible with your computer's settings !", website_url="https://tinyurl.com/ClointFusion")
            else:
                print("ClointFusion Self Testing has Failed for few Functions")
                print(TEST_CASES_STATUS_MESSAGE)
                logging.info("ClointFusion Self Testing has Failed for few Functions")
                logging.info(TEST_CASES_STATUS_MESSAGE)
                SUCCESS = False

            if os_name == windows_os:
                try:
                    window_close_windows('Welcome to ClointFusion - Made in India with LOVE')
                except:
                    text_to_speech("Please click on close button.", show=False)
                    print("Please click red 'Close' button")
            else:
                print("Please click red 'Close' button")
            return TEST_CASES_STATUS_MESSAGE, SUCCESS
        else:
            text_to_speech("I hope you enjoyed our tour and learned something new", show=False)
            text_to_speech("If you want to repeat, press the start button. Otherwise, click the close button.", show=False)
            print("\n____________________________________________________________\n")
            text_to_speech("I printed out all of the syntax I used on this tour")
            text_to_speech("Feel free to scroll through and take notes.")
            print("\n____________________________________________________________\n")
            text_to_speech("There are plenty other functions to explore. ")
            text_to_speech("For example, we have a large selection of, Excel functions as well as, some of the top utility functions.")
            text_to_speech("We hope that, this tour has sparked your interest, in learning more about automation and ClointFusion.")
            print("\n____________________________________________________________\n")
            return tour_comp, tour_comp



def clointfusion_self_test(tour=False):
    global os_name, python_exe_path
    WHILE_TRUE = True #Colab Settings
    start_time = time.monotonic()
    last_updated_on_month = ""
    tour_comp = False
    status_msg, success = "", ""
    console_window = ""
    if os_name == windows_os:
        console_window = window_get_active_window()
        if "command" in console_window:
            console_window = "command"
    try:
    
# Layout and data
        if not tour:
            layout = [ [sg.Text("ClointFusion's Automated Compatibility Self-Test",justification='c',font='Courier 18',text_color='orange')],
                    [sg.Button("Sign-In With Google", key='SSO', tooltip='Sign-In with Gmail ID')],
                    [sg.Text("Thanks for improving ClointFusion by continuing with this self-test.",justification='c',text_color='yellow',font='Courier 12')],
                    [sg.Text('Its highly recommended to close all open files/folders/browsers before running this self test.',size=(0, 1),justification='l',text_color='red',font='Courier 12')],
                    [sg.Text('This Automated Self Test, takes around 4-5 minutes...Kindly do not move the mouse or type anything.',size=(0, 1),justification='l',text_color='red',font='Courier 12')],
                    [sg.Output(size=(140,20), key='-OUTPUT-')],
                    [sg.Button('Start',bind_return_key=True,button_color=('white','green'),font='Courier 14',disabled=True, tooltip='Sign-In with Gmail to Enable this button'), sg.Button('Close',button_color=('white','firebrick'),font='Courier 14', tooltip='Close this window & exit')],
                    [sg.Button('Skip for Now',button_color=('white', 'orange'),font='Courier 14',disabled= False, tooltip=  'Click this button to skip Self-Test')]]
        else:
            layout = [ [sg.Text("ClointFusion's Automated Compatibility Self-Test",justification='c',font='Courier 18',text_color='orange')],
                    [sg.Button("Sign-In With Google", key='SSO', tooltip='Sign-In with Gmail ID')],
                    [sg.Text("Thanks for improving ClointFusion by continuing with this self-test.",justification='c',text_color='yellow',font='Courier 12')],
                    [sg.Text('Its highly recommended to close all open files/folders/browsers before running this self test.',size=(0, 1),justification='l',text_color='red',font='Courier 12')],
                    [sg.Text('This Automated Self Test, takes around 4-5 minutes...Kindly do not move the mouse or type anything.',size=(0, 1),justification='l',text_color='red',font='Courier 12')],
                    [sg.Output(size=(140,20), key='-OUTPUT-')],
                    [sg.Button('Start',bind_return_key=True,button_color=('white','green'),font='Courier 14',disabled=False, tooltip='Sign-In with Gmail to Enable this button'), sg.Button('Close',button_color=('white','firebrick'),font='Courier 14', tooltip='Close this window & exit')],
                    [sg.Button('Skip for Now',button_color=('white', 'orange'),font='Courier 14',disabled= False, tooltip=  'Click this button to skip Self-Test')]]
        
        
        
        if os_name == windows_os:
            window = sg.Window('Welcome to ClointFusion - Made in India with LOVE', layout, return_keyboard_events=True,use_default_focus=False,disable_minimize=True,grab_anywhere=False, disable_close=False,element_justification='c',keep_on_top=False,finalize=True,icon=cf_icon_file_path)
        else:
            try:
                window = sg.Window('Welcome to ClointFusion - Made in India with LOVE', layout, return_keyboard_events=True,use_default_focus=False,disable_minimize=False,grab_anywhere=False, disable_close=False,element_justification='c',keep_on_top=False,finalize=True,icon=cf_icon_file_path)
            except:
                WHILE_TRUE = False
        instructions = False
        name_st, _, _, _ = selft.get_details()

# Application Starts
        while WHILE_TRUE:
            if not instructions:
                if not tour:
                    text_to_speech("Thank you for downloading and installing the ClointFusion Python package.", show=False)    
                    text_to_speech("This is Bol, a ClointFusion-powered voice-based assistant.", show=False)    
                    text_to_speech("I'm here to give you a quick tour of the ClointFusion package, and to check your machine for compatibility.", show=False)
                    text_to_speech("Please sign in and confirm. Then press the start button.", show=False)
                else:
                    text_to_speech(f"Hi, {name_st}, Thanks for showing interest in ClointFusion's tour.", show=False)
                    text_to_speech("As you know, this is Bol, a ClointFusion-powered voice-based assistant.", show=False)
                    text_to_speech("Click on the start button to get started.", show=False)
                instructions = True
            
            event, _ = window.read()

# SSO    
            if event == 'SSO':
                selft.sso()
                window['Start'].update(disabled=False)
                window['SSO'].update(disabled=True)
                window['Skip for Now'].update(disabled=False)
                if os_name  == linux_os:
                    clear_screen()

# Skip
            if event == 'Skip for Now':
                try:
                    text_to_speech("You have chosen to skip ClointFusion's Self-Test. Some of the functions may not work properly", show=False)     
                    pg.alert("You have chosen to skip ClointFusion's Self-Test.\n\nSome of the functions may not work properly")
                    message_toast("ClointFusion Self-Test is Skipped")
                except:
                    put_text("You have chosen to skip ClointFusion's Self-Test.\n\nSome of the functions may not work properly !")

                last_updated_on_month = 2
                break

# Start
            if event == 'Start':
                window['-OUTPUT-'].update('')
                window['Start'].update(disabled=True)
                # window['Close'].update(disabled=True)
                window['Skip for Now'].update(disabled=True)
                _folder_write_text_file(os.path.join(clointfusion_directory,"Config_Files",'Running_ClointFusion_Self_Tests.txt'),str(True))
                _init_cf_quick_test_log_file(temp_current_working_dir)
                
                if not tour:
                    print("Starting ClointFusion's Automated Self Testing Module")
                    print('This may take several minutes to complete...')
                    print('During this test, some excel file, notepad, browser etc may be opened & closed automatically')
                    print('Please sitback & relax while all the test-cases are run...')
                    print()
                    text_to_speech(f"Thank you, {name_st}, for starting the self-test, and helping to improve ClointFusion.", show=False)
                    text_to_speech(f"This is a fully automatic self-test that requires no human intervention.", show=False)
                    text_to_speech(f"Sit back and relax. I'll notify you when I'm finished.", show=False)
                    status_msg, success = clointfusion_self_demo_tour(temp_current_working_dir, start_time, console_window, tour)

                    if str(status_msg).strip() == "" and success:
                        window['Close'].update(disabled=False)
                        break
                    else:
                        try:
                            pg.alert("Please resolve below errors and try again:\n\n" + status_msg)
                        except:
                            put_text("Please resolve below errors and try again:\n\n" + status_msg)
                        sys.exit(0)
                else:
                    text_to_speech("Welcome to ClointFusion's Automated Self Tour.")
                    text_to_speech('Please sitback & relax while i explain you the clointfusion functions.')
                    print()
                    tour_comp, _ = clointfusion_self_demo_tour(temp_current_working_dir, start_time, console_window, tour)
                    window['Close'].update(disabled=False)
                    window['Start'].update(disabled=False)

# Close
            if event in (sg.WINDOW_CLOSED, 'Close', sg.WIN_CLOSED):
                if tour and instructions and not tour_comp:
                    text_to_speech("Please, let us know, if there is anything further we can do, to improve. Please contact us at, clointfusion@cloint.com, if you have any questions.")
                    pause_program(10)
                break

# Exception
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        try:
            pg.alert('Error in Clointfusion Self Test = '+str(ex))
        except:
            print("Error in Clointfusion Self Test = " +str(ex))

        exc_type, exc_value, exc_tb = sys.exc_info()
        print(traceback.format_exception(exc_type, exc_value, exc_tb,limit=None, chain=True))
        _rerun_clointfusion_first_run(str(ex))

# Finally 
    finally:
        try:
            if not tour:
                if last_updated_on_month == 2 :
                    window.close()
                elif success:
                    try:
                        window.close()
                    except:
                        pass
                    pause_program(5)
                    if os_name == windows_os:
                        os.system(f'{python_exe_path} -i -c "import ClointFusion as cf; print(\'Awesome !!!, your now using the latest ClointFusion.\'); print(\'Try cf.browser_activate() \')"')
                        window_activate_and_maximize_windows(console_window)
                    else:
                        os.system(f'sudo python{python_version} -i -c "import ClointFusion as cf; print(\'Awesome !!!, your now using the latest ClointFusion.\'); print(\'Try cf.browser_activate() \')"')
                        time.sleep(2)
                else:
                    sys.exit(1)
            else:
                if tour_comp == True:
                    text_to_speech("Thank you very much. I hope you found this tour useful. You are welcome to return at any time. I will gladly give you the tour again.", show=False)
                    window.close()

        except Exception as ex:
            selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
            print(str(ex))

def update_log_excel_file(message=""):
    """
    Given message will be updated in the excel log file.

    Parameters:
        message  (str) : message to update.

    Retursn:
        returns a boolean true if updated sucessfully
    """
    global status_log_excel_filepath
    try:
        if not message:
            message = gui_get_any_input_from_user("message to Update Log file")

        df = pd.DataFrame({'Timestamp': [datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")], 'Status':[message]})
        append_df_to_excel(status_log_excel_filepath, df, index=False,startrow=None,header=None)

        return True
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in update_log_excel_file="+str(ex))
        return False

# --------- Self-Test Related Functions Ends ---------


# --------- CLI Commands ---------

CONTEXT_SETTINGS = dict(help_option_names=['-h', '--help'])
@click.command(context_settings=CONTEXT_SETTINGS)
def cli_speed_test():
    """CLI for testing internet bandwidth using speedtest.net"""
    try:
        if os_name != mac_os:
            print(os.system("speedtest-cli"))
        else:
            try:
                print(os.system("speedtest-cli"))
            except:
                print("This feature is curently not supported on macOS. Please contribute to make the tomorrow better.")
                print(random.choice(contribution_messages))
                selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in cli_speed_test="+str(ex))

@click.command(context_settings=CONTEXT_SETTINGS)
def cli_auto_liker():
    """CLI for auto liking CF's specific posts on Socail Media"""
    try:
        if os_name == windows_os:
            cmd = f'{python_exe_path} "{site_packages_path}\ClointFusion\cf_auto_liker.py"'
            os.system(cmd)
        else:
            print("Auto Liker option is only available on Windows. We regret the inconvenience.")
            print(f"Please contribute to make this feature available on {os_name.upper()} system.")
            print(random.choice(contribution_messages))
            selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
            
            
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in cli_speed_test="+str(ex))

@click.command(context_settings=CONTEXT_SETTINGS)
def cli_colab_launcher():
    """ClointFusion CLI for Colab Launcher"""
    global python_exe_path
    try: 
        if os_name == windows_os:  
            print("Launching Google Colabs, actively maintained by Jay Trivedi, Research Intern@ClointFusion : https://www.linkedin.com/in/jay-trivedi-09aa791a4/ \n")
            cmd = f'{python_exe_path} "{site_packages_path}\ClointFusion\Colab_Launcher.py"' 
            os.system(cmd)
        else:
            print("Colab Launcher option is only available on Windows. We regret the inconvenience.")
            print(f"Please contribute to make this feature available on {os_name.upper()} system.")
            print(random.choice(contribution_messages))
            selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
            

    except Exception as ex:
        # selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in cli_colab_launcher " + str(ex))

@click.command(context_settings=CONTEXT_SETTINGS)
@click.option('--profile', '-p', multiple=True,help="ClointFusion Command Line Interface's basic command")
def cli_dost(profile):
    """ClointFusion CLI for DOST GUI Launcher"""
    global python_exe_path
    try:
        print("Launching ClointFusion's Drag/Drop based BOT Builder. Thanks to contribution by Murali, Research Intern@ClointFusion : https://www.linkedin.com/in/murali-manohar-varma-220a03207 \n")
        
        if os_name == windows_os:
            # if profile:
            #     cursr.execute("UPDATE CF_IMP_VALUES set DOST_PROFILE = ? where ID = 1", (profile))
            #     connct.commit()
            #     cmd = f'{python_exe_path} "{site_packages_path}\ClointFusion\DOST_CLIENT.pyw" "{profile}"'
            #     os.system(cmd)
            # else:
            #     data = cursr.execute("SELECT dost_profile from CF_IMP_VALUES")
            #     for row in data:
            #         profile =  row[0]
            #     cmd = f'{python_exe_path} "{site_packages_path}\ClointFusion\DOST_CLIENT.pyw" "{profile}"'
            #     os.system(cmd)
            import webbrowser
            webbrowser.open('https://dost.clointfusion.com/')

        elif os_name == linux_os:
            # Commands for linux
            cmd = f'sudo python{python_version} "{site_packages_path}/ClointFusion/DOST_CLIENT.pyw"'
            os.system(cmd)
        else:
            print("DOST option not available on macOS. We regret the inconvenience.")
            print(f"Please contribute to make this feature available on {os_name.upper()} system.")
            print(random.choice(contribution_messages))
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in cli_dost "+str(ex))

@click.command(context_settings=CONTEXT_SETTINGS)
def cli_bol():
    """ClointFusion CLI for DOST GUI Launcher"""
    global python_exe_path
    try:
        print("Launching ClointFusion powered Virtual Assistant : Bol\n")
        
        if os_name == windows_os:
            cmd = f'{python_exe_path} "{site_packages_path}\ClointFusion\Bol.pyw"'
            os.system(cmd)
        else:
            # # Commands for linux
            # cmd = f'sudo python{python_version} "{site_packages_path}/ClointFusion/Bol.pyw"'
            # os.system(cmd)
            print("BOL option is only available on Windows. We regret the inconvenience.")
            print(f"Please contribute to make this feature available on {os_name.upper()} system.")
            print(random.choice(contribution_messages))
            
        
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in cli_bol "+str(ex))

@click.command(context_settings=CONTEXT_SETTINGS)
def cli_whm():
    """ClointFusion CLI for WHM Launcher"""
    try:
        if os_name == windows_os:
            cmd = f'{pythonw_exe_path} "{site_packages_path}\ClointFusion\BRE_WHM.pyw"'
            os.system(cmd)
        else:
            print("Tray option is only available on Windows. We regret the inconvenience.")
            print(f"Please contribute to make this feature available on {os_name.upper()} system.")
            print(random.choice(contribution_messages))
            
            # # Commands for linux
            # cmd = f'chmod +x "{site_packages_path}\ClointFusion\BRE_WHM.pyw"'
            # os.system(cmd)
            # cmd = f'nohup sudo python3 "{site_packages_path}\ClointFusion\BRE_WHM.pyw" &'
            # os.system(cmd)
            
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in cli_whm "+str(ex))

@click.command(context_settings=CONTEXT_SETTINGS)
def cli_vlookup():
    """ClointFusion CLI for Excel VLookUp"""
    try:
        excel_vlook_up()
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in cli_vlookup="+str(ex))

@click.command(context_settings=CONTEXT_SETTINGS)
@click.option("--excel_path","-e", prompt="Please provide path of excel file having 3 columns as header: Mobile Number, Name, Message", help="Please provide path of excel file having 3 columns as header: Mobile Number, Name, Message")
def cli_send_whatsapp_msg(excel_path):
    """Sends WhatsApp Message using CF's Helium"""
    try:
        if os_name == windows_os:
            
            if os.path.isfile(Path(excel_path)):  # Check if file exists 
                cmd = f'{python_exe_path} "{site_packages_path}\ClointFusion\WA_BOT.pyw" "{excel_path}"'
                os.system(cmd)
            else:
                print("File not found. You can simply drag the file into the terminal.")
                text_to_speech("File not found. You can simply drag the file into the terminal.", show=False)
                print("or Please provide valid Absolute Path of excel file having 3 columns as header: Mobile Number, Name, Message")
                text_to_speech("or Please provide valid Absolute Path of excel file having 3 columns as header: Mobile Number, Name, Message", show=False)            
        else:
            print("WhatsApp bot is only available on Windows. We regret the inconvenience.")
            print(f"Please contribute to make this feature available on {os_name.upper()} system.")
            print(random.choice(contribution_messages))
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in send_whatsapp_msg", str(ex))

@click.command(context_settings=CONTEXT_SETTINGS)
def cli_bre_whm():
    """ClointFusion CLI for BRE and WHM"""
    if os_name == windows_os:
        try:
            from datetime import date
            from rich.table import Table
            from pivottablejs import pivot_ui
            
            style = "bold white on blue"
            console.print("Your Work Hour Report for TODAY ({}):".format(datetime.datetime.now().strftime('%dth %B,%Y %I:%M:%S %p %A')),style=style,justify='center')

            try:
                if os_name == windows_os:
                    import ctypes
                    lib = ctypes.windll.kernel32
                    t = lib.GetTickCount64()
                    t = int(str(t)[:-3])
                    mins, sec = divmod(t, 60)
                    hour, mins = divmod(mins, 60)
                    days, hour = divmod(hour, 24)
                    console.print('_' * 20,justify='center')  # Underscore
                    console.print(f"System Uptime: {days} days, {hour:02} Hours, {mins:02} Minutes, {sec:02} Seconds",justify='center')            
                
            except Exception as ex:
                selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
                print("Error in click cli_bre_whm = " + str(ex))

            
            event_table = """ CREATE TABLE IF NOT EXISTS CFEVENTS (
                TIME_STAMP TEXT NOT NULL,
                Event_Name TEXT NULL,
                X TEXT NULL,
                Y TEXT NULL,
                KEY TEXT NULL,
                Button_Name TEXT NULL,
                Click_Count TEXT NULL,
                Window_Name TEXT NULL,
                Mouse_RGB TEXT NULL,
                SNIP_File_Path TEXT NULL
                ); """

            cursr.execute(event_table)
            connct.commit()

            cursr.execute('Update CFEVENTS set Window_Name="Desktop" where Window_Name = ""')
            connct.commit()

            #Print this week's report
            week_day=datetime.datetime.now().isocalendar()[2]
            start_date=datetime.datetime.now() - datetime.timedelta(days=week_day)
            
            dates = []

            dates=[str((start_date + datetime.timedelta(days=i)).date().strftime("%dth %b, %A")) for i in range(7)]
            dates_df=[str((start_date + datetime.timedelta(days=i)).date().strftime("%Y-%m-%d")) for i in range(7)]

            df_mc_lst = []
            df_kp_lst = []
            min_max_sum_lst = []
            value_cnt = []

            for dt in dates_df:
                df = pd.read_sql('Select COUNT(Event_Name) as CNT from CFEVENTS where DATE(datetime(TIME_STAMP)) = date("{}") AND Event_Name = "Mouse Click"'.format(dt),connct)
                df_mc_lst.append(df['CNT'].values[0])

            for dt in dates_df:
                df = pd.read_sql('Select COUNT(Event_Name) as CNT from CFEVENTS where DATE(datetime(TIME_STAMP)) = date("{}") AND Event_Name = "Key Press"'.format(dt),connct)
                df_kp_lst.append(df['CNT'].values[0])

            for dt in dates_df:
                df = pd.read_sql('Select MIN(TIME_STAMP) as log_in, MAX(TIME_STAMP) as log_out from CFEVENTS where DATE(datetime(TIME_STAMP)) = date("{}")'.format(dt),connct)
                try:
                    log_in_tm = datetime.datetime.strptime(str(df['log_in'].values[0]),"%Y-%m-%d %H:%M:%S")
                    log_in_tm = datetime.datetime.strftime(log_in_tm,"%H:%M:%S %p")
                    
                    log_out_tm = datetime.datetime.strptime(str(df['log_out'].values[0]),"%Y-%m-%d %H:%M:%S")
                    log_out_tm = datetime.datetime.strftime(log_out_tm,"%H:%M:%S %p")
                    min_max_sum_lst.append(str(log_in_tm) + "," + str(log_out_tm))
                except:
                    min_max_sum_lst.append("No Data, No Data")

            for dt in dates_df:
                df = pd.read_sql('Select Window_Name from CFEVENTS where DATE(datetime(TIME_STAMP)) = date("{}")'.format(dt),connct)       
                            
                if len(df.Window_Name.value_counts()) > 0:
                    value_cnt.append(str(df['Window_Name'].value_counts(ascending=False,dropna=True,normalize=True).mul(100).round(1).astype(str) + '%'))
                else:
                    value_cnt.append("No Data")
                
            console.print('―' * 20,justify='center')  # U+2015, Horizontal Bar
            table = Table(title="This Week's Work Report",show_lines=True)

            table.add_column("Date", justify="center", style="bold cyan", width=8)
            table.add_column("Highlights",justify="center",style="bold magenta",width=20)
            table.add_column("Details",justify="left",style="bold yellow")
            
            for i in reversed(range(7)):
                table.add_row(dates[i],"In=" + str(min_max_sum_lst[i]).split(",")[0] + "\nClicks="+str(df_mc_lst[i]) + "\nKey Stokes=" + str(df_kp_lst[i]) + "\nOut=" + str(min_max_sum_lst[i]).split(",")[1],value_cnt[i])

            console.print(table,justify='center')

            # Charts       
            df = pd.read_sql("SELECT DISTINCT(DATE(TIME_STAMP)),TIME_STAMP, Event_Name,Window_Name as 'Software/Program' from CFEVENTS group by Window_Name order by TIME_STAMP ASC",connct)

            delay_lst = []
            for ind in df.index:
                try:
                    current_line_time_stamp = df['TIME_STAMP'][ind]
                    next_line_time_stamp = df['TIME_STAMP'][ind + 1]

                    delay = parser.parse(next_line_time_stamp) - parser.parse(current_line_time_stamp)
                    delay = int(delay.total_seconds())
                
                    delay = str(datetime.timedelta(seconds=delay))
                    delay_lst.append(delay)
                except Exception:
                    # delay = parser.parse(str(datetime.datetime.now().strftime("%H:%M:%S"))) - parser.parse(current_line_time_stamp)
                    delay_lst.append("No Data")

            df["Time Spent"] = delay_lst
            df.drop('TIME_STAMP', axis=1, inplace=True)
            df=df[~df['Time Spent'].isin (["0:00:00"])] # str(datetime.datetime.strptime('01:00', '%H:%M'))]

            print()
            yes_no = console.input("Would you like to see [bold red]detailed report (Y/N) [/] ?")
            if yes_no in ["Yes", "y", "Y","yes"]:
                    pivot_ui(df,rows=["(DATE(TIME_STAMP))"], cols=['Time Spent','Software/Program','Event_Name'])
                    webbrowser.open_new_tab('pivottablejs.html')

            console.print('―' * 20,justify='center')  # U+2015, Horizontal Bar
            table = Table(title="Last Week's Work Report",show_lines=True)
            console.print("All data is being stored locally on your own Computer and is in your Control.",style=style,justify='center',crop=False)
            print()
        except Exception as ex:
            selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
            print("You may need to restart your computer !")
            print("Error in cli_bre_whm="+str(ex))
    else:
        print("Work Hour Monitor currently only available on Windows PC.")
        print(f"Please contribute to make this feature available on {os_name.upper()} system.")
        print(random.choice(contribution_messages))

@click.command(context_settings=CONTEXT_SETTINGS)
@click.option('--message', '-m', multiple=True,help="ClointFusion Command Line Interface's basic command")
def cli_cf(message):
    """ClointFusion Command Line Interface's basic command"""
    click.echo('\n'.join(message))
    click.echo("Below commands are available for TERMINAL use :   ( W - Windows, U - Ubuntu )\n\n1)   dost         - [W,  ] || Build RPA Bots without Code.\n2)   bol          - [W, U] || Voice based assistant powered by ClointFusion.\n3)   cf_work      - [W,  ] || Get your computer usage report.\n4)   cf_tray      - [W,  ] || Launch ClointFusion tray icon.\n5)   cf_st        - [W, U] || Check your internet speed.\n6)   cf_wm        - [W,  ] || Sends WhatsApp messages by providing path of excel file having 3 columns: Mobile Number, Name, Message.\n7)   cf_py        - [W, U] || Opens a Python interpreter with 'import ClointFusion as cf' pre-loaded.\n8)   cf_like      - [W,  ] || CLI for auto liking CF's specific posts on Social Media.\n9)   cf_tour      - [W, U] || CLI for guided tour of ClointFusion.\n10)  cf_vlookup   - [W, U] || Performs excel_vlook_up on the given excel files for the desired columns.\n11)  cf_sm        - [W, U] || Opens all our ClointFusion's Social Media in Default Browser.")

@click.command(context_settings=CONTEXT_SETTINGS)
def cli_call_sm():
    """Opens all our Social Media in Google Chrome"""
    try:
        call_social_media()
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in cli_call_sm="+str(ex))
        

@click.command(context_settings=CONTEXT_SETTINGS)
def cli_cf_py():
    """Open python interpreter with preloaded clointfusion as cf."""
    
    functions_list = ['message_counter_down_timer', 'message_pop_up', 'message_flash', 'message_toast', 'mouse_click', 'mouse_move', 'mouse_drag_from_to', 'mouse_search_snip_return_coordinates_x_y', 'key_press', 'key_write_enter', 'key_hit_enter', 'browser_activate', 'folder_read_text_file', 'folder_write_text_file', 'folder_create', 'folder_create_text_file', 'folder_get_all_filenames_as_list', 'folder_delete_all_files', 'file_rename', 'window_show_desktop', 'window_get_all_opened_titles_windows', 'window_activate_and_maximize_windows', 'window_minimize_windows', 'window_close_windows', 'launch_any_exe_bat_application', 'string_extract_only_alphabets', 'string_extract_only_numbers', 'string_remove_special_characters', 'scrape_save_contents_to_notepad', 'scrape_get_contents_by_search_copy_paste', 'pause_program', 'show_emoji', 'clear_screen', 'print_with_magic_color', 'text_to_speech', 'speech_to_text']
    ch_function_1 = random.choice(functions_list)
    ch_function_2 = random.choice(functions_list)
    ch_function_3 = random.choice(functions_list)
    
    if os_name == windows_os:
        os.system(f'{python_exe_path} -i -c "import ClointFusion as cf; print(\'Try some of our functions | cf.{ch_function_1}() | or | cf.{ch_function_2}() | or | cf.{ch_function_3}() |\')"')
    elif os_name == linux_os:
        os.system(f'sudo python{python_version} -i -c "import ClointFusion as cf; print(\'Try some of our functions | cf.{ch_function_1}() | or | cf.{ch_function_2}() | or | cf.{ch_function_3}() |\')"')
    else:
        try:
            os.system(f'python{python_version} -i -c "import ClointFusion as cf; print(\'Try some of our functions | cf.{ch_function_1}() | or | cf.{ch_function_2}() | or | cf.{ch_function_3}() |\')"')
        except:
            print("This command is not available on macOS.")
            print(f"Please contribute to make this feature available on {os_name.upper()} system.")
            print(random.choice(contribution_messages))
            selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
            

@click.command(context_settings=CONTEXT_SETTINGS)
def cli_cf_tour():
    """CLI for the guided tour."""
    if os_name != mac_os:
        _perform_self_test(True)
    else:
        try:
            _perform_self_test(True)
        except:
            print("This command is not available on macOS.")
            print(f"Please contribute to make this feature available on {os_name.upper()} system.")
            print(random.choice(contribution_messages))
            selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
            
   
# --------- CLI Commands Ends ---------


# --------- TEST FOR CLI ---------

def cli_bre_whm_test():
    """ClointFusion CLI for BRE and WHM"""
    try:
        from rich.table import Table
        
        style = "bold white on blue"
        console.print("Your Work Hour Report for TODAY ({}):".format(datetime.datetime.now().strftime('%dth %B,%Y %I:%M:%S %p %A')),style=style,justify='center')

        try:
            if os_name == windows_os:
                import ctypes
                lib = ctypes.windll.kernel32
                t = lib.GetTickCount64()
                t = int(str(t)[:-3])
                mins, sec = divmod(t, 60)
                hour, mins = divmod(mins, 60)
                days, hour = divmod(hour, 24)
                console.print('_' * 20,justify='center')  # Underscore
                console.print(f"System Uptime: {days} days, {hour:02} Hours, {mins:02} Minutes, {sec:02} Seconds",justify='center')

        except Exception as ex:
            selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
            print("Error in click cli_bre_whm = " + str(ex))

        
        event_table = """ CREATE TABLE IF NOT EXISTS CFEVENTS (
            TIME_STAMP TEXT NOT NULL,
            Event_Name TEXT NULL,
            X TEXT NULL,
            Y TEXT NULL,
            KEY TEXT NULL,
            Button_Name TEXT NULL,
            Click_Count TEXT NULL,
            Window_Name TEXT NULL,
            Mouse_RGB TEXT NULL,
            SNIP_File_Path TEXT NULL
            ); """

        cursr.execute(event_table)
        connct.commit()
        
        # cursr = connct.cursor()
        cursr.execute('Update CFEVENTS set Window_Name="Desktop" where Window_Name = ""')
        connct.commit()

        #Print this week's report
        week_day=datetime.datetime.now().isocalendar()[2]
        start_date=datetime.datetime.now() - datetime.timedelta(days=week_day)
        
        dates = []

        dates=[str((start_date + datetime.timedelta(days=i)).date().strftime("%dth %b, %A")) for i in range(7)]
        dates_df=[str((start_date + datetime.timedelta(days=i)).date().strftime("%Y-%m-%d")) for i in range(7)]

        df_mc_lst = []
        df_kp_lst = []
        min_max_sum_lst = []
        value_cnt = []

        for dt in dates_df:
            df = pd.read_sql('Select COUNT(Event_Name) as CNT from CFEVENTS where DATE(datetime(TIME_STAMP)) = date("{}") AND Event_Name = "Mouse Click"'.format(dt),connct)
            df_mc_lst.append(df['CNT'].values[0])

        for dt in dates_df:
            df = pd.read_sql('Select COUNT(Event_Name) as CNT from CFEVENTS where DATE(datetime(TIME_STAMP)) = date("{}") AND Event_Name = "Key Press"'.format(dt),connct)
            df_kp_lst.append(df['CNT'].values[0])

        for dt in dates_df:
            df = pd.read_sql('Select MIN(TIME_STAMP) as log_in, MAX(TIME_STAMP) as log_out from CFEVENTS where DATE(datetime(TIME_STAMP)) = date("{}")'.format(dt),connct)
            try:
                log_in_tm = datetime.datetime.strptime(str(df['log_in'].values[0]),"%Y-%m-%d %H:%M:%S")
                log_in_tm = datetime.datetime.strftime(log_in_tm,"%H:%M:%S %p")
                
                log_out_tm = datetime.datetime.strptime(str(df['log_out'].values[0]),"%Y-%m-%d %H:%M:%S")
                log_out_tm = datetime.datetime.strftime(log_out_tm,"%H:%M:%S %p")
                min_max_sum_lst.append(str(log_in_tm) + "," + str(log_out_tm))
            except:
                min_max_sum_lst.append("No Data, No Data")

        for dt in dates_df:
            df = pd.read_sql('Select Window_Name from CFEVENTS where DATE(datetime(TIME_STAMP)) = date("{}")'.format(dt),connct)       
                        
            if len(df.Window_Name.value_counts()) > 0:
                value_cnt.append(str(df['Window_Name'].value_counts(ascending=False,dropna=True,normalize=True).mul(100).round(1).astype(str) + '%'))
            else:
                value_cnt.append("No Data")
            
        console.print('―' * 20,justify='center')  # U+2015, Horizontal Bar
        table = Table(title="This Week's Work Report",show_lines=True)

        table.add_column("Date", justify="center", style="bold cyan", width=8)
        table.add_column("Highlights",justify="center",style="bold magenta",width=20)
        table.add_column("Details",justify="left",style="bold yellow")
        
        for i in reversed(range(7)):
            table.add_row(dates[i],"In=" + str(min_max_sum_lst[i]).split(",")[0] + "\nClicks="+str(df_mc_lst[i]) + "\nKey Stokes=" + str(df_kp_lst[i]) + "\nOut=" + str(min_max_sum_lst[i]).split(",")[1],value_cnt[i])

        console.print(table,justify='center')

        # Charts       
        df = pd.read_sql("SELECT DISTINCT(DATE(TIME_STAMP)),TIME_STAMP, Event_Name,Window_Name as 'Software/Program' from CFEVENTS group by Window_Name order by TIME_STAMP ASC",connct)

        delay_lst = []
        for ind in df.index:
            try:
                current_line_time_stamp = df['TIME_STAMP'][ind]
                next_line_time_stamp = df['TIME_STAMP'][ind + 1]

                delay = parser.parse(next_line_time_stamp) - parser.parse(current_line_time_stamp)
                delay = int(delay.total_seconds())
            
                delay = str(datetime.timedelta(seconds=delay))
                delay_lst.append(delay)
            except Exception:
                # delay = parser.parse(str(datetime.datetime.now().strftime("%H:%M:%S"))) - parser.parse(current_line_time_stamp)
                delay_lst.append("No Data")

        df["Time Spent"] = delay_lst
        df.drop('TIME_STAMP', axis=1, inplace=True)
        df=df[~df['Time Spent'].isin (["0:00:00"])] # str(datetime.datetime.strptime('01:00', '%H:%M'))]

        print()

        console.print('―' * 20,justify='center')  # U+2015, Horizontal Bar
        table = Table(title="Last Week's Work Report",show_lines=True)
        console.print("All data is being stored locally on your own Computer and is in your Control.",style=style,justify='center',crop=False)
        print()

    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        print("You may need to restart your computer !")
        print("Error in cli_bre_whm="+str(ex))

def cli_speed_test_test():
    """CLI for testing internet bandwidth using speedtest.net"""
    try:
        os.system("speedtest-cli")
    except Exception as ex:
        selft.crash_report(traceback.format_exception(*sys.exc_info(),limit=None, chain=True))
        text_to_speech("Error in cli_speed_test_test="+str(ex))

def cli_cf_test():
    print("Below commands are available for TERMINAL use :   ( W - Windows, U - Ubuntu )\n\n1)   dost         - [W,  ] || Build RPA Bots without Code.\n2)   bol          - [W, U] || Voice based assistant powered by ClointFusion.\n3)   cf_work      - [W,  ] || Get your computer usage report.\n4)   cf_tray      - [W,  ] || Launch ClointFusion tray icon.\n5)   cf_st        - [W, U] || Check your internet speed.\n6)   cf_wm        - [W,  ] || Sends WhatsApp messages by providing path of excel file having 3 columns: Mobile Number, Name, Message.\n7)   cf_py        - [W, U] || Opens a Python interpreter with 'import ClointFusion as cf' pre-loaded.\n8)   cf_like      - [W,  ] || CLI for auto liking CF's specific posts on Social Media.\n9)   cf_tour      - [W, U] || CLI for guided tour of ClointFusion.\n10)  cf_vlookup   - [W, U] || Performs excel_vlook_up on the given excel files for the desired columns.\n11)  cf_sm        - [W, U] || Opens all our ClointFusion's Social Media in Default Browser.")

# --------- TEST FOR CLI Ends ---------

# --------- 4. All default services ---------

# All new functions to be added before this line
# ########################
# ClointFusion's DEFAULT SERVICES
if os_name != mac_os:
    if os_name == windows_os:
        db_file_path = r'{}\BRE_WHM.db'.format(str(config_folder_path))
    elif os_name == linux_os:
        db_file_path = folder_create_text_file(config_folder_path, 'BRE_WHM.db', custom=True)
    else:
        pass

    try:
        connct = sqlite3.connect(db_file_path,check_same_thread=False)
        cursr = connct.cursor()
    except Exception as ex:
        print("Error in connecting to DB="+str(ex))


    data = cursr.execute("SELECT updating from CF_IMP_VALUES")

    for row in data:
        updating =  row[0]
        if updating == "False":

            _welcome_to_clointfusion()

            _init_log_file()
            update_log_excel_file(bot_name +'- BOT initiated')
            _ask_user_semi_automatic_mode()
            enable_semi_automatic_mode = False # By DEFAULT
        else:
            sys.exit()
    

    # ########################

    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", category=PendingDeprecationWarning)
        warnings.filterwarnings("ignore", category=DeprecationWarning)
else:
    _welcome_to_clointfusion()
    try:
        print_with_magic_color("We do not yet support macOS completely. Feel free to have a look and contribute to make us strong.")
        print_with_magic_color(random.choice(contribution_messages))
    except:
        print("We do not yet support macOS completely. Feel free to have a look and contribute to make us strong.")
        print(random.choice(contribution_messages))
